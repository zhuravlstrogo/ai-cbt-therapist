# -*- coding: utf-8 -*-
"""
Weekly check-in module for tracking therapy progress
Provides scheduled weekly assessments and manual progress evaluation
"""

import os
import json
import random
import hashlib
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
from telebot import types
from openpyxl import load_workbook, Workbook
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger

# Import LLM client for summaries
from openrouter import OpenRouterClient
from config import MODEL_SIMPLE, TEMPERATURE, TOP_P, TOP_K

# Path to Excel file for saving check-in results
CHECKIN_FILE = 'check_in.xlsx'

# Store user check-in states
user_checkin_states: Dict[int, Dict[str, Any]] = {}

# Scheduler instance
scheduler: Optional[AsyncIOScheduler] = None

# Cache for LLM responses
llm_cache: Dict[str, Dict[str, Any]] = {}
CACHE_TTL_HOURS = 24

# Greeting variations
GREETING_VARIATIONS = [
    "Привет! Как у тебя дела?",
    "Привет! Как твои дела на этой неделе?",
    "Привет! Как прошла твоя неделя?",
    "Привет! Как ты себя чувствуешь сегодня?",
    "Привет! Рады тебя видеть! Как дела?",
    "Привет! Как твоё настроение сегодня?",
    "Привет! Как ваши дела?"
]


def get_cache_key(user_id: int, data_type: str, data_hash: str) -> str:
    """Generate cache key for LLM responses"""
    return f"{user_id}_{data_type}_{data_hash}"


def get_cached_response(cache_key: str) -> Optional[str]:
    """Get cached LLM response if still valid"""
    if cache_key in llm_cache:
        cached_data = llm_cache[cache_key]
        if datetime.now() - cached_data['timestamp'] < timedelta(hours=CACHE_TTL_HOURS):
            return cached_data['response']
    return None


def set_cached_response(cache_key: str, response: str) -> None:
    """Store LLM response in cache"""
    llm_cache[cache_key] = {
        'response': response,
        'timestamp': datetime.now()
    }


def ensure_checkin_file_exists():
    """Create check-in Excel file if it doesn't exist"""
    if not os.path.exists(CHECKIN_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'CheckIn'

        # Define headers
        headers = [
            'User ID', 'Username', 'User Name', 'Check-in Date',
            'Days Since Start', 'Question 1 Response', 'Question 2 Response',
            'Problems Ratings', 'Goal Progress', 'Weekly Summary',
            'Crisis Detected', 'Crisis Type'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        wb.save(CHECKIN_FILE)
        print(f"Created {CHECKIN_FILE}")


def get_user_start_date(user_id: int) -> Optional[datetime]:
    """Get user's therapy start date from messages.xlsx"""
    try:
        if not os.path.exists('messages.xlsx'):
            return None

        df = pd.read_excel('messages.xlsx')
        user_rows = df[df['User ID'] == user_id]

        if len(user_rows) > 0:
            # Find the earliest date for this user
            # Use 'Protocol Choice' column which contains timestamps
            dates = pd.to_datetime(user_rows['Protocol Choice'], errors='coerce')
            valid_dates = dates.dropna()
            if len(valid_dates) > 0:
                return valid_dates.min().to_pydatetime()

    except Exception as e:
        print(f"Error getting user start date: {e}")

    return None


def get_last_checkin_date(user_id: int) -> Optional[datetime]:
    """Get user's last check-in date"""
    try:
        ensure_checkin_file_exists()
        df = pd.read_excel(CHECKIN_FILE)

        user_rows = df[df['User ID'] == user_id]
        if len(user_rows) > 0:
            dates = pd.to_datetime(user_rows['Check-in Date'], errors='coerce')
            valid_dates = dates.dropna()
            if len(valid_dates) > 0:
                return valid_dates.max().to_pydatetime()

    except Exception as e:
        print(f"Error getting last check-in date: {e}")

    return None


def should_do_checkin(user_id: int) -> Tuple[bool, str]:
    """
    Check if user should do check-in
    Returns: (should_do, reason)
    """
    start_date = get_user_start_date(user_id)
    if not start_date:
        return False, "Нет данных о начале терапии"

    last_checkin = get_last_checkin_date(user_id)

    # If no previous check-in and at least 7 days since start
    if not last_checkin:
        days_since_start = (datetime.now() - start_date).days
        if days_since_start >= 7:
            return True, f"Прошло {days_since_start} дней с начала терапии"
        else:
            return False, f"До первого check-in осталось {7 - days_since_start} дней"

    # Check if week passed since last check-in
    days_since_last = (datetime.now() - last_checkin).days
    if days_since_last >= 7:
        return True, f"Прошло {days_since_last} дней с последнего check-in"

    # Allow manual check-in if there's new data (simplified check)
    return True, "Доступен новый check-in"


def get_user_insights_last_week(user_id: int) -> List[str]:
    """Get user's insights from exercises in the last 7 days"""
    try:
        if not os.path.exists('exercises.xlsx'):
            return []

        df = pd.read_excel('exercises.xlsx')

        # Filter by user and last 7 days
        user_exercises = df[df['User ID'] == user_id]
        if 'Date Time' in user_exercises.columns:
            dates = pd.to_datetime(user_exercises['Date Time'], errors='coerce')
            week_ago = datetime.now() - timedelta(days=7)
            recent = user_exercises[dates > week_ago]

            # Extract insights from "Final Answers" column
            insights = []
            if 'Final Answers' in recent.columns:
                for answers_json in recent['Final Answers'].dropna():
                    try:
                        answers = json.loads(answers_json) if isinstance(answers_json, str) else answers_json
                        if isinstance(answers, dict) and 'insight' in answers:
                            insights.append(answers['insight'])
                    except:
                        continue

            return insights

    except Exception as e:
        print(f"Error getting user insights: {e}")

    return []


def get_problem_dynamics(user_id: int) -> Dict[str, List[int]]:
    """Get dynamics of problem ratings from previous check-ins"""
    try:
        ensure_checkin_file_exists()
        df = pd.read_excel(CHECKIN_FILE)

        user_rows = df[df['User ID'] == user_id].sort_values('Check-in Date')
        dynamics = {}

        for _, row in user_rows.iterrows():
            if pd.notna(row['Problems Ratings']):
                try:
                    ratings = json.loads(row['Problems Ratings'])
                    for problem, rating in ratings.items():
                        if problem not in dynamics:
                            dynamics[problem] = []
                        dynamics[problem].append(rating)
                except:
                    continue

        return dynamics

    except Exception as e:
        print(f"Error getting problem dynamics: {e}")

    return {}


async def start_check_in(bot, chat_id: int, user_id: int, username: str, scheduled: bool = False):
    """Start the check-in process"""
    try:
        # Get user info
        from greeting import user_states
        user_name = 'Друг'
        user_problems = []

        if user_id in user_states:
            user_name = user_states[user_id].get('user_name', 'Друг')
            user_problems = user_states[user_id].get('problems', [])

        # Initialize check-in state
        user_checkin_states[user_id] = {
            'step': 1,
            'username': username,
            'user_name': user_name,
            'responses': {
                'q1_response': '',
                'q2_response': '',
                'problem_ratings': {},
                'goal_progress': 0
            },
            'problems': user_problems,
            'current_problem_idx': 0,
            'start_date': datetime.now(),
            'scheduled': scheduled
        }

        # Send first question with variation
        greeting = random.choice(GREETING_VARIATIONS)

        from universal_menu import get_menu_button
        markup = get_menu_button()

        await bot.send_message(chat_id, greeting, reply_markup=markup)

        print(f"Started check-in for {username} (ID: {user_id}), scheduled={scheduled}")

    except Exception as e:
        print(f"Error starting check-in: {e}")


async def handle_checkin_text_input(bot, message):
    """Handle text input during check-in process"""
    try:
        user_id = message.from_user.id
        username = message.from_user.username or 'Unknown'
        chat_id = message.chat.id

        if user_id not in user_checkin_states:
            return

        state = user_checkin_states[user_id]
        step = state['step']

        if step == 1:
            # Save response to question 1
            state['responses']['q1_response'] = message.text
            state['step'] = 2

            # Ask question 2
            text = "Как ты сейчас себя чувствуешь?"

            from universal_menu import get_menu_button
            markup = get_menu_button()

            await bot.send_message(chat_id, text, reply_markup=markup)

        elif step == 2:
            # Save response to question 2
            state['responses']['q2_response'] = message.text
            state['step'] = 3
            state['current_problem_idx'] = 0

            # Move to problem ratings if user has problems
            if state['problems']:
                await show_problem_rating(bot, chat_id, user_id)
            else:
                # Skip to goal progress
                state['step'] = 4
                await show_goal_progress(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling check-in text input: {e}")


async def show_problem_rating(bot, chat_id: int, user_id: int):
    """Show current problem for rating (0-3 scale)"""
    try:
        if user_id not in user_checkin_states:
            return

        state = user_checkin_states[user_id]

        if state['current_problem_idx'] >= len(state['problems']):
            # All problems rated, move to goal progress
            state['step'] = 4
            await show_goal_progress(bot, chat_id, user_id)
            return

        current_problem = state['problems'][state['current_problem_idx']]

        # Create rating buttons
        markup = types.InlineKeyboardMarkup()

        # Rating buttons (0-3)
        btn_0 = types.InlineKeyboardButton("0️⃣ не мешает", callback_data=f"checkin_rate:{state['current_problem_idx']}:0")
        btn_1 = types.InlineKeyboardButton("1️⃣ немного", callback_data=f"checkin_rate:{state['current_problem_idx']}:1")
        btn_2 = types.InlineKeyboardButton("2️⃣ заметно", callback_data=f"checkin_rate:{state['current_problem_idx']}:2")
        btn_3 = types.InlineKeyboardButton("3️⃣ сильно", callback_data=f"checkin_rate:{state['current_problem_idx']}:3")

        markup.row(btn_0, btn_1)
        markup.row(btn_2, btn_3)

        # Menu button
        btn_menu = types.InlineKeyboardButton("↩️ Главное меню", callback_data="menu:show")
        markup.add(btn_menu)

        text = (
            f"Оцени, насколько сейчас тебя беспокоит:\n\n"
            f"**{current_problem}**\n\n"
            f"0 - не мешает ️ 1 - немного ️ 2 - заметно ️ 3 - сильно мешает"
        )

        await bot.send_message(chat_id, text, reply_markup=markup, parse_mode='Markdown')

    except Exception as e:
        print(f"Error showing problem rating: {e}")


async def handle_problem_rating(bot, callback_query, problem_idx: int, rating: int):
    """Handle problem rating selection"""
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        # Answer callback immediately
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_checkin_states:
            return

        state = user_checkin_states[user_id]
        problem_idx = int(problem_idx)
        rating_value = int(rating)

        # Save rating
        problem = state['problems'][problem_idx]
        state['responses']['problem_ratings'][problem] = rating_value

        # Move to next problem
        state['current_problem_idx'] += 1

        # Show next problem or move to goal progress
        await show_problem_rating(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling problem rating: {e}")


async def show_goal_progress(bot, chat_id: int, user_id: int):
    """Show goal progress question (0-10 scale)"""
    try:
        if user_id not in user_checkin_states:
            return

        state = user_checkin_states[user_id]

        # Get user's goal
        from greeting import user_states
        goal = "твоей цели"
        if user_id in user_states and 'goal' in user_states[user_id]:
            goal = f"цели: {user_states[user_id]['goal']}"

        # Create progress buttons (0-10)
        markup = types.InlineKeyboardMarkup()

        # Add buttons in rows of 3
        for i in range(0, 11, 3):
            row_buttons = []
            for j in range(3):
                if i + j <= 10:
                    btn = types.InlineKeyboardButton(
                        str(i + j),
                        callback_data=f"checkin_goal:{i + j}"
                    )
                    row_buttons.append(btn)
            if row_buttons:
                markup.row(*row_buttons)

        # Menu button
        btn_menu = types.InlineKeyboardButton("↩️ Главное меню", callback_data="menu:show")
        markup.add(btn_menu)

        text = (
            f"Насколько ты продвинулся: {goal}?\n\n"
            "Оцени от 0 до 10:\n"
            "0 - совсем не продвинулся\n"
            "10 - полностью достиг цели"
        )

        await bot.send_message(chat_id, text, reply_markup=markup)

    except Exception as e:
        print(f"Error showing goal progress: {e}")


async def handle_goal_progress(bot, callback_query, progress: int):
    """Handle goal progress selection"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        # Answer callback immediately
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_checkin_states:
            return

        state = user_checkin_states[user_id]
        state['responses']['goal_progress'] = int(progress)

        # All questions answered, generate summary
        await generate_and_show_summary(bot, chat_id, user_id, username)

    except Exception as e:
        print(f"Error handling goal progress: {e}")


async def generate_weekly_summary(user_id: int, responses: Dict, user_name: str) -> str:
    """Generate weekly summary using LLM"""
    try:
        # Get insights from exercises
        insights = get_user_insights_last_week(user_id)

        # Get problem dynamics
        dynamics = get_problem_dynamics(user_id)

        # Create cache key
        cache_data = {
            'responses': responses,
            'insights': insights,
            'dynamics': dynamics
        }
        data_str = json.dumps(cache_data, ensure_ascii=False, sort_keys=True)
        data_hash = hashlib.md5(data_str.encode()).hexdigest()[:8]
        cache_key = get_cache_key(user_id, 'weekly_summary', data_hash)

        # Check cache
        cached = get_cached_response(cache_key)
        if cached:
            print(f"Using cached weekly summary for user {user_id}")
            return cached

        # Prepare data for analysis
        insights_text = ""
        if insights:
            insights_text = "Инсайты из упражнений за неделю:\n" + "\n".join(f"- {i}" for i in insights)

        dynamics_text = ""
        if dynamics:
            dynamics_text = "Динамика проблем:\n"
            for problem, ratings in dynamics.items():
                if len(ratings) > 1:
                    change = ratings[-1] - ratings[-2]
                    direction = "⬆️" if change > 0 else "⬇️" if change < 0 else "➡️"
                    dynamics_text += f"- {problem}: {ratings[-2]} {direction} {ratings[-1]}\n"

        system_prompt = """Ты опытный психотерапевт с 15-летним стажем.
Проанализируй недельный прогресс клиента и создай поддерживающее саммари.

Ключи в анализ:
1. Цитируй инсайты из упражнений (если есть)
2. Поддерживай динамику (что улучшилось или стабилизировалось)
3. Мягкие наблюдения по паттернам (без критики)
4. Мотивирующее предложение следующего шага

Тон: теплый, поддерживающий, профессиональный.
Объем: 3-5 предложений."""

        user_prompt = f"""Имя клиента: {user_name}

Ответы на вопросы check-in:
- Как дела: {responses['q1_response']}
- Как себя чувствует: {responses["q2_response"]}
- Прогресс по цели: {responses["goal_progress"]}/10

{insights_text}

{dynamics_text}

Текущие оценки проблем:
{json.dumps(responses['problem_ratings'], ensure_ascii=False, indent=2)}

Создай поддерживающее саммари на неделю."""

        client = OpenRouterClient()
        response, usage = client.get_simple_response(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            model=MODEL_SIMPLE,
            temperature=TEMPERATURE,
            top_p=TOP_P,
            top_k=TOP_K
        )

        # Cache the response
        set_cached_response(cache_key, response.strip())
        print(f"Cached weekly summary for user {user_id}")

        return response.strip()

    except Exception as e:
        print(f"Error generating weekly summary: {e}")
        return "Продолжай двигаться 2 своём темпе. Каждый шаг важен."


async def check_crisis_indicators(responses: Dict) -> Tuple[bool, Optional[str]]:
    """Check for crisis indicators in user responses"""
    try:
        # Import safety check module
        from safety_check import check_text_safety

        # Combine all text responses for analysis
        all_text = f"{responses['q1_response']} {responses['q2_response']}"

        # Use unified safety check
        crisis_detected, crisis_type, confidence = await check_text_safety(
            text=all_text,
            context="checkin"
        )

        return crisis_detected, crisis_type

    except Exception as e:
        print(f"Error checking crisis indicators: {e}")
        return False, None


async def show_crisis_support(bot, chat_id: int, user_name: str, crisis_type: str):
    """Show crisis support message and resources"""
    try:
        # Use unified crisis support from safety_check module
        from safety_check import show_crisis_support as show_unified_support

        await show_unified_support(
            bot=bot,
            chat_id=chat_id,
            user_name=user_name,
            crisis_type=crisis_type,
            context="checkin",
            continue_after=False
        )

    except Exception as e:
        print(f"Error showing crisis support: {e}")


async def generate_and_show_summary(bot, chat_id: int, user_id: int, username: str):
    """Generate and show check-in summary"""
    try:
        if user_id not in user_checkin_states:
            return

        state = user_checkin_states[user_id]
        responses = state['responses']
        user_name = state['user_name']

        # Save check-in results first
        save_check_in_results(user_id, username, user_name, responses)

        # Send loading message
        loading_text = "Анализирую твой прогресс... ⏳"
        await bot.send_message(chat_id, loading_text)

        # Check for crisis indicators
        crisis_detected, crisis_type = await check_crisis_indicators(responses)

        if crisis_detected and crisis_type:
            # Show crisis support
            await show_crisis_support(bot, chat_id, user_name, crisis_type)

            # Log crisis detection
            from safety_check import log_crisis_detection
            await log_crisis_detection(
                user_id=user_id,
                username=username,
                crisis_type=crisis_type,
                context="checkin",
                text_sample=f"{responses['q1_response'][:100]}...",
                file_path=CHECKIN_FILE
            )

            # Update Excel with crisis flag
            update_crisis_flag(user_id, crisis_type)
        else:
            # Generate and show weekly summary
            summary = await generate_weekly_summary(user_id, responses, user_name)

            # Format summary message
            text = (
                f"📋 **Твой недельный прогресс**\n\n"
                f"{summary}\n\n"
                f"Прогресс 📍 цели: {responses['goal_progress']}/10"
            )

            # Add navigation buttons
            markup = types.InlineKeyboardMarkup()

            btn_exercises = types.InlineKeyboardButton(
                "<️ Выбрать упражнение",
                callback_data="menu:select_exercise"
            )
            btn_diary = types.InlineKeyboardButton(
                "📝 Дневник",
                callback_data="menu:diary"
            )
            btn_menu = types.InlineKeyboardButton(
                "↩️ Главное меню",
                callback_data="menu:show"
            )

            markup.add(btn_exercises)
            markup.add(btn_diary)
            markup.add(btn_menu)

            await bot.send_message(
                chat_id,
                text,
                reply_markup=markup,
                parse_mode='Markdown'
            )

        # Clean up state
        del user_checkin_states[user_id]
        print(f"Completed check-in for {username} (ID: {user_id})")

    except Exception as e:
        print(f"Error generating and showing summary: {e}")

        # Fallback message
        markup = types.InlineKeyboardMarkup()
        btn_menu = types.InlineKeyboardButton("↩️ Главное меню", callback_data="menu:show")
        markup.add(btn_menu)

        await bot.send_message(
            chat_id,
            "Не удалось создать анализ. Попробуй позже.",
            reply_markup=markup
        )

        # Clean up state
        if user_id in user_checkin_states:
            del user_checkin_states[user_id]


def save_check_in_results(user_id: int, username: str, user_name: str, responses: Dict):
    """Save check-in results to Excel"""
    try:
        ensure_checkin_file_exists()

        # Calculate days since start
        start_date = get_user_start_date(user_id)
        days_since_start = (datetime.now() - start_date).days if start_date else 0

        # Load workbook
        wb = load_workbook(CHECKIN_FILE)
        ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        # Add data
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = user_name
        ws[f'D{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'E{next_row}'] = days_since_start
        ws[f'F{next_row}'] = responses['q1_response']
        ws[f'G{next_row}'] = responses['q2_response']
        ws[f'H{next_row}'] = json.dumps(responses['problem_ratings'], ensure_ascii=False)
        ws[f'I{next_row}'] = responses['goal_progress']
        # Weekly summary will be added later
        ws[f'K{next_row}'] = False  # Crisis detected (default)
        ws[f'L{next_row}'] = ''  # Crisis type

        wb.save(CHECKIN_FILE)
        print(f"Check-in saved for {username} (ID: {user_id})")

    except Exception as e:
        print(f"Error saving check-in results: {e}")


def update_crisis_flag(user_id: int, crisis_type: str):
    """Update crisis detection flag in Excel"""
    try:
        wb = load_workbook(CHECKIN_FILE)
        ws = wb.active

        # Find the last row for this user
        for row in range(ws.max_row, 1, -1):
            if ws[f'A{row}'].value == user_id:
                ws[f'K{row}'] = True
                ws[f'L{row}'] = crisis_type
                break

        wb.save(CHECKIN_FILE)
        print(f"Updated crisis flag for user {user_id}: {crisis_type}")

    except Exception as e:
        print(f"Error updating crisis flag: {e}")


async def schedule_weekly_checkins(bot):
    """Schedule weekly check-ins for all active users"""
    global scheduler

    try:
        # Initialize scheduler if not exists
        if scheduler is None:
            scheduler = AsyncIOScheduler(timezone='Europe/Moscow')
            scheduler.start()
            print("APScheduler started")

        # Check all users daily for needed check-ins
        async def check_users_for_checkin():
            """Check all users if they need check-in"""
            try:
                # Get all unique users from messages.xlsx
                if not os.path.exists('messages.xlsx'):
                    return

                df = pd.read_excel('messages.xlsx')
                unique_users = df['User ID'].unique()

                for user_id in unique_users:
                    should_check, reason = should_do_checkin(user_id)
                    if should_check:
                        # Get user info
                        username = df[df['User ID'] == user_id]['Username'].iloc[0]

                        # Find chat_id (assuming it's same as user_id for private chats)
                        chat_id = user_id

                        print(f"Scheduling check-in for user {user_id}: {reason}")
                        await start_check_in(bot, chat_id, user_id, username, scheduled=True)

            except Exception as e:
                print(f"Error in scheduled check-in: {e}")

        # Schedule daily check at 10:00 AM
        scheduler.add_job(
            check_users_for_checkin,
            trigger=IntervalTrigger(days=1, start_date=datetime.now().replace(hour=10, minute=0, second=0)),
            id='daily_checkin_check',
            replace_existing=True
        )

        print("Weekly check-in scheduler initialized")

    except Exception as e:
        print(f"Error initializing scheduler: {e}")


async def show_check_in_progress(bot, chat_id: int, user_id: int, username: str):
    """Show check-in progress or start new check-in"""
    try:
        should_check, reason = should_do_checkin(user_id)

        if should_check:
            # Start new check-in
            await start_check_in(bot, chat_id, user_id, username, scheduled=False)
        else:
            # Show last check-in results
            last_checkin = get_last_checkin_date(user_id)

            if last_checkin:
                days_ago = (datetime.now() - last_checkin).days
                text = (
                    f"📍 Последняя оценка прогресса была {days_ago} дней назад.\n\n"
                    f"{reason}\n\n"
                    "Хочешь провести новую оценку сейчас?"
                )

                markup = types.InlineKeyboardMarkup()
                btn_yes = types.InlineKeyboardButton(
                    "✅ Да, начать",
                    callback_data="checkin:start"
                )
                btn_no = types.InlineKeyboardButton(
                    "↩️ В главное меню",
                    callback_data="menu:show"
                )
                markup.add(btn_yes)
                markup.add(btn_no)

                await bot.send_message(chat_id, text, reply_markup=markup)
            else:
                # No previous check-ins
                text = f"Ещё нет данных для оценки прогресса.\n{reason}"

                markup = types.InlineKeyboardMarkup()
                btn_menu = types.InlineKeyboardButton(
                    "↩️ Главное меню",
                    callback_data="menu:show"
                )
                markup.add(btn_menu)

                await bot.send_message(chat_id, text, reply_markup=markup)

    except Exception as e:
        print(f"Error showing check-in progress: {e}")


def register_checkin_handlers(bot):
    """Register check-in handlers"""

    @bot.callback_query_handler(func=lambda call: call.data.startswith('checkin'))
    async def checkin_callback_handler(callback_query):
        """Handle check-in callbacks"""
        try:
            data_parts = callback_query.data.split(':')
            action = data_parts[0].replace('checkin', '').lstrip('_')

            user_id = callback_query.from_user.id
            username = callback_query.from_user.username or 'Unknown'
            chat_id = callback_query.message.chat.id

            if action == 'start':
                await bot.answer_callback_query(callback_query.id)
                await start_check_in(bot, chat_id, user_id, username, scheduled=False)

            elif action == 'rate' and len(data_parts) == 3:
                # Handle problem rating
                problem_idx = data_parts[1]
                rating = data_parts[2]
                await handle_problem_rating(bot, callback_query, problem_idx, rating)

            elif action == 'goal' and len(data_parts) == 2:
                # Handle goal progress
                progress = data_parts[1]
                await handle_goal_progress(bot, callback_query, progress)

            elif action == 'later':
                # Remind later
                await bot.answer_callback_query(
                    callback_query.id,
                    "Хорошо, напомню позже. Береги себя!",
                    show_alert=True
                )
                # Return to main menu
                from universal_menu import show_main_menu
                from greeting import user_states

                user_name = 'Друг'
                form = 'ты'
                if user_id in user_states:
                    user_name = user_states[user_id].get('user_name', 'Друг')
                    form = user_states[user_id].get('form', 'ты')

                await show_main_menu(bot, chat_id, user_id, username, user_name, form)

        except Exception as e:
            print(f"Error handling check-in callback: {e}")
            await bot.answer_callback_query(callback_query.id)


# Initialize the module
ensure_checkin_file_exists()