# -*- coding: utf-8 -*-
"""
Mindfulness-Based Cognitive Therapy (MBCT) Practice Module
Provides guided mindfulness practices for users with tracking and feedback
"""

import asyncio
import os
from datetime import datetime
from telebot import types
from openpyxl import load_workbook, Workbook

# File paths
MVST_EXCEL_FILE = 'mvst.xlsx'

# List of mindfulness practices
PRACTICES = [
    {
        'id': 1,
        'name': '3-минутная дыхательная пауза',
        'short_name': 'Breathing Space',
        'description': 'Дыхательная пауза (Breathing Space) — короткая «перезагрузка»: замечаем «что есть», фокусируемся на дыхании, расширяем внимание.',
        'emoji': '🌬️'
    },
    {
        'id': 2,
        'name': 'Сканирование тела',
        'short_name': 'Body Scan',
        'description': 'Сканирование тела (Body Scan) — внимание проходит от макушки к стопам, замечаем ощущения безоценочно, развиваем заземление.',
        'emoji': '🧘'
    },
    {
        'id': 3,
        'name': 'Осознанное дыхание',
        'short_name': 'Mindful Breathing',
        'description': 'Осознанное дыхание (2–5 минут) — наблюдаем вдох-выдох, мягко возвращаем внимание.',
        'emoji': '🫁'
    },
    {
        'id': 4,
        'name': 'Осознанная ходьба/движение',
        'short_name': 'Mindful Walking',
        'description': 'Осознанная ходьба/движение — замечаем ощущения в теле при каждом шаге; тренируем присутствие в движении.',
        'emoji': '🚶'
    },
    {
        'id': 5,
        'name': 'Мысли как мысли',
        'short_name': 'Decentering',
        'description': 'Мысли как мысли (Decentering) — воспринимаем мысли как события ума (облака/листья на воде), не сливаясь с ними.',
        'emoji': '☁️'
    },
    {
        'id': 6,
        'name': 'Повернуться к трудности',
        'short_name': 'Turning Toward Difficulty',
        'description': 'Повернуться к трудности (Turning Toward Difficulty) — мягко встречаем неприятное ощущение/эмоцию, дышим рядом, расширяем внимание.',
        'emoji': '💛'
    }
]

# Store user mindfulness states
user_mvst_states = {}


def init_mvst_excel():
    """Initialize MVST Excel file with headers"""
    if not os.path.exists(MVST_EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Practices'
        ws['A1'] = 'User ID'
        ws['B1'] = 'Username'
        ws['C1'] = 'Practice Name'
        ws['D1'] = 'Practice Type'
        ws['E1'] = 'Practice Start Time'
        ws['F1'] = 'User Input During Practice'
        ws['G1'] = 'What Was Noticed'
        ws['H1'] = 'What Was Useful'
        ws['I1'] = 'What Was Difficult'
        ws['J1'] = 'Date Time'
        wb.save(MVST_EXCEL_FILE)


def save_practice_to_excel(user_id, username, practice_name, practice_type):
    """Save practice selection to mvst.xlsx"""
    try:
        if not os.path.exists(MVST_EXCEL_FILE):
            init_mvst_excel()

        wb = load_workbook(MVST_EXCEL_FILE)
        ws = wb.active

        next_row = ws.max_row + 1

        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = practice_name
        ws[f'D{next_row}'] = practice_type
        ws[f'E{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'J{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save(MVST_EXCEL_FILE)
        print(f"Practice saved: {username} - {practice_name}")

    except Exception as e:
        print(f"Error saving practice to Excel: {e}")


def save_practice_user_input_to_excel(user_id, practice_name, user_input):
    """Save user input during practice to mvst.xlsx"""
    try:
        if not os.path.exists(MVST_EXCEL_FILE):
            init_mvst_excel()

        wb = load_workbook(MVST_EXCEL_FILE)
        ws = wb.active

        # Find the last row for this user/practice and update it
        for row in range(ws.max_row, 0, -1):
            if (ws[f'A{row}'].value == user_id and
                ws[f'C{row}'].value == practice_name):
                ws[f'F{row}'] = user_input
                break

        wb.save(MVST_EXCEL_FILE)
        print(f"Practice user input saved: {practice_name}")

    except Exception as e:
        print(f"Error saving practice user input to Excel: {e}")


def save_practice_final_answers_to_excel(user_id, practice_name, noticed, useful, difficult):
    """Save final answers (noticed, useful, difficult) to mvst.xlsx"""
    try:
        if not os.path.exists(MVST_EXCEL_FILE):
            init_mvst_excel()

        wb = load_workbook(MVST_EXCEL_FILE)
        ws = wb.active

        # Find the last row for this user/practice and update it
        for row in range(ws.max_row, 0, -1):
            if (ws[f'A{row}'].value == user_id and
                ws[f'C{row}'].value == practice_name):
                ws[f'G{row}'] = noticed
                ws[f'H{row}'] = useful
                ws[f'I{row}'] = difficult
                break

        wb.save(MVST_EXCEL_FILE)
        print(f"Practice final answers saved: {practice_name}")

    except Exception as e:
        print(f"Error saving practice final answers to Excel: {e}")


async def show_mindfulness_practices(bot, chat_id, user_id, username):
    """
    Show list of mindfulness practices
    """
    try:
        # Initialize user state
        user_mvst_states[user_id] = {
            'practices': PRACTICES,
            'selected_practice': None,
            'username': username,
            'completed_practices': [],
            'current_step': 'selection'  # selection, practice, questions, completion
        }

        header_text = "🌙 Майндфулнесс-практика (MBCT)\n\nВыбери практику для начала:"
        await bot.send_message(chat_id, header_text)

        # Pause for 1 second
        await asyncio.sleep(1)

        # Show practice cards
        for practice in PRACTICES:
            emoji = practice['emoji']
            card_text = f"{emoji} {practice['name']}\n{practice['short_name']}\n\n{practice['description']}"

            markup = types.InlineKeyboardMarkup()
            btn_start = types.InlineKeyboardButton(
                f"Начать: {practice['name']}",
                callback_data=f"mvst_select:{practice['id']}"
            )
            markup.add(btn_start)

            await bot.send_message(chat_id, card_text, reply_markup=markup)

        # Add menu button
        from universal_menu import get_menu_button
        menu_markup = get_menu_button()
        await bot.send_message(chat_id, "Выбери практику или вернись в меню", reply_markup=menu_markup)

    except Exception as e:
        print(f"Error showing mindfulness practices: {e}")


async def handle_practice_select(bot, callback_query, practice_id):
    """
    Handle practice selection
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_mvst_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_mvst_states[user_id]
        practice_id = int(practice_id)

        # Find selected practice
        selected_practice = None
        for practice in state['practices']:
            if practice['id'] == practice_id:
                selected_practice = practice
                break

        if not selected_practice:
            await bot.answer_callback_query(callback_query.id)
            return

        state['selected_practice'] = selected_practice
        state['current_step'] = 'practice'

        # Save practice selection
        save_practice_to_excel(user_id, username, selected_practice['name'], selected_practice['short_name'])

        await bot.answer_callback_query(callback_query.id)

        # Show practice description
        emoji = selected_practice['emoji']
        practice_text = f"{emoji} {selected_practice['name']}\n\n{selected_practice['description']}"
        await bot.send_message(chat_id, practice_text)

        # Pause for 1 second
        await asyncio.sleep(1)

        # Show prompt for user input
        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(
            chat_id,
            "Поделись, что ты замечаешь во время практики, или просто начни — я помогу тебе дальше:",
            reply_markup=markup
        )

        # Mark that we're awaiting practice input
        state['awaiting_practice_input'] = True

    except Exception as e:
        print(f"Error handling practice selection: {e}")
        await bot.answer_callback_query(callback_query.id)


async def show_final_questions(bot, chat_id, user_id):
    """
    Show final questions after practice
    """
    try:
        if user_id not in user_mvst_states:
            return

        state = user_mvst_states[user_id]
        state['current_final_question'] = 0
        state['final_answers'] = {}
        state['current_step'] = 'questions'

        # Show first question
        await show_final_question(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error showing final questions: {e}")


async def show_final_question(bot, chat_id, user_id):
    """
    Show current final question
    """
    try:
        if user_id not in user_mvst_states:
            return

        state = user_mvst_states[user_id]
        question_idx = state['current_final_question']

        questions = [
            "Что ты заметил(а) в ходе практики?",
            "Что было полезно?",
            "Что вызвало сложности?"
        ]

        if question_idx >= len(questions):
            # All questions answered - show completion options
            await show_practice_completion_options(bot, chat_id, user_id)
            return

        question = questions[question_idx]
        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(chat_id, question, reply_markup=markup)

        state['awaiting_final_answer'] = True

    except Exception as e:
        print(f"Error showing final question: {e}")


async def show_practice_completion_options(bot, chat_id, user_id):
    """
    Show completion options after all final questions are answered
    """
    try:
        if user_id not in user_mvst_states:
            return

        state = user_mvst_states[user_id]

        markup = types.InlineKeyboardMarkup()

        btn_mark_complete = types.InlineKeyboardButton(
            "✅ Отметить как завершённое",
            callback_data="mvst_mark_complete"
        )

        markup.add(btn_mark_complete)
        await bot.send_message(chat_id, "Отлично! Ты выполнил(а) практику.", reply_markup=markup)

    except Exception as e:
        print(f"Error showing practice completion options: {e}")


async def finish_practice(bot, chat_id, user_id):
    """
    Finish practice and save answers, then show next practice options
    """
    try:
        if user_id not in user_mvst_states:
            return

        state = user_mvst_states[user_id]
        username = state.get('username', 'Unknown')
        selected_practice = state['selected_practice']

        noticed = state['final_answers'].get(0, '')
        useful = state['final_answers'].get(1, '')
        difficult = state['final_answers'].get(2, '')

        # Save final answers
        save_practice_final_answers_to_excel(user_id, selected_practice['name'], noticed, useful, difficult)

        # Show next practice options
        await show_next_practice_options(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error finishing practice: {e}")


async def show_next_practice_options(bot, chat_id, user_id):
    """
    Show all remaining practices after completing one
    """
    try:
        if user_id not in user_mvst_states:
            return

        state = user_mvst_states[user_id]
        practices = state.get('practices', [])
        current_practice = state.get('selected_practice')

        # Track completed practices in state
        if 'completed_practices' not in state:
            state['completed_practices'] = []

        # Mark current practice as completed
        if current_practice and current_practice['id'] not in state['completed_practices']:
            state['completed_practices'].append(current_practice['id'])

        # Find remaining practices
        remaining_practices = [p for p in practices if p['id'] not in state['completed_practices']]

        if remaining_practices:
            # Show remaining practices
            await bot.send_message(
                chat_id,
                "Отлично! ✨ Вот другие практики для тебя:"
            )

            # Display each remaining practice with selection button
            for practice in practices:
                if practice['id'] not in state['completed_practices']:
                    emoji = practice['emoji']
                    card_text = f"{emoji} {practice['name']}\n{practice['short_name']}\n\n{practice['description']}"

                    # Create button
                    markup = types.InlineKeyboardMarkup()
                    btn_select = types.InlineKeyboardButton(
                        f"Начать: {practice['name']}",
                        callback_data=f"mvst_select:{practice['id']}"
                    )
                    markup.add(btn_select)

                    await bot.send_message(chat_id, card_text, reply_markup=markup)

            # Add final menu button
            markup_final = types.InlineKeyboardMarkup()
            btn_menu = types.InlineKeyboardButton(
                "📍 Главное меню",
                callback_data="menu:show"
            )
            markup_final.add(btn_menu)
            await bot.send_message(
                chat_id,
                "Или вернись в главное меню:",
                reply_markup=markup_final
            )
        else:
            # All practices completed
            markup = types.InlineKeyboardMarkup()
            btn_menu = types.InlineKeyboardButton(
                "📍 Главное меню",
                callback_data="menu:show"
            )
            markup.add(btn_menu)

            await bot.send_message(
                chat_id,
                "Поздравляю! 🎉 Ты выполнил(а) все практики!",
                reply_markup=markup
            )

    except Exception as e:
        print(f"Error showing next practice options: {e}")


def validate_practice_input(text):
    """
    Validate practice input text.
    Returns: (is_valid, feedback_message)
    Optional validation - allow short answers for practices
    """
    text = text.strip()

    # For mindfulness practices, we're more lenient - allow optional input
    if not text:
        return True, None

    return True, None


async def handle_practice_text_input(bot, message):
    """
    Handle practice text input
    """
    try:
        user_id = message.from_user.id
        username = message.from_user.username or 'Unknown'
        text = message.text

        if user_id not in user_mvst_states:
            return

        state = user_mvst_states[user_id]

        # Check if awaiting practice input
        if state.get('awaiting_practice_input'):
            await handle_practice_input(bot, message, user_id, username, text, state)
            return

        # Check if awaiting final answer
        if state.get('awaiting_final_answer'):
            await handle_final_answer_input(bot, message, user_id, username, text, state)
            return

    except Exception as e:
        print(f"Error handling practice text input: {e}")


async def handle_practice_input(bot, message, user_id, username, text, state):
    """
    Handle input during practice
    """
    try:
        # Validate input (lenient for practices)
        is_valid, feedback = validate_practice_input(text)

        if not is_valid and feedback:
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(message.chat.id, feedback, reply_markup=markup)
            return

        # Store text temporarily
        state['pending_practice_input'] = text

        # Show preview
        preview_text = f"📝 Вот что ты написал(а):\n\n{text}\n\nВсё верно?" if text else "Готов(а) продолжить?"

        markup = types.InlineKeyboardMarkup()

        btn_confirm = types.InlineKeyboardButton(
            "✅ Подтвердить",
            callback_data="mvst_input_confirm:yes"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="mvst_input_confirm:edit"
        )
        btn_menu = types.InlineKeyboardButton(
            "📍 Главное меню",
            callback_data="menu:show"
        )

        markup.add(btn_confirm)
        if text:
            markup.add(btn_edit)
        markup.add(btn_menu)

        await bot.send_message(message.chat.id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling practice input: {e}")


async def handle_practice_input_confirm(bot, callback_query, action):
    """
    Handle practice input confirmation
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_mvst_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_mvst_states[user_id]
        pending_input = state.get('pending_practice_input', '')

        if action == "yes":
            # Save the input if provided
            selected_practice = state['selected_practice']
            if pending_input:
                save_practice_user_input_to_excel(user_id, selected_practice['name'], pending_input)

            await bot.answer_callback_query(callback_query.id, "Спасибо! Продолжаем.")

            state['awaiting_practice_input'] = False
            state['pending_practice_input'] = None

            # Move to final questions
            await show_final_questions(bot, chat_id, user_id)

        elif action == "edit":
            # Ask to re-enter
            state['pending_practice_input'] = None
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                chat_id,
                "Окей, введи свой ответ заново или нажми подтвердить для продолжения:"
            )

    except Exception as e:
        print(f"Error handling practice input confirm: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_final_answer_input(bot, message, user_id, username, text, state):
    """
    Handle final answer input
    """
    try:
        # Validate input (lenient for practices)
        is_valid, feedback = validate_practice_input(text)

        if not is_valid and feedback:
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(message.chat.id, feedback, reply_markup=markup)
            return

        # Store answer temporarily
        state['pending_final_answer'] = text

        # Show preview
        preview_text = f"📝 Вот что ты написал(а):\n\n{text}\n\nВсё верно?" if text else "Готов(а) продолжить?"

        markup = types.InlineKeyboardMarkup()

        btn_confirm = types.InlineKeyboardButton(
            "✅ Подтвердить",
            callback_data="mvst_answer_confirm:yes"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="mvst_answer_confirm:edit"
        )
        btn_menu = types.InlineKeyboardButton(
            "📍 Меню",
            callback_data="menu:show"
        )

        markup.add(btn_confirm)
        if text:
            markup.add(btn_edit)
        markup.add(btn_menu)

        await bot.send_message(message.chat.id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling final answer input: {e}")


async def handle_answer_confirm(bot, callback_query, action):
    """
    Handle final answer confirmation
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_mvst_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_mvst_states[user_id]
        pending_answer = state.get('pending_final_answer', '')
        question_idx = state['current_final_question']

        if action == "yes":
            # Save the answer
            state['final_answers'][question_idx] = pending_answer

            await bot.answer_callback_query(callback_query.id, "Спасибо! Записано.")

            # Move to next final question
            state['current_final_question'] += 1
            state['awaiting_final_answer'] = False
            state['pending_final_answer'] = None

            # Show next question or finish
            await show_final_question(bot, chat_id, user_id)

        elif action == "edit":
            # Ask to re-enter
            state['pending_final_answer'] = None
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                chat_id,
                "Окей, введи свой ответ заново:"
            )

    except Exception as e:
        print(f"Error handling answer confirm: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_mark_practice_complete(bot, callback_query):
    """
    Handle marking practice as completed
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_mvst_states:
            await bot.answer_callback_query(callback_query.id)
            return

        await bot.answer_callback_query(callback_query.id, "✅ Практика завершена!")

        # Finish the practice (save data)
        await finish_practice(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error marking practice complete: {e}")
        await bot.answer_callback_query(callback_query.id)


def register_mvst_handlers(bot):
    """
    Register mindfulness practice handlers
    """
    @bot.callback_query_handler(func=lambda call: call.data.startswith('mvst_select:'))
    async def mvst_select_handler(callback_query):
        """Handle practice selection"""
        practice_id = callback_query.data.split(':')[1]
        await handle_practice_select(bot, callback_query, practice_id)

    @bot.callback_query_handler(func=lambda call: call.data.startswith('mvst_input_confirm:'))
    async def mvst_input_confirm_handler(callback_query):
        """Handle practice input confirmation"""
        action = callback_query.data.split(':')[1]
        await handle_practice_input_confirm(bot, callback_query, action)

    @bot.callback_query_handler(func=lambda call: call.data.startswith('mvst_answer_confirm:'))
    async def mvst_answer_confirm_handler(callback_query):
        """Handle final answer confirmation"""
        action = callback_query.data.split(':')[1]
        await handle_answer_confirm(bot, callback_query, action)

    @bot.callback_query_handler(func=lambda call: call.data == 'mvst_mark_complete')
    async def mvst_mark_complete_handler(callback_query):
        """Handle marking practice as completed"""
        await handle_mark_practice_complete(bot, callback_query)
