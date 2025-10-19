# -*- coding: utf-8 -*-
"""
Diary module for emotion and thought recording
Allows users to record free-form emotional reflections and saves to Excel
"""

import os
from datetime import datetime
from telebot import types
from openpyxl import load_workbook, Workbook

# Path to the diary data file
DIARY_FILE = 'diary.xlsx'

# Store user diary states
# Format: {user_id: {'awaiting_entry': bool, 'user_name': str, 'username': str}}
user_diary_states = {}


def init_diary_file():
    """Initialize diary Excel file with headers if it doesn't exist"""
    try:
        if not os.path.exists(DIARY_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Diary'

            # Add headers
            ws['A1'] = 'User ID'
            ws['B1'] = 'Username'
            ws['C1'] = 'User Name'
            ws['D1'] = 'Entry Type'
            ws['E1'] = 'Entry Text'
            ws['F1'] = 'Progress Rating (0-10)'
            ws['G1'] = 'Date Time'

            wb.save(DIARY_FILE)
            print(f"Diary file initialized: {DIARY_FILE}")
    except Exception as e:
        print(f"Error initializing diary file: {e}")


def save_diary_entry(user_id, username, user_name, entry_text, progress_rating=None):
    """
    Save diary entry to Excel file

    Args:
        user_id (int): Telegram user ID
        username (str): Telegram username
        user_name (str): User's name (from greeting)
        entry_text (str): The diary entry text
        progress_rating (int/str): User's progress rating (0-10), optional
    """
    try:
        if not os.path.exists(DIARY_FILE):
            init_diary_file()

        wb = load_workbook(DIARY_FILE)
        ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        # Add diary entry
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = user_name
        ws[f'D{next_row}'] = 'diary_entry'
        ws[f'E{next_row}'] = entry_text
        ws[f'F{next_row}'] = progress_rating if progress_rating else ''
        ws[f'G{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save(DIARY_FILE)
        print(f"Diary entry saved for user {username}: {entry_text[:50]}...")

    except Exception as e:
        print(f"Error saving diary entry: {e}")


async def show_diary_prompt(bot, chat_id, user_id, username, user_name):
    """
    Show diary entry prompt to user

    Args:
        bot: Telegram bot instance
        chat_id: Chat ID
        user_id: User ID
        username: Username
        user_name: User's name
    """
    try:
        # Store state - awaiting diary entry text
        user_diary_states[user_id] = {
            'stage': 'awaiting_text',
            'entry_text': None,
            'progress_rating': None,
            'user_name': user_name,
            'username': username
        }

        text = (
            "📖 Дневник: Эмоции и мысли\n\n"
            "Напиши, как ты сейчас себя чувствуешь, какие мысли/эмоции есть и что на них повлияло.\n"
            "Также оцени, насколько ты продвинулся(ась) к своей цели от 0 до 10.\n\n"
            "Ты можешь писать в свободной форме - это может быть несколько слов или целый рассказ. 💭\n\n"
            "Отправь своё сообщение:"
        )

        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(chat_id, text, reply_markup=markup)

    except Exception as e:
        print(f"Error showing diary prompt: {e}")


async def handle_diary_entry(bot, message):
    """
    Handle incoming diary entry text from user

    Args:
        bot: Telegram bot instance
        message: Telegram message object
    """
    try:
        user_id = message.from_user.id

        if user_id not in user_diary_states:
            return

        state = user_diary_states[user_id]

        # Check if we're awaiting text entry
        if state.get('stage') != 'awaiting_text':
            return

        entry_text = message.text

        # Store the entry text
        state['entry_text'] = entry_text
        state['stage'] = 'preview'

        # Show preview with confirmation buttons
        await show_diary_preview(bot, message.chat.id, user_id, entry_text)

    except Exception as e:
        print(f"Error handling diary entry: {e}")


async def show_diary_preview(bot, chat_id, user_id, entry_text):
    """
    Show preview of diary entry with confirmation buttons

    Args:
        bot: Telegram bot instance
        chat_id: Chat ID
        user_id: User ID
        entry_text: The diary entry text to preview
    """
    try:
        text = (
            "📋 Предпросмотр твоей записи:\n\n"
            f"{entry_text}\n\n"
            "Что ты хочешь сделать?"
        )

        markup = types.InlineKeyboardMarkup()

        btn_confirm = types.InlineKeyboardButton(
            "✅ Подтвердить",
            callback_data="diary:confirm"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="diary:edit"
        )
        btn_back = types.InlineKeyboardButton(
            "⬅️ Вернуться",
            callback_data="diary:back"
        )

        markup.row(btn_confirm)
        markup.row(btn_edit)
        markup.row(btn_back)

        await bot.send_message(chat_id, text, reply_markup=markup)

    except Exception as e:
        print(f"Error showing diary preview: {e}")


async def handle_diary_confirm(bot, callback_query):
    """
    Handle diary entry confirmation

    Args:
        bot: Telegram bot instance
        callback_query: Callback query from button press
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_diary_states:
            try:
                await bot.answer_callback_query(callback_query.id, "Ошибка состояния", show_alert=True)
            except Exception:
                pass  # Query may have expired
            return

        state = user_diary_states[user_id]
        entry_text = state.get('entry_text')

        if not entry_text:
            try:
                await bot.answer_callback_query(callback_query.id, "Ошибка: нет текста записи", show_alert=True)
            except Exception:
                pass  # Query may have expired
            return

        # Save to Excel
        save_diary_entry(
            user_id,
            state['username'],
            state['user_name'],
            entry_text,
            state.get('progress_rating')
        )

        # Check for crisis indicators in diary entry
        from safety_check import check_text_safety, show_crisis_support, log_crisis_detection

        crisis_detected, crisis_type, confidence = await check_text_safety(
            text=entry_text,
            context="diary"
        )

        # Clear state
        del user_diary_states[user_id]

        # Answer callback with error handling
        try:
            await bot.answer_callback_query(callback_query.id, "✅ Запись сохранена!", show_alert=False)
        except Exception:
            pass  # Query may have expired

        # Send confirmation message first
        await bot.send_message(
            chat_id,
            "✅ Твоя запись сохранена в дневнике 💭\n\n"
            "Спасибо, что делишься своими чувствами. "
            "Это первый шаг к лучшему пониманию себя."
        )

        # Check if crisis was detected
        if crisis_detected and crisis_type:
            # Log crisis detection
            await log_crisis_detection(
                user_id=user_id,
                username=state['username'],
                crisis_type=crisis_type,
                context="diary",
                text_sample=entry_text[:200],
                file_path='diary.xlsx'
            )

            # Show crisis support
            await show_crisis_support(
                bot=bot,
                chat_id=chat_id,
                user_name=state['user_name'],
                crisis_type=crisis_type,
                context="diary",
                continue_after=False  # Don't show continue option, go to menu
            )
        else:
            # No crisis - show main menu as usual
            from universal_menu import show_main_menu
            from greeting import user_states
            form_of_address = 'ты'
            if user_id in user_states:
                form_of_address = user_states[user_id].get('form', 'ты')
            await show_main_menu(bot, chat_id, user_id, state['username'], state['user_name'], form_of_address)

    except Exception as e:
        print(f"Error confirming diary entry: {e}")


async def handle_diary_edit(bot, callback_query):
    """
    Handle diary entry edit - return to text input stage

    Args:
        bot: Telegram bot instance
        callback_query: Callback query from button press
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_diary_states:
            try:
                await bot.answer_callback_query(callback_query.id, "Ошибка состояния", show_alert=True)
            except Exception:
                pass  # Query may have expired
            return

        state = user_diary_states[user_id]
        state['stage'] = 'awaiting_text'

        # Answer callback with error handling
        try:
            await bot.answer_callback_query(callback_query.id, "Введи новую запись", show_alert=False)
        except Exception:
            pass  # Query may have expired

        # Send message prompting for new entry
        text = (
            "✏️ Давай напишем заново.\n\n"
            "Напиши, как ты сейчас себя чувствуешь, какие мысли/эмоции есть и что на них повлияло.\n"
            "Также оцени, насколько ты продвинулся(ась) к своей цели от 0 до 10."
        )

        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(chat_id, text, reply_markup=markup)

    except Exception as e:
        print(f"Error editing diary entry: {e}")


async def handle_diary_back(bot, callback_query):
    """
    Handle diary entry cancellation and return to menu

    Args:
        bot: Telegram bot instance
        callback_query: Callback query from button press
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_diary_states:
            try:
                await bot.answer_callback_query(callback_query.id, "Ошибка состояния", show_alert=True)
            except Exception:
                pass  # Query may have expired
            return

        state = user_diary_states[user_id]

        # Clear state
        del user_diary_states[user_id]

        # Answer callback with error handling
        try:
            await bot.answer_callback_query(callback_query.id, "Отменено", show_alert=False)
        except Exception:
            pass  # Query may have expired

        # Show main menu
        from universal_menu import show_main_menu
        from greeting import user_states
        form_of_address = 'ты'
        if user_id in user_states:
            form_of_address = user_states[user_id].get('form', 'ты')
        await show_main_menu(bot, chat_id, user_id, state['username'], state['user_name'], form_of_address)

    except Exception as e:
        print(f"Error handling diary back: {e}")
