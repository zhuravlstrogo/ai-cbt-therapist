# -*- coding: utf-8 -*-
"""
Protocol Unknown module for AI Psychologist bot
Handles questionnaire flow when user doesn't know which protocol they need
"""

import os
import asyncio
from datetime import datetime
from openpyxl import load_workbook, Workbook
from telebot import types
import json
from typing import Dict, Any, Optional, List

# Store user questionnaire states
# Format: {user_id: {'stage': str, 'user_name': str, 'symptoms': list, 'duration': str, 'impact': int, 'details': list}}
questionnaire_states = {}

# Excel file path
EXCEL_FILE = 'messages.xlsx'

# Symptom categories with their details
SYMPTOM_DETAILS = {
    'anxiety': {
        'name': 'Тревога',
        'details': [
            'Частое напряжение',
            'Трудно остановить поток мыслей',
            'Избегаю ситуаций',
            'Тело "на взводе"'
        ]
    },
    'apathy': {
        'name': 'Потеря интереса',
        'details': [
            'Потеря интереса',
            'Усталость, апатия',
            'Чувство вины',
            'Безразличие к происходящему'
        ]
    },
    'low_mood': {
        'name': 'Сниженное настроение',
        'details': [
            'Грусть большую часть дня',
            'Негативные мысли о будущем',
            'Слезливость',
            'Ощущение безнадежности'
        ]
    },
    'sleep': {
        'name': 'Проблемы со сном',
        'details': [
            'Долго засыпаю',
            'Просыпаюсь ночью',
            'Думаю о сне слишком много',
            'Сплю, но не восстанавливаюсь'
        ]
    },
    'procrastination': {
        'name': 'Прокрастинация',
        'details': [
            'Откладываю важные дела',
            'Трудно начать задачу',
            'Чувствую вину за откладывание',
            'Теряю мотивацию'
        ]
    },
    'communication': {
        'name': 'Трудности в общении',
        'details': [
            'Избегаю конфликтов',
            'Трудно выражать свои чувства',
            'Не могу сказать "нет"',
            'Чувствую непонимание'
        ]
    },
    'self_criticism': {
        'name': 'Самокритичность',
        'details': [
            'Постоянно себя критикую',
            'Чувство вины без причины',
            'Считаю себя неудачником',
            'Не верю в свои силы'
        ]
    },
    'irritability': {
        'name': 'Раздражительность',
        'details': [
            'Легко выхожу из себя',
            'Злюсь на близких',
            'Трудно сдерживать эмоции',
            'Чувствую постоянное напряжение'
        ]
    },
    'obsessive': {
        'name': 'Навязчивые мысли',
        'details': [
            'Мысли крутятся по кругу',
            'Выполняю ритуалы',
            'Постоянно проверяю',
            'Не могу отпустить мысли'
        ]
    },
    'panic': {
        'name': 'Панические атаки',
        'details': [
            'Внезапный страх',
            'Учащенное сердцебиение',
            'Страх потерять контроль',
            'Избегаю мест где были атаки'
        ]
    },
    'social_anxiety': {
        'name': 'Неуверенность в компаниях',
        'details': [
            'Страх оценки других',
            'Избегаю выступлений',
            'Тревога в новых компаниях',
            'Боюсь показаться глупым'
        ]
    },
    'trauma': {
        'name': 'Травматичный опыт',
        'details': [
            'Воспоминания о событии',
            'Кошмары',
            'Избегаю напоминаний',
            'Чувство отстраненности'
        ]
    },
    'eating': {
        'name': 'Проблемы с питанием',
        'details': [
            'Контролирую питание',
            'Недоволен телом',
            'Переедаю при стрессе',
            'Чувство вины после еды'
        ]
    },
    'somatic': {
        'name': 'Ощущения в теле',
        'details': [
            'Боли без причины',
            'Напряжение в теле',
            'Головокружение',
            'Проблемы с дыханием'
        ]
    },
    'perfectionism': {
        'name': 'Перфекционизм',
        'details': [
            'Страх ошибок',
            'Все должно быть идеально',
            'Откладываю из-за страха',
            'Критикую себя за неидеальность'
        ]
    },
    'grief': {
        'name': 'Переживание утраты',
        'details': [
            'Тоска по утраченному',
            'Трудно принять потерю',
            'Чувство пустоты',
            'Избегаю напоминаний'
        ]
    },
    'stress': {
        'name': 'Стресс, выгорание',
        'details': [
            'Постоянная усталость',
            'Потеря интереса к работе',
            'Чувство перегрузки',
            'Цинизм и раздражение'
        ]
    },
    'resilience': {
        'name': 'Укрепление устойчивости',
        'details': [
            'Хочу лучше справляться',
            'Развить эмоциональный интеллект',
            'Научиться управлять стрессом',
            'Повысить уверенность'
        ]
    }
}


def save_questionnaire_to_excel(user_id: int, username: str, data: Dict[str, Any]):
    """Save questionnaire responses to Excel file"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Messages'
            # Add headers if needed
            headers = ['User ID', 'Username', 'User Name', 'Message Text', 'Message Type',
                      'Protocol Choice', 'Date Time', 'Questionnaire Data']
            for i, header in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=header)

        # Find next empty row
        next_row = ws.max_row + 1

        # Save questionnaire data
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'D{next_row}'] = f"Questionnaire completed"
        ws[f'E{next_row}'] = 'questionnaire_response'
        ws[f'G{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'H{next_row}'] = json.dumps(data, ensure_ascii=False)

        wb.save(EXCEL_FILE)
        print(f"Questionnaire data saved to Excel for user {username}")
    except Exception as e:
        print(f"Error saving questionnaire to Excel: {e}")


async def start_questionnaire(bot, chat_id, user_id, username, user_name):
    """Start the questionnaire flow for protocol selection"""
    try:
        # Initialize user state
        questionnaire_states[user_id] = {
            'stage': 'symptoms_selection',
            'user_name': user_name,
            'symptoms': [],
            'duration': None,
            'impact': None,
            'details': []
        }

        # Send first question with symptoms buttons
        intro_text = f"Поняла 😊\n{user_name}, давай разберёмся вместе. Скажи, пожалуйста, что тебя беспокоит в последнее время?"

        await bot.send_message(chat_id, intro_text)

        # Small delay for better UX
        await asyncio.sleep(1)

        # Create inline keyboard with symptom options
        markup = types.InlineKeyboardMarkup(row_width=1)

        symptoms = [
            ("😟 Тревога, беспокойство", "symptom_anxiety"),
            ("😞 Потеря интереса, апатия", "symptom_apathy"),
            ("😔 Сниженное настроение", "symptom_low_mood"),
            ("💤 Проблемы со сном", "symptom_sleep"),
            ("⏳ Прокрастинация, снижение сил и мотивации", "symptom_procrastination"),
            ("💬 Трудности в общении", "symptom_communication"),
            ("💔 Самокритичность, чувство вины", "symptom_self_criticism"),
            ("😤 Раздражительность, вспышки гнева", "symptom_irritability"),
            ("🌀 Навязчивые мысли, действия", "symptom_obsessive"),
            ("💥 Панические атаки", "symptom_panic"),
            ("🎭 Неуверенность в компаниях людей", "symptom_social_anxiety"),
            ("🌧 Пережитый травматичный опыт", "symptom_trauma"),
            ("🍽 Проблемы с питанием или принятием своего тела", "symptom_eating"),
            ("🩺 Неприятные ощущения в теле, не связанные с состоянием физического здоровья", "symptom_somatic"),
            ("🎯 Перфекционизм", "symptom_perfectionism"),
            ("🌻 Болезненное переживание утраты или жизненные перемены", "symptom_grief"),
            ("🔄 Стресс, усталость, выгорание", "symptom_stress"),
            ("💡 Просто хочу укрепить устойчивость", "symptom_resilience")
        ]

        for text, callback in symptoms:
            markup.add(types.InlineKeyboardButton(text, callback_data=callback))

        # Add continue button (will be enabled after selection)
        markup.add(types.InlineKeyboardButton("✅ Продолжить", callback_data="symptoms_continue"))

        # Add menu button
        markup.add(types.InlineKeyboardButton("📱 Главное меню", callback_data="menu:show"))

        await bot.send_message(
            chat_id,
            "Выбери все, что подходит (можно несколько вариантов):",
            reply_markup=markup
        )

        print(f"Questionnaire started for user {username}")

    except Exception as e:
        print(f"Error starting questionnaire: {e}")
        await bot.send_message(chat_id, "Произошла ошибка. Попробуйте начать заново с /start")


async def handle_symptom_selection(bot, callback_query, user_id, username):
    """Handle symptom selection from the first question"""
    try:
        if user_id not in questionnaire_states:
            await bot.answer_callback_query(
                callback_query.id,
                "Сессия истекла. Пожалуйста, начните заново с /start"
            )
            return

        data = callback_query.data

        if data == "symptoms_continue":
            # Check if at least one symptom is selected
            if not questionnaire_states[user_id]['symptoms']:
                await bot.answer_callback_query(
                    callback_query.id,
                    "Пожалуйста, выберите хотя бы один пункт",
                    show_alert=True
                )
                return

            # Move to duration question
            await ask_duration_question(bot, callback_query.message.chat.id, user_id)
            await bot.answer_callback_query(callback_query.id)

        elif data.startswith("symptom_"):
            symptom_key = data.replace("symptom_", "")
            symptoms_list = questionnaire_states[user_id]['symptoms']

            # Toggle symptom selection
            if symptom_key in symptoms_list:
                symptoms_list.remove(symptom_key)
                await bot.answer_callback_query(callback_query.id, "Убрано из списка")
            else:
                symptoms_list.append(symptom_key)
                await bot.answer_callback_query(callback_query.id, "Добавлено в список")

            # Update the keyboard to show checkmarks on selected items
            await update_symptoms_keyboard(bot, callback_query.message, user_id)

    except Exception as e:
        print(f"Error handling symptom selection: {e}")
        await bot.answer_callback_query(callback_query.id, "Произошла ошибка")


async def update_symptoms_keyboard(bot, message, user_id):
    """Update the symptoms selection keyboard to show selected items"""
    try:
        if user_id not in questionnaire_states:
            return

        selected_symptoms = questionnaire_states[user_id]['symptoms']

        # Create updated inline keyboard
        markup = types.InlineKeyboardMarkup(row_width=1)

        symptoms = [
            ("😟 Тревога, беспокойство", "symptom_anxiety", "anxiety"),
            ("😞 Потеря интереса, апатия", "symptom_apathy", "apathy"),
            ("😔 Сниженное настроение", "symptom_low_mood", "low_mood"),
            ("💤 Проблемы со сном", "symptom_sleep", "sleep"),
            ("⏳ Прокрастинация, снижение сил и мотивации", "symptom_procrastination", "procrastination"),
            ("💬 Трудности в общении", "symptom_communication", "communication"),
            ("💔 Самокритичность, чувство вины", "symptom_self_criticism", "self_criticism"),
            ("😤 Раздражительность, вспышки гнева", "symptom_irritability", "irritability"),
            ("🌀 Навязчивые мысли, действия", "symptom_obsessive", "obsessive"),
            ("💥 Панические атаки", "symptom_panic", "panic"),
            ("🎭 Неуверенность в компаниях людей", "symptom_social_anxiety", "social_anxiety"),
            ("🌧 Пережитый травматичный опыт", "symptom_trauma", "trauma"),
            ("🍽 Проблемы с питанием или принятием своего тела", "symptom_eating", "eating"),
            ("🩺 Неприятные ощущения в теле, не связанные с состоянием физического здоровья", "symptom_somatic", "somatic"),
            ("🎯 Перфекционизм", "symptom_perfectionism", "perfectionism"),
            ("🌻 Болезненное переживание утраты или жизненные перемены", "symptom_grief", "grief"),
            ("🔄 Стресс, усталость, выгорание", "symptom_stress", "stress"),
            ("💡 Просто хочу укрепить устойчивость", "symptom_resilience", "resilience")
        ]

        for text, callback, key in symptoms:
            # Add checkmark if selected
            if key in selected_symptoms:
                button_text = "✅ " + text
            else:
                button_text = text
            markup.add(types.InlineKeyboardButton(button_text, callback_data=callback))

        # Add continue button - enable it only if at least one symptom is selected
        if selected_symptoms:
            continue_text = f"✅ Продолжить ({len(selected_symptoms)} выбрано)"
        else:
            continue_text = "✅ Продолжить"

        markup.add(types.InlineKeyboardButton(continue_text, callback_data="symptoms_continue"))

        # Add menu button
        markup.add(types.InlineKeyboardButton("📱 Главное меню", callback_data="menu:show"))

        # Edit the message with updated keyboard
        await bot.edit_message_reply_markup(
            chat_id=message.chat.id,
            message_id=message.message_id,
            reply_markup=markup
        )

    except Exception as e:
        print(f"Error updating symptoms keyboard: {e}")


async def ask_duration_question(bot, chat_id, user_id):
    """Ask how long the user has been experiencing difficulties"""
    try:
        questionnaire_states[user_id]['stage'] = 'duration_selection'

        question_text = "Как давно ты замечаешь эти трудности?"

        markup = types.InlineKeyboardMarkup(row_width=1)
        durations = [
            ("☐ Несколько дней", "duration_days"),
            ("☐ Пару недель", "duration_weeks"),
            ("☐ Несколько месяцев", "duration_months"),
            ("☐ Более полугода", "duration_half_year")
        ]

        for text, callback in durations:
            markup.add(types.InlineKeyboardButton(text, callback_data=callback))

        # Add menu button
        markup.add(types.InlineKeyboardButton("📱 Главное меню", callback_data="menu:show"))

        await bot.send_message(chat_id, question_text, reply_markup=markup)

    except Exception as e:
        print(f"Error asking duration question: {e}")


async def handle_duration_selection(bot, callback_query, user_id, username):
    """Handle duration selection"""
    try:
        if user_id not in questionnaire_states:
            await bot.answer_callback_query(
                callback_query.id,
                "Сессия истекла. Пожалуйста, начните заново с /start"
            )
            return

        data = callback_query.data
        duration_map = {
            'duration_days': 'Несколько дней',
            'duration_weeks': 'Пару недель',
            'duration_months': 'Несколько месяцев',
            'duration_half_year': 'Более полугода'
        }

        if data in duration_map:
            questionnaire_states[user_id]['duration'] = duration_map[data]
            await bot.answer_callback_query(callback_query.id)

            # Move to impact question
            await ask_impact_question(bot, callback_query.message.chat.id, user_id)

    except Exception as e:
        print(f"Error handling duration selection: {e}")


async def ask_impact_question(bot, chat_id, user_id):
    """Ask about the impact level on life"""
    try:
        questionnaire_states[user_id]['stage'] = 'impact_selection'

        question_text = "Насколько сильно это мешает твоей жизни?\n(0 — не мешает, 3 — очень мешает)"

        markup = types.InlineKeyboardMarkup(row_width=4)
        impacts = [
            ("0️⃣", "impact_0"),
            ("1️⃣", "impact_1"),
            ("2️⃣", "impact_2"),
            ("3️⃣", "impact_3")
        ]

        buttons = [types.InlineKeyboardButton(text, callback_data=callback)
                  for text, callback in impacts]
        markup.add(*buttons)

        # Add menu button on a new row
        markup.add(types.InlineKeyboardButton("📱 Главное меню", callback_data="menu:show"))

        await bot.send_message(chat_id, question_text, reply_markup=markup)

    except Exception as e:
        print(f"Error asking impact question: {e}")


async def handle_impact_selection(bot, callback_query, user_id, username):
    """Handle impact level selection"""
    try:
        if user_id not in questionnaire_states:
            await bot.answer_callback_query(
                callback_query.id,
                "Сессия истекла. Пожалуйста, начните заново с /start"
            )
            return

        data = callback_query.data
        impact_map = {
            'impact_0': 0,
            'impact_1': 1,
            'impact_2': 2,
            'impact_3': 3
        }

        if data in impact_map:
            questionnaire_states[user_id]['impact'] = impact_map[data]
            await bot.answer_callback_query(callback_query.id)

            # Move to details question based on primary symptom
            await ask_details_question(bot, callback_query.message.chat.id, user_id)

    except Exception as e:
        print(f"Error handling impact selection: {e}")


async def ask_details_question(bot, chat_id, user_id):
    """Ask about specific manifestations based on selected symptoms"""
    try:
        questionnaire_states[user_id]['stage'] = 'details_selection'

        # Get primary symptom (first selected)
        symptoms = questionnaire_states[user_id]['symptoms']
        if not symptoms:
            await bot.send_message(chat_id, "Произошла ошибка. Попробуйте начать заново.")
            return

        primary_symptom = symptoms[0]

        # Get details for primary symptom
        if primary_symptom not in SYMPTOM_DETAILS:
            # Fallback to general details
            details_list = [
                'Физическое напряжение',
                'Негативные мысли',
                'Изменения в поведении',
                'Эмоциональные колебания'
            ]
        else:
            details_list = SYMPTOM_DETAILS[primary_symptom]['details']

        question_text = "Какие проявления ты замечаешь чаще всего?"

        markup = types.InlineKeyboardMarkup(row_width=1)

        for i, detail in enumerate(details_list):
            callback = f"detail_{i}"
            markup.add(types.InlineKeyboardButton(f"☐ {detail}", callback_data=callback))

        # Add continue button
        markup.add(types.InlineKeyboardButton("✅ Готово", callback_data="details_continue"))

        # Add menu button
        markup.add(types.InlineKeyboardButton("📱 Главное меню", callback_data="menu:show"))

        # Store available details for reference
        questionnaire_states[user_id]['available_details'] = details_list

        await bot.send_message(chat_id, question_text, reply_markup=markup)

    except Exception as e:
        print(f"Error asking details question: {e}")


async def handle_details_selection(bot, callback_query, user_id, username):
    """Handle details selection"""
    try:
        if user_id not in questionnaire_states:
            await bot.answer_callback_query(
                callback_query.id,
                "Сессия истекла. Пожалуйста, начните заново с /start"
            )
            return

        data = callback_query.data

        if data == "details_continue":
            # Process completed questionnaire
            await process_questionnaire_results(bot, callback_query.message.chat.id, user_id, username)
            await bot.answer_callback_query(callback_query.id)

        elif data.startswith("detail_"):
            detail_index = int(data.replace("detail_", ""))
            available_details = questionnaire_states[user_id].get('available_details', [])

            if detail_index < len(available_details):
                detail_text = available_details[detail_index]
                details_list = questionnaire_states[user_id].get('details', [])

                # Toggle detail selection
                if detail_text in details_list:
                    details_list.remove(detail_text)
                    await bot.answer_callback_query(callback_query.id, "Убрано из списка")
                else:
                    details_list.append(detail_text)
                    questionnaire_states[user_id]['details'] = details_list
                    await bot.answer_callback_query(callback_query.id, "Добавлено в список")

                # Update the keyboard to show selected items
                await update_details_keyboard(bot, callback_query.message, user_id)

    except Exception as e:
        print(f"Error handling details selection: {e}")


async def update_details_keyboard(bot, message, user_id):
    """Update the details selection keyboard to show selected items"""
    try:
        if user_id not in questionnaire_states:
            return

        selected_details = questionnaire_states[user_id].get('details', [])
        available_details = questionnaire_states[user_id].get('available_details', [])

        markup = types.InlineKeyboardMarkup(row_width=1)

        for i, detail in enumerate(available_details):
            callback = f"detail_{i}"
            # Add checkmark if selected
            if detail in selected_details:
                button_text = f"✅ {detail}"
            else:
                button_text = f"☐ {detail}"
            markup.add(types.InlineKeyboardButton(button_text, callback_data=callback))

        # Add continue button with count
        if selected_details:
            continue_text = f"✅ Готово ({len(selected_details)} выбрано)"
        else:
            continue_text = "✅ Готово"

        markup.add(types.InlineKeyboardButton(continue_text, callback_data="details_continue"))

        # Add menu button
        markup.add(types.InlineKeyboardButton("📱 Главное меню", callback_data="menu:show"))

        # Edit the message with updated keyboard
        await bot.edit_message_reply_markup(
            chat_id=message.chat.id,
            message_id=message.message_id,
            reply_markup=markup
        )

    except Exception as e:
        print(f"Error updating details keyboard: {e}")


def extract_protocol_methods(protocol_name: str) -> List[str]:
    """Extract intervention methods from protocol_and_interventions_map.md"""
    try:
        import os

        map_file = 'protocol_and_interventions_map.md'
        if not os.path.exists(map_file):
            return []

        with open(map_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # Find the protocol section
        lines = content.split('\n')
        protocol_found = False
        methods = []

        for i, line in enumerate(lines):
            # Look for the protocol header
            if protocol_name in line and line.startswith('##'):
                protocol_found = True
                continue

            if protocol_found:
                # Stop at next protocol section
                if line.startswith('##') and i > 0:
                    break

                # Look for "Интервенции:" section
                if 'Интервенции:' in line:
                    # Extract methods after this line
                    for j in range(i + 1, len(lines)):
                        if lines[j].startswith('##'):
                            break
                        if lines[j].strip().startswith('*') or lines[j].strip().startswith(str(len(methods) + 1)):
                            # Clean the method text
                            method = lines[j].strip()
                            method = method.lstrip('*').strip()
                            method = method.lstrip('0123456789').lstrip('️⃣').strip()
                            if method and any(c.isalpha() for c in method):
                                methods.append(method)

        return methods[:5]  # Return first 5 methods

    except Exception as e:
        print(f"Error extracting protocol methods: {e}")
        return []


async def process_questionnaire_results(bot, chat_id, user_id, username):
    """Process questionnaire results and get AI recommendation"""
    try:
        if user_id not in questionnaire_states:
            await bot.send_message(chat_id, "Произошла ошибка. Попробуйте начать заново.")
            return

        state = questionnaire_states[user_id]

        # Prepare data for AI analysis
        symptoms_names = [SYMPTOM_DETAILS.get(s, {}).get('name', s) for s in state['symptoms']]

        # Save to Excel
        save_questionnaire_to_excel(user_id, username, {
            'symptoms': symptoms_names,
            'duration': state['duration'],
            'impact': state['impact'],
            'details': state.get('details', [])
        })

        # Send processing message
        await bot.send_message(chat_id, "Анализирую твои ответы... 🤔")

        # Call AI analysis
        recommendation = await analyze_with_ai(state)

        if recommendation:
            # Store recommendation in user state for confirmation
            questionnaire_states[user_id]['recommendation'] = recommendation

            # Show confirmation message with summary
            await show_situation_summary(bot, chat_id, user_id, recommendation)
        else:
            # Fallback if AI analysis fails
            await bot.send_message(
                chat_id,
                "Спасибо за ответы! Я проанализировал информацию. "
                "Давай выберем подходящий протокол из списка доступных."
            )
            # Import and call protocol selection from protocol_known
            import protocol_known
            await protocol_known.send_protocol_selection(bot, chat_id)

            # Clear user state
            del questionnaire_states[user_id]

    except Exception as e:
        print(f"Error processing questionnaire results: {e}")
        await bot.send_message(chat_id, "Произошла ошибка при обработке результатов.")


async def show_situation_summary(bot, chat_id, user_id, recommendation):
    """Show situation summary and ask for confirmation"""
    try:
        category = recommendation.get('category', 'Не определено')
        duration = recommendation.get('duration', 'Не указано')
        impact = recommendation.get('impact', 'Не определено')
        goal = recommendation.get('goal', 'Не определена')

        summary_text = (
            "🧾 Вот как я вижу твою ситуацию:\n\n"
            f"Основная трудность: *{category}*\n\n"
            f"Длительность: *{duration}*\n\n"
            f"Влияние: *{impact}*\n\n"
            f"Цель: *{goal}*\n\n"
            "Все верно?"
        )

        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("✅ Да, верно", callback_data="confirm_situation"))
        markup.add(types.InlineKeyboardButton("✏️ Изменить", callback_data="change_situation"))

        await bot.send_message(chat_id, summary_text, parse_mode='Markdown', reply_markup=markup)

    except Exception as e:
        print(f"Error showing situation summary: {e}")


async def handle_situation_confirmation(bot, callback_query, user_id, username):
    """Handle confirmation of situation summary"""
    try:
        if user_id not in questionnaire_states:
            await bot.answer_callback_query(
                callback_query.id,
                "Сессия истекла. Пожалуйста, начните заново с /start"
            )
            return

        data = callback_query.data

        if data == "confirm_situation":
            # Show final recommendation with protocol methods
            recommendation = questionnaire_states[user_id].get('recommendation')

            if recommendation:
                protocol_name = recommendation.get('protocol_name', '')
                protocol_id = recommendation.get('protocol_id', '')

                # Extract methods for this protocol
                methods = extract_protocol_methods(protocol_name)

                if methods:
                    methods_text = '\n'.join([f"• {method}" for method in methods])
                else:
                    methods_text = "• Когнитивная реструктуризация\n• Поведенческие эксперименты\n• Техники релаксации"

                final_text = (
                    f"На основе твоих ответов, я рекомендую начать с протокола:\n\n"
                    f"*{protocol_name}*\n\n"
                    f"Этот подход включает в себя техники:\n{methods_text}\n\n"
                    f"Хочешь, я покажу, какие упражнения входят в программу?"
                )

                markup = types.InlineKeyboardMarkup()
                if protocol_id:
                    markup.add(types.InlineKeyboardButton("▶️ Да, покажи", callback_data=f"ps:{protocol_id}"))
                markup.add(types.InlineKeyboardButton("🔙 Хочу выбрать другой", callback_data="show_all_protocols"))

                await bot.answer_callback_query(callback_query.id)
                await bot.send_message(callback_query.message.chat.id, final_text, parse_mode='Markdown', reply_markup=markup)

                # Clear user state
                del questionnaire_states[user_id]

        elif data == "change_situation":
            # Restart questionnaire
            await bot.answer_callback_query(callback_query.id, "Давайте пройдем опрос заново")

            # Get user name from state
            user_name = questionnaire_states[user_id].get('user_name', username)

            # Clear and restart
            del questionnaire_states[user_id]
            await start_questionnaire(bot, callback_query.message.chat.id, user_id, username, user_name)

    except Exception as e:
        print(f"Error handling situation confirmation: {e}")


async def analyze_with_ai(state: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Analyze questionnaire results with AI to recommend a protocol"""
    try:
        from openrouter import OpenRouterClient

        # Import protocol list from protocol_known
        from protocol_known import PROTOCOLS

        # Prepare survey results for prompt
        symptoms_names = [SYMPTOM_DETAILS.get(s, {}).get('name', s) for s in state['symptoms']]

        # Map impact level to text
        impact_levels = {
            0: "не мешает",
            1: "немного мешает",
            2: "умеренно мешает",
            3: "сильно мешает"
        }
        impact_text = impact_levels.get(state['impact'], f"{state['impact']}/3")

        survey_text = f"""
        Main symptoms: {', '.join(symptoms_names)}
        Duration: {state['duration']}
        Impact level: {impact_text} ({state['impact']}/3)
        Specific manifestations: {', '.join(state.get('details', []))}
        """

        # Create list of available protocols with mapping
        protocols_info = []
        for display_name, search_term, protocol_id in PROTOCOLS:
            protocols_info.append(f"- {search_term} (ID: {protocol_id})")
        available_protocols = "\n".join(protocols_info)

        prompt = f"""You are a cognitive behavioral therapy psychologist with 15 years of experience.
        Based on the patient's survey:
        {survey_text}

        Choose the most suitable protocol from this exact list of available protocols:
        {available_protocols}

        IMPORTANT:
        1. You must select ONLY from the protocols in the list above
        2. Return the protocol name EXACTLY as it appears in the list (the Russian text before the ID)
        3. Determine the main difficulty category based on symptoms
        4. Define a clear therapeutic goal based on the selected protocol

        Consider:
        - Primary difficulty category
        - Symptom severity and duration ({state['duration']}, impact: {state['impact']}/3)
        - Impact on daily functioning
        - Specific manifestations

        Provide a complete structured analysis."""

        # Define JSON schema for structured output
        json_schema = {
            "name": "protocol_recommendation",
            "schema": {
                "type": "object",
                "properties": {
                    "category": {
                        "type": "string",
                        "description": "Main difficulty category in Russian (e.g., 'Тревожность', 'Депрессия', 'Стресс')"
                    },
                    "impact": {
                        "type": "string",
                        "description": "Impact level description in Russian (e.g., 'умеренное влияние на жизнь')"
                    },
                    "goal": {
                        "type": "string",
                        "description": "Main therapeutic goal in Russian based on the protocol"
                    },
                    "protocol_name": {
                        "type": "string",
                        "description": "EXACT name of the recommended protocol from the provided list (in Russian, without ID)"
                    },
                    "protocol_id": {
                        "type": "string",
                        "description": "Protocol ID (p1-p18) corresponding to the selected protocol"
                    }
                },
                "required": ["category", "impact", "goal", "protocol_name", "protocol_id"],
                "additionalProperties": False
            }
        }

        client = OpenRouterClient()
        result, usage = client.get_structured_response(
            prompt=prompt,
            json_schema=json_schema,
            system_message="You are an expert CBT psychologist. Always provide ALL responses in Russian. Be specific and therapeutic in your goal formulation."
        )

        # Add duration from state to result
        if result:
            result['duration'] = state['duration']

        print(f"AI Analysis completed. Tokens used: {usage}")
        return result

    except Exception as e:
        print(f"Error in AI analysis: {e}")
        return None


# Handler registration function for main bot
def register_handlers(bot):
    """Register all handlers for protocol_unknown module"""

    @bot.callback_query_handler(func=lambda call: call.data.startswith('symptom_') or call.data == 'symptoms_continue')
    async def symptom_callback(callback_query):
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or str(user_id)
        await handle_symptom_selection(bot, callback_query, user_id, username)

    @bot.callback_query_handler(func=lambda call: call.data.startswith('duration_'))
    async def duration_callback(callback_query):
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or str(user_id)
        await handle_duration_selection(bot, callback_query, user_id, username)

    @bot.callback_query_handler(func=lambda call: call.data.startswith('impact_'))
    async def impact_callback(callback_query):
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or str(user_id)
        await handle_impact_selection(bot, callback_query, user_id, username)

    @bot.callback_query_handler(func=lambda call: call.data.startswith('detail_') or call.data == 'details_continue')
    async def details_callback(callback_query):
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or str(user_id)
        await handle_details_selection(bot, callback_query, user_id, username)

    @bot.callback_query_handler(func=lambda call: call.data == 'show_all_protocols')
    async def show_all_protocols_callback(callback_query):
        # Import and call protocol selection from protocol_known
        import protocol_known
        await bot.answer_callback_query(callback_query.id)
        await protocol_known.send_protocol_selection(bot, callback_query.message.chat.id)

    @bot.callback_query_handler(func=lambda call: call.data in ['confirm_situation', 'change_situation'])
    async def situation_confirmation_callback(callback_query):
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or str(user_id)
        await handle_situation_confirmation(bot, callback_query, user_id, username)