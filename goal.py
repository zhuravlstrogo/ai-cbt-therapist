# -*- coding: utf-8 -*-
"""
Goal setting module for AI Psychologist bot
Handles 3-step goal and problem identification process
"""

import os
import re
from datetime import datetime
from telebot import types
from openpyxl import load_workbook, Workbook

# Path to Excel file for saving progress
EXCEL_FILE = 'messages.xlsx'

# Store user goal-setting states
# Format: {user_id: {'step': int, 'goal': str, 'problems': [str], 'problem_ratings': {str: int}, 'current_problem_idx': int}}
user_goal_states = {}

# List of problems for step 2
# IMPORTANT: These names must match EXACTLY with the headers in protocol_and_interventions_map.md
PROBLEMS = [
    ("😟 Тревога, беспокойство", "anxiety"),
    ("😞 Потеря интереса, апатия / Сниженное настроение", "apathy"),
    ("Пониженное настроение", "mood"),
    ("💤 Проблемы со сном", "sleep"),
    ("⏳ Прокрастинация, снижение сил/мотивации", "procrastination"),
    ("💬 Трудности в общении", "communication"),
    ("💔 Самокритичность, чувство вины", "self_criticism"),
    ("😤 Раздражительность, вспышки гнева", "anger"),
    ("🌀 Навязчивые мысли, действия (ОКР)", "ocd"),
    ("💥 Панические атаки", "panic"),
    ("🎭 Неуверенность в компаниях людей (социальная тревога)", "social_anxiety"),
    ("🎯 Перфекционизм", "perfectionism"),
    ("🌻 Переживание утраты / жизненные перемены", "loss"),
    ("🔄 Стресс, усталость, выгорание", "burnout"),
    ("💡 Хочу укрепить устойчивость", "resilience"),
    ("➕ Другая проблема", "other"),
]


def save_goal_results_to_excel(user_id, username, goal, problems, ratings):
    """Save goal-setting results to Excel file"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Messages'

        # Find next empty row
        next_row = ws.max_row + 1

        # Add goal data
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = f"Goal: {goal}"

        # Format problems with ratings
        problems_with_ratings = []
        for problem in problems:
            rating = ratings.get(problem, 'N/A')
            problems_with_ratings.append(f"{problem} (оценка: {rating})")

        ws[f'D{next_row}'] = f"Problems: {'; '.join(problems_with_ratings)}"
        ws[f'E{next_row}'] = 'goal_setting'
        ws[f'G{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save(EXCEL_FILE)
        print(f"Goal setting saved: {username} - Goal: {goal}, Problems: {len(problems)}")

    except Exception as e:
        print(f"Error saving goal results to Excel: {e}")


async def start_goal_setting(bot, chat_id, user_id, username, skip_goal=False, force_change_goal=False, force_change_problems=False):
    """Start the 3-step goal setting process

    Args:
        skip_goal: If True, skip goal setting and go directly to problem selection
        force_change_goal: If True, force changing goal even if it exists
        force_change_problems: If True, force changing problems
    """
    import traceback
    print(f"DEBUG: start_goal_setting called from:\n{traceback.format_stack()[-2]}")

    try:
        # Check if user already has a goal saved
        from greeting import user_states
        existing_goal = None
        existing_problems = None
        existing_ratings = None

        if user_id in user_states:
            existing_goal = user_states[user_id].get('goal')
            existing_problems = user_states[user_id].get('problems', [])
            existing_ratings = user_states[user_id].get('problem_ratings', {})
            print(f"DEBUG: Loaded existing data - goal: {existing_goal}, problems: {existing_problems}, ratings: {existing_ratings}")

            # If not forcing change, skip goal if it exists
            if not force_change_goal and not force_change_problems and existing_goal:
                skip_goal = True

        # Initialize user state
        initial_step = 1 if not skip_goal else (3 if force_change_problems else 2)

        # When changing only goal, preserve existing problems
        # When changing only problems, preserve existing goal
        user_goal_states[user_id] = {
            'step': initial_step,
            'username': username,
            'goal': '' if force_change_goal else (existing_goal or ''),
            'problems': [] if force_change_problems else (existing_problems or []),
            'problem_ratings': {} if force_change_problems else (existing_ratings or {}),
            'current_problem_idx': 0,
            'is_changing': force_change_goal or force_change_problems,
            'change_type': 'goal' if force_change_goal else ('problems' if force_change_problems else None)
        }

        print(f"DEBUG: User state initialized - problems: {user_goal_states[user_id]['problems']}, change_type: {user_goal_states[user_id]['change_type']}, is_changing: {user_goal_states[user_id]['is_changing']}")

        if force_change_goal:
            # Force change goal - clear it and ask for new one
            user_goal_states[user_id]['step'] = 1
            user_goal_states[user_id]['goal'] = ''
            from universal_menu import get_menu_button
            markup = get_menu_button()

            await bot.send_message(
                chat_id,
                "Какую цель терапии ты перед собой ставишь?",
                reply_markup=markup
            )
        elif force_change_problems:
            # Force change problems - clear them and ask for new selection
            user_goal_states[user_id]['step'] = 2
            user_goal_states[user_id]['problems'] = []
            user_goal_states[user_id]['problem_ratings'] = {}
            await show_problem_selection(bot, chat_id, user_id)
        elif skip_goal or existing_goal:
            # Skip to problem selection if goal already exists
            user_goal_states[user_id]['step'] = 2
            await show_problem_selection(bot, chat_id, user_id)
        else:
            # Step 1: Ask for therapy goal
            from universal_menu import get_menu_button
            markup = get_menu_button()

            await bot.send_message(
                chat_id,
                "Какую цель терапии ты перед собой ставишь?",
                reply_markup=markup
            )

        print(f"Goal setting started for user {username} (ID: {user_id}), skip_goal={skip_goal}, force_change_goal={force_change_goal}, force_change_problems={force_change_problems}")

    except Exception as e:
        print(f"Error starting goal setting: {e}")


async def handle_goal_text_input(bot, message):
    """Handle text input for goal (step 1)"""
    try:
        user_id = message.from_user.id
        username = message.from_user.username or 'Unknown'

        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        if state['step'] != 1:
            return

        # Store goal text
        goal_text = message.text
        state['goal'] = goal_text

        # Check for crisis indicators in goal text
        from safety_check import check_text_safety, show_crisis_support, log_crisis_detection
        from greeting import user_states

        crisis_detected, crisis_type, confidence = await check_text_safety(
            text=goal_text,
            context="goal_setting"
        )

        if crisis_detected and crisis_type:
            # Log crisis detection
            await log_crisis_detection(
                user_id=user_id,
                username=username,
                crisis_type=crisis_type,
                context="goal_setting",
                text_sample=goal_text[:200],
                file_path='messages.xlsx'
            )

            # Get user name
            user_name = 'Друг'
            if user_id in user_states:
                user_name = user_states[user_id].get('user_name', 'Друг')

            # Show crisis support
            await show_crisis_support(
                bot=bot,
                chat_id=message.chat.id,
                user_name=user_name,
                crisis_type=crisis_type,
                context="goal_setting",
                continue_after=True  # Allow continuing with goal setting
            )
            return

        # Show preview with action buttons
        markup = types.InlineKeyboardMarkup()

        btn_confirm = types.InlineKeyboardButton(
            "✅ Подтвердить цель",
            callback_data="goal_confirm:step1"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="goal_edit:step1"
        )
        btn_back = types.InlineKeyboardButton(
            "⬅️ Вернуться",
            callback_data="goal_back:step1"
        )

        markup.add(btn_confirm)
        markup.add(btn_edit)
        markup.add(btn_back)

        preview_text = f"📝 Твоя цель:\n\n{goal_text}"
        await bot.send_message(message.chat.id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling goal text input: {e}")


async def handle_goal_callback(bot, callback_query, action, step):
    """Handle goal setting callbacks"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_goal_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_goal_states[user_id]

        if action == "confirm" and step == "step1":
            # Check if this is a change-only-goal operation
            print(f"DEBUG: Confirm goal - is_changing: {state.get('is_changing')}, change_type: {state.get('change_type')}, problems: {state.get('problems')}")
            if state.get('is_changing') and state.get('change_type') == 'goal':
                # User is only changing goal
                # Just save the goal and return to menu (regardless of existing problems)
                print("DEBUG: Only changing goal, saving and returning to menu")
                await bot.answer_callback_query(callback_query.id, "Цель сохранена!")
                
                # Save only the goal
                from greeting import user_states
                if user_id not in user_states:
                    user_states[user_id] = {}
                
                user_states[user_id]['goal'] = state['goal']
                
                # Show confirmation and return to main menu
                from universal_menu import show_main_menu
                
                user_name = 'User'
                form_of_address = 'ты'
                if user_id in user_states:
                    user_name = user_states[user_id].get('user_name', 'User')
                    form_of_address = user_states[user_id].get('form', 'ты')
                
                # Show confirmation with menu button
                from universal_menu import get_menu_button
                markup = get_menu_button()
                
                await bot.send_message(
                    chat_id,
                    f"✅ Цель сохранена: {state['goal']}",
                    reply_markup=markup
                )
                
                # Clean up state
                del user_goal_states[user_id]
            else:
                # Normal flow - move to step 2: problem selection
                state['step'] = 2
                await bot.answer_callback_query(callback_query.id)
                await show_problem_selection(bot, chat_id, user_id)

        elif action == "edit" and step == "step1":
            # Ask for new goal
            state['goal'] = ''
            await bot.answer_callback_query(callback_query.id)
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(chat_id, "Введи новую цель:", reply_markup=markup)

        elif action == "back" and step == "step1":
            # Return to previous screen
            await bot.answer_callback_query(callback_query.id)
            # Show main menu
            from universal_menu import show_main_menu
            from greeting import user_states

            user_name = 'User'
            form_of_address = 'ты'
            if user_id in user_states:
                user_name = user_states[user_id].get('user_name', 'User')
                form_of_address = user_states[user_id].get('form', 'ты')

            # Clean up state
            del user_goal_states[user_id]
            await show_main_menu(bot, chat_id, user_id, username, user_name, form_of_address)
        
        elif action == "continue" and step == "after_safety":
            # Continue goal setting after safety check
            await bot.answer_callback_query(callback_query.id)
            
            # Show preview with action buttons (continue from where we left off)
            markup = types.InlineKeyboardMarkup()

            btn_confirm = types.InlineKeyboardButton(
                "✅ Подтвердить цель",
                callback_data="goal_confirm:step1"
            )
            btn_edit = types.InlineKeyboardButton(
                "✏️ Изменить",
                callback_data="goal_edit:step1"
            )
            btn_back = types.InlineKeyboardButton(
                "⬅️ Вернуться",
                callback_data="goal_back:step1"
            )

            markup.add(btn_confirm)
            markup.add(btn_edit)
            markup.add(btn_back)

            preview_text = f"📝 Твоя цель:\n\n{state['goal']}"
            await bot.send_message(chat_id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling goal callback: {e}")
        await bot.answer_callback_query(callback_query.id)


async def show_problem_selection(bot, chat_id, user_id):
    """Show step 2: problem selection with multiple choice buttons"""
    try:
        # Check if this is a change operation - if so, clear problem ratings
        if user_id in user_goal_states:
            state = user_goal_states[user_id]
            if state.get('is_changing'):
                state['problem_ratings'] = {}

        text = "Выбери проблемы, над которыми хочешь работать (можно несколько):"

        markup = types.InlineKeyboardMarkup()

        # Add problem buttons one per row for clarity
        for display_name, problem_id in PROBLEMS:
            btn = types.InlineKeyboardButton(
                display_name,
                callback_data=f"prob_select:{problem_id}"
            )
            markup.add(btn)

        # Add continue button
        btn_continue = types.InlineKeyboardButton(
            "➡️ Продолжить",
            callback_data="prob_done:proceed"
        )
        markup.add(btn_continue)

        # Add menu button for accessibility
        btn_menu = types.InlineKeyboardButton(
            "📱 Главное меню",
            callback_data="menu:show"
        )
        markup.add(btn_menu)

        await bot.send_message(chat_id, text, reply_markup=markup)

    except Exception as e:
        print(f"Error showing problem selection: {e}")


async def handle_problem_selection(bot, callback_query, problem_id):
    """Handle problem selection toggle"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        # Answer callback IMMEDIATELY to avoid timeout
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        if state['step'] != 2:
            return

        # Check if user selected "other" problem
        if problem_id == "other":
            # Import and start other problem flow
            from other_problem import start_other_problem_flow
            success = await start_other_problem_flow(bot, chat_id, user_id, username)
            if success:
                # Hide current problem selection interface temporarily
                await bot.edit_message_text(
                    "Переходим к описанию проблемы...",
                    chat_id,
                    callback_query.message.message_id
                )
            return

        # Find the problem display name
        problem_display = None
        for display_name, p_id in PROBLEMS:
            if p_id == problem_id:
                problem_display = display_name
                break

        if problem_display is None:
            return

        # Toggle selection
        if problem_display in state['problems']:
            state['problems'].remove(problem_display)
        else:
            state['problems'].append(problem_display)

        # Update the message to show selected problems with checkmarks
        text = "Выбери проблемы, над которыми хочешь работать (можно несколько):\n\n"
        if state['problems']:
            text += "✅ Выбрано:\n"
            for prob in state['problems']:
                text += f"• {prob}\n"

        markup = types.InlineKeyboardMarkup()

        # Add problem buttons with checkmarks for selected ones
        for display_name, p_id in PROBLEMS:
            if display_name in state['problems']:
                btn_text = f"✅ {display_name}"
            else:
                btn_text = display_name
            btn = types.InlineKeyboardButton(
                btn_text,
                callback_data=f"prob_select:{p_id}"
            )
            markup.add(btn)

        # Add continue button
        btn_continue = types.InlineKeyboardButton(
            "➡️ Продолжить",
            callback_data="prob_done:proceed"
        )
        markup.add(btn_continue)

        # Add menu button
        btn_menu = types.InlineKeyboardButton(
            "📱 Главное меню",
            callback_data="menu:show"
        )
        markup.add(btn_menu)

        # Update the message with new markup
        await bot.edit_message_text(
            text,
            chat_id,
            callback_query.message.message_id,
            reply_markup=markup
        )

    except Exception as e:
        print(f"Error handling problem selection: {e}")


async def handle_problems_done(bot, callback_query):
    """Handle move from step 2 to step 3"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        # Answer callback IMMEDIATELY to avoid timeout
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        if state['step'] != 2:
            return

        # Check if at least one problem selected (but allow skipping)
        if not state['problems']:
            # Allow proceeding without selection
            state['step'] = 4  # Skip to completion
            await finish_goal_setting(bot, chat_id, user_id, username)
            return

        # Move to step 3: rating problems
        state['step'] = 3
        state['current_problem_idx'] = 0

        # Clear ratings for any problems that are no longer selected
        # (in case of change operation where some problems were deselected)
        problems_to_remove = [p for p in state['problem_ratings'].keys() if p not in state['problems']]
        for problem in problems_to_remove:
            del state['problem_ratings'][problem]

        # Send step 3 header
        from universal_menu import get_menu_button
        markup = get_menu_button()

        await bot.send_message(
            chat_id,
            "Теперь давай оценим, насколько каждая из выбранных трудностей влияет на твою жизнь.",
            reply_markup=markup
        )

        # Show first problem for rating
        await show_problem_rating(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling problems done: {e}")


async def show_problem_rating(bot, chat_id, user_id):
    """Show current problem for rating (step 3)"""
    try:
        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        if state['step'] != 3:
            return

        if state['current_problem_idx'] >= len(state['problems']):
            # All problems rated - show final preview
            username = state.get('username', 'Unknown')
            await show_final_preview(bot, chat_id, user_id, username)
            return

        current_problem = state['problems'][state['current_problem_idx']]

        # Create rating buttons
        markup = types.InlineKeyboardMarkup()

        # Rating buttons (0-3)
        btn_0 = types.InlineKeyboardButton("0️⃣ не мешает", callback_data=f"rate:{state['current_problem_idx']}:0")
        btn_1 = types.InlineKeyboardButton("1️⃣ немного", callback_data=f"rate:{state['current_problem_idx']}:1")
        btn_2 = types.InlineKeyboardButton("2️⃣ заметно", callback_data=f"rate:{state['current_problem_idx']}:2")
        btn_3 = types.InlineKeyboardButton("3️⃣ сильно", callback_data=f"rate:{state['current_problem_idx']}:3")

        markup.row(btn_0, btn_1, btn_2, btn_3)

        # Back button
        btn_back = types.InlineKeyboardButton("⬅️ Назад", callback_data=f"rate_back:{state['current_problem_idx']}")
        markup.add(btn_back)

        rating_text = f"Проблема {state['current_problem_idx'] + 1} из {len(state['problems'])}:\n\n{current_problem}\n\nКак сильно это влияет на твою жизнь?\n\n0 — не мешает · 1 — немного · 2 — заметно · 3 — сильно мешает"

        await bot.send_message(chat_id, rating_text, reply_markup=markup)

    except Exception as e:
        print(f"Error showing problem rating: {e}")


async def handle_problem_rating(bot, callback_query, problem_idx, rating):
    """Handle problem rating"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        # Answer callback IMMEDIATELY to avoid timeout
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        if state['step'] != 3 or state['current_problem_idx'] != int(problem_idx):
            return

        problem = state['problems'][int(problem_idx)]
        rating_value = int(rating)

        # Store rating
        state['problem_ratings'][problem] = rating_value

        # Move to next problem
        state['current_problem_idx'] += 1

        # Check if all problems are rated
        if state['current_problem_idx'] >= len(state['problems']):
            # All problems rated - show final preview
            await show_final_preview(bot, chat_id, user_id, username)
        else:
            # Show next problem
            await show_problem_rating(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling problem rating: {e}")


async def handle_rating_back(bot, callback_query, problem_idx):
    """Handle back button during rating"""
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        # Answer callback IMMEDIATELY to avoid timeout
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        problem_idx = int(problem_idx)

        if problem_idx == 0:
            # Go back to step 2 (problem selection)
            state['step'] = 2
            state['problems'] = []
            state['problem_ratings'] = {}

            await show_problem_selection(bot, chat_id, user_id)
        else:
            # Go to previous problem for re-rating
            state['current_problem_idx'] = problem_idx - 1

            await show_problem_rating(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling rating back: {e}")


async def show_final_preview(bot, chat_id, user_id, username):
    """Show final preview with goal, problems and ratings"""
    try:
        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        # Get user name from greeting state
        from greeting import user_states
        user_name = 'друг'
        if user_id in user_states:
            user_name = user_states[user_id].get('user_name', 'друг')

        # Format problems with ratings
        problems_text = []
        for problem in state['problems']:
            rating = state['problem_ratings'].get(problem, 'N/A')
            problems_text.append(f"• {problem}: {rating}")

        problems_list = "\n".join(problems_text) if problems_text else "Проблемы не выбраны"

        # Customize message based on what was changed
        change_type = state.get('change_type')
        if change_type == 'goal':
            # Only goal was changed
            preview_message = (
                f"🎯 Обновлённая цель терапии:\n\n"
                f"{state['goal']}\n\n"
                f"Твои текущие трудности:\n{problems_list}\n\n"
                f"Всё верно?"
            )
        elif change_type == 'problems':
            # Only problems were changed
            preview_message = (
                f"🧭 Обновлённые трудности и их оценка:\n{problems_list}\n\n"
                f"Твоя цель терапии:\n{state['goal']}\n\n"
                f"Всё верно?"
            )
        else:
            # Normal flow or both changed
            preview_message = (
                f"🧾 Вот как я вижу твою ситуацию, {user_name}:\n\n"
                f"Трудности и их оценка:\n{problems_list}\n\n"
                f"Цель терапии: {state['goal']}\n\n"
                f"Всё верно?"
            )

        # Create buttons
        markup = types.InlineKeyboardMarkup()

        btn_yes = types.InlineKeyboardButton(
            "✅ Подтвердить",
            callback_data="preview_confirm:yes"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="preview_edit:choose"
        )

        markup.add(btn_yes)
        markup.add(btn_edit)

        await bot.send_message(chat_id, preview_message, reply_markup=markup)

    except Exception as e:
        print(f"Error showing final preview: {e}")


async def handle_preview_confirm(bot, callback_query, action):
    """Handle final preview confirmation or edit"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_goal_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_goal_states[user_id]

        if action == "yes":
            # All confirmed - start exercise recommendations
            await bot.answer_callback_query(callback_query.id, "Спасибо! Данные сохранены.")

            # Save goal and problems to user_states for future use
            from greeting import user_states
            if user_id not in user_states:
                user_states[user_id] = {}

            user_states[user_id]['goal'] = state['goal']
            user_states[user_id]['problems'] = state['problems']
            user_states[user_id]['problem_ratings'] = state['problem_ratings']

            # Import exercise module
            from exercise import show_exercise_recommendations

            # Show exercise recommendations
            await show_exercise_recommendations(bot, chat_id, user_id, username, state['problem_ratings'])

        elif action == "choose":
            # Ask what to change
            await bot.answer_callback_query(callback_query.id, show_alert=False)
            change_markup = types.InlineKeyboardMarkup()

            btn_goal = types.InlineKeyboardButton(
                "🎯 Изменить цель",
                callback_data="preview_change:goal"
            )
            btn_problems = types.InlineKeyboardButton(
                "📋 Изменить проблемы",
                callback_data="preview_change:problems"
            )

            change_markup.add(btn_goal)
            change_markup.add(btn_problems)

            change_text = "Что ты хочешь изменить?"
            await bot.send_message(chat_id, change_text, reply_markup=change_markup)

    except Exception as e:
        print(f"Error handling preview confirm: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_preview_change(bot, callback_query, change_type):
    """Handle what to change in preview"""
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        # Answer callback IMMEDIATELY to avoid timeout
        await bot.answer_callback_query(callback_query.id, show_alert=False)

        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        # Preserve the original change intent
        original_change_type = state.get('change_type')

        if change_type == "goal":
            # Go back to step 1 - ask for new goal
            state['step'] = 1
            state['goal'] = ''
            # If originally changing only problems, now changing both
            if original_change_type == 'problems':
                state['change_type'] = None  # Now changing both
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(chat_id, "Введи новую цель терапии:", reply_markup=markup)

        elif change_type == "problems":
            # Go back to step 2 - select problems again
            state['step'] = 2
            state['problems'] = []
            state['problem_ratings'] = {}
            # If originally changing only goal, now changing both
            if original_change_type == 'goal':
                state['change_type'] = None  # Now changing both
            await show_problem_selection(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling preview change: {e}")


async def finish_goal_setting(bot, chat_id, user_id, username):
    """Finish goal setting process and save results"""
    try:
        if user_id not in user_goal_states:
            return

        state = user_goal_states[user_id]

        # Save to Excel
        save_goal_results_to_excel(
            user_id,
            username,
            state['goal'],
            state['problems'],
            state['problem_ratings']
        )

        # Save goal and problems to user_states for future use
        from greeting import user_states
        if user_id not in user_states:
            user_states[user_id] = {}

        user_states[user_id]['goal'] = state['goal']
        user_states[user_id]['problems'] = state['problems']
        user_states[user_id]['problem_ratings'] = state['problem_ratings']

        # Show completion message
        completion_text = "✅ Спасибо! Я записал твои ответы.\n\n"
        completion_text += f"🎯 Твоя цель: {state['goal']}\n\n"

        if state['problems']:
            completion_text += "📋 Проблемы, над которыми будем работать:\n"
            for problem in state['problems']:
                rating = state['problem_ratings'].get(problem, 'N/A')
                completion_text += f"  • {problem} (оценка: {rating})\n"
        else:
            completion_text += "Ты выбрал не выбирать конкретные проблемы.\n"

        await bot.send_message(chat_id, completion_text)

        # Return to main menu
        from universal_menu import show_main_menu
        from greeting import user_states

        user_name = 'User'
        form_of_address = 'ты'
        if user_id in user_states:
            user_name = user_states[user_id].get('user_name', 'User')
            form_of_address = user_states[user_id].get('form', 'ты')

        # Clean up state
        del user_goal_states[user_id]

        await show_main_menu(bot, chat_id, user_id, username, user_name, form_of_address)

    except Exception as e:
        print(f"Error finishing goal setting: {e}")
