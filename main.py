import asyncio
import os
import io
from datetime import datetime
from telebot.async_telebot import AsyncTeleBot
from telebot import types
import speech_recognition as sr
from pydub import AudioSegment
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from greeting import (
    send_greeting_messages,
    handle_name_input,
    handle_form_of_address_choice,
    handle_ready_to_start,
    reset_user_greeting_state,
    user_states,
    update_excel_headers
)
import goal
import universal_menu
from diary import init_diary_file, handle_diary_entry

# Load environment variables
load_dotenv()

# Initialize bot
BOT_TOKEN = os.getenv('BOT_TOKEN')
if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN not found in environment variables. Please create a .env file with BOT_TOKEN=your_token")
bot = AsyncTeleBot(BOT_TOKEN)

# Initialize speech recognizer
recognizer = sr.Recognizer()

# Excel file path
EXCEL_FILE = 'messages.xlsx'


def init_excel_file():
    """Initialize Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Messages'
        ws['A1'] = 'User ID'
        ws['B1'] = 'Username'
        ws['C1'] = 'User Name'
        ws['D1'] = 'Message Text'
        ws['E1'] = 'Message Type'
        ws['F1'] = 'Form of Address'  # 'ты' or 'Вы'
        ws['G1'] = 'Protocol Choice'
        ws['H1'] = 'Date Time'
        wb.save(EXCEL_FILE)
    else:
        # Update headers if file exists but doesn't have new columns
        update_excel_headers()


def save_message_to_excel(username, text, user_id=None, message_type='user_message'):
    """Save message to Excel file"""
    try:
        # Load existing workbook or create new one
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            init_excel_file()
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        # Add message data with new columns
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'D{next_row}'] = text
        ws[f'E{next_row}'] = message_type
        ws[f'H{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"Message saved to Excel: {username} - {text[:50]}...")
    except Exception as e:
        print(f"Error saving message to Excel: {e}")


async def process_voice_message(message):
    """Process voice message and return transcribed text"""
    try:
        # Get voice file
        file_info = await bot.get_file(message.voice.file_id)
        voice_file = await bot.download_file(file_info.file_path)

        # Convert voice to wav format
        audio = AudioSegment.from_ogg(io.BytesIO(voice_file))
        wav_data = io.BytesIO()
        audio.export(wav_data, format="wav")
        wav_data.seek(0)

        # Recognize speech
        with sr.AudioFile(wav_data) as source:
            audio_data = recognizer.record(source)
            text = recognizer.recognize_google(audio_data, language='ru-RU')
            return text
    except Exception as e:
        print(f"Error processing voice message: {e}")
        return ""


@bot.message_handler(commands=['start'])
async def start(message):
    """Handle /start command - initiate greeting sequence"""
    user_id = message.from_user.id
    username = message.from_user.username or 'Unknown'

    # Reset user state if they were in the middle of greeting process
    reset_user_greeting_state(user_id)

    # Send greeting messages
    await send_greeting_messages(bot, message.chat.id, user_id, username)


@bot.message_handler(commands=['menu'])
async def menu_command(message):
    """Handle /menu command - show main menu"""
    user_id = message.from_user.id
    username = message.from_user.username or 'Unknown'

    # Get user name and form of address from greeting state or use default
    from greeting import user_states
    user_name = 'Друг'
    form_of_address = 'ты'
    if user_id in user_states:
        user_name = user_states[user_id].get('user_name', 'Друг')
        form_of_address = user_states[user_id].get('form', 'ты')

    # Show universal menu
    await universal_menu.show_main_menu(bot, message.chat.id, user_id, username, user_name, form_of_address)



@bot.message_handler(content_types=['text'])
async def handle_text(message):
    """Handle text messages"""
    text = message.text
    user_id = message.from_user.id
    username = message.from_user.username or 'Unknown'

    # Check if user is in greeting process (awaiting name input) - MUST BE FIRST
    if user_id in user_states and user_states[user_id].get('stage') == 'awaiting_name':
        # Handle name input for greeting
        success = await handle_name_input(bot, message, user_id, username)
        if success:
            return

    # Check if user is in diary entry mode
    from diary import user_diary_states
    if user_id in user_diary_states and user_diary_states[user_id].get('stage') == 'awaiting_text':
        # Handle diary entry
        await handle_diary_entry(bot, message)
        return

    # Check if user is in "other problem" flow
    from other_problem import user_other_problem_states, handle_other_problem_text
    if user_id in user_other_problem_states:
        handled = await handle_other_problem_text(bot, message)
        if handled:
            return

    # Check if user is in check-in process (steps 1-2)
    from check_in import user_checkin_states, handle_checkin_text_input
    if user_id in user_checkin_states:
        state = user_checkin_states[user_id]
        if state.get('step') in [1, 2]:
            # Handle check-in text input
            await handle_checkin_text_input(bot, message)
            return

    # Check if user is in goal setting (step 1) - only if not awaiting name
    from goal import user_goal_states
    if user_id in user_goal_states and user_goal_states[user_id].get('step') == 1:
        # Make sure we're not in the middle of greeting flow
        if user_id not in user_states or user_states[user_id].get('stage') != 'awaiting_name':
            print(f"DEBUG: Handling goal text input for user {username}, text: {text}")
            # Handle goal text input
            await goal.handle_goal_text_input(bot, message)
            return
        else:
            print(f"DEBUG: Skipping goal handling - user {username} is awaiting name input")

    # Check if user is in exercise execution mode
    from exercise import user_exercise_states
    if user_id in user_exercise_states:
        state = user_exercise_states[user_id]
        if state.get('awaiting_exercise_text') or state.get('awaiting_step_input') or state.get('awaiting_final_answer'):
            # Handle exercise/step/answer text input
            import exercise
            await exercise.handle_exercise_text_input(bot, message)
            return

    # Check if user is in mindfulness practice mode
    from mvst import user_mvst_states
    if user_id in user_mvst_states:
        state = user_mvst_states[user_id]
        if state.get('awaiting_practice_input') or state.get('awaiting_final_answer'):
            # Handle mindfulness practice/answer text input
            import mvst
            await mvst.handle_practice_text_input(bot, message)
            return

    # Regular message handling
    print(f"Text message from {username}: {text}")
    save_message_to_excel(username, text, user_id)

    # Add menu button for accessibility
    from universal_menu import get_menu_button
    markup = get_menu_button()

    await bot.send_message(message.chat.id, f"Получено текстовое сообщение: {text}", reply_markup=markup)


@bot.message_handler(content_types=['voice'])
async def handle_voice(message):
    """Handle voice messages"""
    try:
        transcribed_text = await process_voice_message(message)
        username = message.from_user.username or 'Unknown'

        from universal_menu import get_menu_button
        markup = get_menu_button()

        if transcribed_text:
            print(f"Voice message from {username} transcribed to: {transcribed_text}")
            save_message_to_excel(username, transcribed_text, message.from_user.id, 'voice_message')
            await bot.send_message(message.chat.id, f"Распознано голосовое сообщение: {transcribed_text}", reply_markup=markup)
        else:
            print(f"Error transcribing voice message from {username}")
            await bot.send_message(message.chat.id, "Ошибка при распознавании голосового сообщения", reply_markup=markup)
    except Exception as e:
        print(f"Error handling voice message: {e}")
        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(message.chat.id, "Произошла ошибка при обработке голосового сообщения", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith('form_address:'))
async def handle_form_of_address_selection(call):
    """Handle form of address (ты/Вы) button clicks"""
    user_id = call.from_user.id
    username = call.from_user.username or 'Unknown'

    # Process the form of address choice
    await handle_form_of_address_choice(bot, call, user_id, username)


@bot.callback_query_handler(func=lambda call: call.data == 'ready_to_start')
async def handle_ready_start(call):
    """Handle 'Ready to start?' button click"""
    user_id = call.from_user.id
    username = call.from_user.username or 'Unknown'

    # Process the ready to start click
    await handle_ready_to_start(bot, call, user_id, username)


@bot.callback_query_handler(func=lambda call: call.data.startswith('ps:'))
async def handle_specific_protocol_selection(call):
    """Handle specific protocol selection from the list"""
    user_id = call.from_user.id
    username = call.from_user.username or 'Unknown'

    # Extract protocol ID from callback data (ps:p1, ps:p2, etc.)
    protocol_id = call.data.replace('ps:', '')

    # Process the selected protocol
    await goal.handle_protocol_selection(bot, call, protocol_id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('ex_start:'))
async def handle_exercise_start(call):
    """Handle exercise start button click"""
    # Parse callback data: ex_start:protocol_id:exercise_index
    parts = call.data.split(':')
    if len(parts) == 3:
        protocol_id = parts[1]
        exercise_index = parts[2]
        await goal.handle_exercise_start(bot, call, protocol_id, exercise_index)


@bot.callback_query_handler(func=lambda call: call.data.startswith('ex_skip:'))
async def handle_exercise_skip(call):
    """Handle exercise skip button click"""
    # Parse callback data: ex_skip:protocol_id:exercise_index
    parts = call.data.split(':')
    if len(parts) == 3:
        protocol_id = parts[1]
        exercise_index = parts[2]
        await goal.handle_exercise_skip(bot, call, protocol_id, exercise_index)


# Goal setting callbacks (new 3-step process)
@bot.callback_query_handler(func=lambda call: call.data.startswith('goal_confirm:') or call.data.startswith('goal_edit:') or call.data.startswith('goal_back:'))
async def handle_goal_step1_actions(call):
    """Handle goal step 1 actions (confirm, edit, back)"""
    # Parse callback data: goal_confirm:step1, goal_edit:step1, goal_back:step1
    parts = call.data.split(':')
    if len(parts) == 2:
        action = parts[0].replace('goal_', '')
        step = parts[1]
        await goal.handle_goal_callback(bot, call, action, step)


@bot.callback_query_handler(func=lambda call: call.data.startswith('prob_select:'))
async def handle_problem_select(call):
    """Handle problem selection toggle"""
    # Parse callback data: prob_select:problem_id
    problem_id = call.data.replace('prob_select:', '')
    await goal.handle_problem_selection(bot, call, problem_id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('prob_done:'))
async def handle_problems_continue(call):
    """Handle move from problem selection to rating"""
    await goal.handle_problems_done(bot, call)


@bot.callback_query_handler(func=lambda call: call.data.startswith('rate:'))
async def handle_problem_rate(call):
    """Handle problem rating (0-3)"""
    # Parse callback data: rate:problem_idx:rating
    parts = call.data.split(':')
    if len(parts) == 3:
        problem_idx = parts[1]
        rating = parts[2]
        await goal.handle_problem_rating(bot, call, problem_idx, rating)


@bot.callback_query_handler(func=lambda call: call.data.startswith('rate_back:'))
async def handle_rating_back(call):
    """Handle back button during rating"""
    # Parse callback data: rate_back:problem_idx
    problem_idx = call.data.replace('rate_back:', '')
    await goal.handle_rating_back(bot, call, problem_idx)


@bot.callback_query_handler(func=lambda call: call.data.startswith('preview_confirm:') or call.data.startswith('preview_edit:'))
async def handle_preview_confirm(call):
    """Handle final preview confirmation"""
    # Parse callback data: preview_confirm:yes or preview_edit:choose
    parts = call.data.split(':')
    if len(parts) == 2:
        if call.data.startswith('preview_confirm:'):
            action = parts[1]
        else:
            action = parts[1]
        await goal.handle_preview_confirm(bot, call, action)


@bot.callback_query_handler(func=lambda call: call.data.startswith('preview_change:'))
async def handle_preview_change(call):
    """Handle what to change in preview"""
    # Parse callback data: preview_change:goal or preview_change:problems
    change_type = call.data.replace('preview_change:', '')
    await goal.handle_preview_change(bot, call, change_type)


# Exercise callbacks
@bot.callback_query_handler(func=lambda call: call.data.startswith('ex_select:'))
async def handle_exercise_select(call):
    """Handle exercise selection"""
    # Parse callback data: ex_select:idx
    exercise_idx = call.data.replace('ex_select:', '')
    import exercise
    await exercise.handle_exercise_select(bot, call, exercise_idx)


@bot.callback_query_handler(func=lambda call: call.data == 'ex_start_exec')
async def handle_exercise_start(call):
    """Handle exercise start"""
    import exercise
    await exercise.handle_exercise_start(bot, call)


@bot.callback_query_handler(func=lambda call: call.data == 'ex_change_select')
async def handle_exercise_change(call):
    """Handle exercise selection change"""
    import exercise
    await exercise.handle_exercise_change_select(bot, call)


@bot.callback_query_handler(func=lambda call: call.data.startswith('ex_text_confirm:'))
async def handle_exercise_text_confirm(call):
    """Handle exercise text confirmation"""
    action = call.data.replace('ex_text_confirm:', '')
    import exercise
    await exercise.handle_exercise_text_confirm(bot, call, action)


@bot.callback_query_handler(func=lambda call: call.data.startswith('ex_step_confirm:'))
async def handle_step_confirm(call):
    """Handle step confirmation during exercise execution"""
    action = call.data.replace('ex_step_confirm:', '')
    import exercise
    await exercise.handle_step_confirm(bot, call, action)


@bot.callback_query_handler(func=lambda call: call.data.startswith('ex_answer_confirm:'))
async def handle_answer_confirm(call):
    """Handle final answer confirmation"""
    action = call.data.replace('ex_answer_confirm:', '')
    import exercise
    await exercise.handle_answer_confirm(bot, call, action)


@bot.callback_query_handler(func=lambda call: call.data == 'ex_mark_complete')
async def handle_mark_exercise_complete(call):
    """Handle marking exercise as completed"""
    import exercise
    await exercise.handle_mark_exercise_complete(bot, call)


# Other problem callbacks
@bot.callback_query_handler(func=lambda call: call.data.startswith('other_suggest:'))
async def handle_other_suggest(call):
    """Handle selection of suggested problem from LLM"""
    problem_id = call.data.replace('other_suggest:', '')
    from other_problem import handle_other_problem_callback
    await handle_other_problem_callback(bot, call, 'other_suggest', problem_id)


@bot.callback_query_handler(func=lambda call: call.data.startswith('other_custom:'))
async def handle_other_custom(call):
    """Handle custom problem name request"""
    data = call.data.replace('other_custom:', '')
    from other_problem import handle_other_problem_callback
    await handle_other_problem_callback(bot, call, 'other_custom', data)


@bot.callback_query_handler(func=lambda call: call.data.startswith('other_another:'))
async def handle_other_another(call):
    """Handle request to add another problem"""
    data = call.data.replace('other_another:', '')
    from other_problem import handle_other_problem_callback
    await handle_other_problem_callback(bot, call, 'other_another', data)


@bot.callback_query_handler(func=lambda call: call.data.startswith('other_done:'))
async def handle_other_done(call):
    """Handle completion of other problem flow"""
    data = call.data.replace('other_done:', '')
    from other_problem import handle_other_problem_callback
    await handle_other_problem_callback(bot, call, 'other_done', data)


@bot.callback_query_handler(func=lambda call: call.data.startswith('other_confirm_selected:'))
async def handle_other_confirm_selected(call):
    """Handle confirmation of selected problem suggestions"""
    data = call.data.replace('other_confirm_selected:', '')
    from other_problem import handle_other_problem_callback
    await handle_other_problem_callback(bot, call, 'other_confirm_selected', data)


# Diary callbacks
@bot.callback_query_handler(func=lambda call: call.data.startswith('diary:'))
async def handle_diary_callback(call):
    """Handle diary button clicks"""
    action = call.data.replace('diary:', '')
    from diary import handle_diary_confirm, handle_diary_edit, handle_diary_back

    if action == 'confirm':
        await handle_diary_confirm(bot, call)
    elif action == 'edit':
        await handle_diary_edit(bot, call)
    elif action == 'back':
        await handle_diary_back(bot, call)


async def main():
    """Main function to run the bot"""
    print("Starting bot in polling mode...")
    init_excel_file()
    init_diary_file()

    # Initialize MVST Excel file
    from mvst import init_mvst_excel
    init_mvst_excel()

    # Register universal menu handlers
    universal_menu.register_menu_handlers(bot)

    # Register MVST (mindfulness) handlers
    from mvst import register_mvst_handlers
    register_mvst_handlers(bot)

    # Register check-in handlers
    from check_in import register_checkin_handlers, schedule_weekly_checkins
    register_checkin_handlers(bot)

    # Initialize weekly check-in scheduler
    await schedule_weekly_checkins(bot)

    # Register safety handlers
    from safety_check import register_safety_handlers
    register_safety_handlers(bot)

    await bot.infinity_polling()


if __name__ == '__main__':
    asyncio.run(main())
