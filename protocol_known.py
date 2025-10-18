# -*- coding: utf-8 -*-
"""
Protocol selection module for AI Psychologist bot
Handles protocol selection when user knows which protocol they need
"""

import os
import re
from datetime import datetime
from telebot import types
from openpyxl import load_workbook, Workbook

# Path to the protocols and interventions map file
PROTOCOL_MAP_FILE = 'protocol_and_interventions_map.md'
INTERVENTIONS_FILE = 'interventions.md'
EXCEL_FILE = 'messages.xlsx'

# Store user exercise states
# Format: {user_id: {'protocol_id': str, 'protocol_name': str, 'exercises': [str], 'current_index': int}}
user_exercise_states = {}

# List of all 18 protocols with their display names and search terms
# Format: (display_name, search_term, short_id)
PROTOCOLS = [
    ("1️⃣ Управление тревогой", "Тревожность / генерализованное тревожное расстройство", "p1"),
    ("2️⃣ Настроение", "Депрессия / сниженное настроение", "p2"),
    ("3️⃣ Качество сна", "Бессонница", "p3"),
    ("4️⃣ Прокрастинация и мотивация", "Прокрастинация / низкая мотивация", "p4"),
    ("5️⃣ Коммуникация и уверенность", "Коммуникация и уверенность / трудности в отношениях", "p5"),
    ("6️⃣ Самооценка", "Самооценка и внутренний критик", "p6"),
    ("7️⃣ Контроль эмоций", "Контроль эмоций / раздражительность / гнев", "p7"),
    ("8️⃣ Отношения и границы", "Зависимость от одобрения / трудности с границами", "p8"),
    ("9️⃣ Навязчивости (ОКР)", "Обсессивно-компульсивное расстройство", "p9"),
    ("🔟 Панические атаки", "Паническое расстройство / панические атаки", "p10"),
    ("1️⃣1️⃣ Социальная тревожность", "Социальная тревожность", "p11"),
    ("1️⃣2️⃣ Травмирующий опыт", "Посттравматическое стрессовое расстройство", "p12"),
    ("1️⃣3️⃣ Пищевое поведение", "Расстройства пищевого поведения / образ тела", "p13"),
    ("1️⃣4️⃣ Психосоматика", "Психосоматика / тревога о здоровье", "p14"),
    ("1️⃣5️⃣ Перфекционизм", "Перфекционизм", "p15"),
    ("1️⃣6️⃣ Адаптация к изменениям", "Утрата / адаптация к переменам", "p16"),
    ("1️⃣7️⃣ Стресс и выгорание", "Стресс / выгорание", "p17"),
    ("1️⃣8️⃣ Поддержание устойчивости", "Профилактика рецидивов / устойчивость", "p18"),
]

# Create a mapping from short_id to search_term for quick lookup
PROTOCOL_ID_MAP = {item[2]: item[1] for item in PROTOCOLS}


def extract_protocol_data(protocol_search_term):
    """
    Search for protocol in protocol_and_interventions_map.md and extract goal and exercises.

    Args:
        protocol_search_term (str): The protocol name to search for (e.g., "Перфекционизм")

    Returns:
        dict: {'goal': str, 'exercises': [str]} or None if not found
    """
    try:
        if not os.path.exists(PROTOCOL_MAP_FILE):
            print(f"Error: {PROTOCOL_MAP_FILE} not found")
            return None

        with open(PROTOCOL_MAP_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        # Find the section containing the protocol
        # Look for pattern like "## ... protocol_search_term"
        lines = content.split('\n')
        protocol_start_idx = None

        # Find the line with the protocol
        for idx, line in enumerate(lines):
            if protocol_search_term in line and line.startswith('##'):
                protocol_start_idx = idx
                break

        if protocol_start_idx is None:
            print(f"Protocol '{protocol_search_term}' not found in {PROTOCOL_MAP_FILE}")
            return None

        # Extract goal: find "Цель: " and get text until next line
        goal = None
        goal_idx = None
        for idx in range(protocol_start_idx, len(lines)):
            if lines[idx].startswith('Цель:'):
                goal = lines[idx].replace('Цель:', '').strip()
                goal_idx = idx
                break

        if goal is None:
            print(f"Goal not found for protocol '{protocol_search_term}'")
            return None

        # Extract exercises: find "Интервенции:" and collect all lines starting with "*"
        exercises = []
        exercises_idx = None
        for idx in range(goal_idx, len(lines)):
            if 'Интервенции:' in lines[idx]:
                exercises_idx = idx
                break

        if exercises_idx is None:
            print(f"Exercises section not found for protocol '{protocol_search_term}'")
            return None

        # Collect all exercises (lines starting with "*")
        for idx in range(exercises_idx + 1, len(lines)):
            line = lines[idx].strip()

            # Stop if we hit another section (another "##" or empty lines followed by section)
            if line.startswith('##'):
                break

            # Extract exercise if line starts with "*"
            if line.startswith('*'):
                exercise_text = line.lstrip('*').strip()
                # Only add non-empty exercises that contain at least one letter
                if exercise_text and any(c.isalpha() for c in exercise_text):
                    exercises.append(exercise_text)

        return {
            'goal': goal,
            'exercises': exercises
        }

    except Exception as e:
        print(f"Error extracting protocol data: {e}")
        return None


def extract_intervention_description(exercise_name):
    """
    Search for intervention description in interventions.md file.

    Args:
        exercise_name (str): Exercise name to search for

    Returns:
        str: Full description or None if not found
    """
    try:
        if not os.path.exists(INTERVENTIONS_FILE):
            print(f"Error: {INTERVENTIONS_FILE} not found")
            return None

        with open(INTERVENTIONS_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        # Clean exercise name - extract main part before parentheses and remove punctuation
        # Example: "Письмо себе с добротой." -> "Письмо себе с добротой"
        search_term = exercise_name.split('(')[0].strip()
        # Remove trailing punctuation
        search_term = search_term.rstrip('.!?,;:')

        # Handle exercises with slash - take the first part for search
        # Example: "Дыхание 4-7-8 / мышечная релаксация" -> search for "Дыхание"
        search_parts = search_term.split('/')
        main_search_term = search_parts[0].strip()

        # Find the section containing the intervention
        lines = content.split('\n')
        intervention_start_idx = None
        best_match_score = 0
        best_match_idx = None

        # Look for the intervention header with fuzzy matching
        for idx, line in enumerate(lines):
            if line.startswith('###'):
                # Extract the title part from the header line
                # Example: "### 🧠 1. Запись тревожных мыслей (Thought Record)" -> "Запись тревожных мыслей"
                line_clean = line.replace('###', '').strip()

                # Remove emoji and number prefix if present
                line_clean = re.sub(r'^[^\w]*\d+\.\s*', '', line_clean)  # Remove emoji and number
                line_clean = line_clean.split('(')[0].strip()  # Remove parenthetical content

                # Try multiple matching strategies
                score = 0

                # Strategy 1: Full search term match
                search_words = set(search_term.lower().split())
                line_words = set(line_clean.lower().split())
                common_words = search_words.intersection(line_words)

                if len(common_words) > 0:
                    score = len(common_words) / max(len(search_words), 1)

                # Strategy 2: Main term match (for slashed terms)
                if '/' in search_term:
                    main_words = set(main_search_term.lower().split())
                    main_common = main_words.intersection(line_words)
                    if len(main_common) > 0:
                        alt_score = len(main_common) / max(len(main_words), 1)
                        score = max(score, alt_score)

                # Strategy 3: Check if key terms are present
                # Special handling for specific known patterns
                if "дыхание" in search_term.lower() and "дыхание" in line_clean.lower():
                    score = max(score, 0.7)
                if "mindfulness" in search_term.lower() and "mindfulness" in line_clean.lower():
                    score = max(score, 0.7)

                # Boost score for exact substring match
                if search_term.lower() in line_clean.lower() or line_clean.lower() in search_term.lower():
                    score += 0.5
                elif main_search_term.lower() in line_clean.lower():
                    score += 0.3

                if score > best_match_score and score >= 0.5:  # Lower threshold to 50%
                    best_match_score = score
                    best_match_idx = idx
                    intervention_start_idx = idx
                    print(f"Found match for '{search_term}' -> '{line_clean}' with score {score:.2f}")

        if intervention_start_idx is None:
            print(f"Intervention '{search_term}' not found in {INTERVENTIONS_FILE} (best score: {best_match_score:.2f})")
            return None

        # Extract full description until next section (###) or end
        description_lines = []
        for idx in range(intervention_start_idx, len(lines)):
            line = lines[idx]

            # Stop at next section
            if idx != intervention_start_idx and line.startswith('###'):
                break

            # Skip empty lines at the beginning
            if not description_lines and not line.strip():
                continue

            description_lines.append(line)

        return '\n'.join(description_lines)

    except Exception as e:
        print(f"Error extracting intervention description: {e}")
        return None


def escape_markdown(text):
    """
    Escape special characters for Telegram Markdown parsing.

    Args:
        text (str): Text to escape

    Returns:
        str: Escaped text safe for Telegram Markdown
    """
    # Characters that need to be escaped in Telegram Markdown
    special_chars = ['*', '_', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']

    for char in special_chars:
        text = text.replace(char, f'\\{char}')

    return text


def clean_markdown_for_telegram(text):
    """
    Clean and prepare markdown text for Telegram.
    Removes problematic formatting while preserving readability.

    Args:
        text (str): Markdown text from interventions.md

    Returns:
        str: Cleaned text suitable for Telegram
    """
    # Remove the heading markers but keep the title
    text = re.sub(r'^###\s*', '', text, flags=re.MULTILINE)

    # Remove single # on its own line
    text = re.sub(r'^\s*#\s*$', '', text, flags=re.MULTILINE)

    # Replace bullet points with simple dashes
    text = re.sub(r'^\s*\*\s+', '• ', text, flags=re.MULTILINE)

    # Replace numbered lists
    text = re.sub(r'^(\d+)\.\s+', r'\1. ', text, flags=re.MULTILINE)

    # Remove horizontal rules and symbol-only lines
    text = re.sub(r'^\s*\*\s*\*\s*\*\s*$', '', text, flags=re.MULTILINE)

    # Filter out lines that contain only symbols (no letters)
    lines = text.split('\n')
    filtered_lines = []
    for line in lines:
        # Keep the line if it's empty or contains at least one letter
        if not line.strip() or any(c.isalpha() for c in line):
            filtered_lines.append(line)
        else:
            # Skip lines like "• * *" or "---" etc.
            continue

    text = '\n'.join(filtered_lines)

    # Remove excessive blank lines
    text = re.sub(r'\n{3,}', '\n\n', text)

    return text.strip()


def save_exercise_progress_to_excel(user_id, username, protocol_name, exercise_name, action):
    """Save exercise progress to Excel file"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Messages'

        # Find next empty row
        next_row = ws.max_row + 1

        # Add exercise progress data
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = f"Protocol: {protocol_name}"
        ws[f'D{next_row}'] = f"Exercise: {exercise_name} - Action: {action}"
        ws[f'E{next_row}'] = 'exercise_progress'
        ws[f'G{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save(EXCEL_FILE)
        print(f"Exercise progress saved: {username} - {exercise_name} - {action}")

    except Exception as e:
        print(f"Error saving exercise progress to Excel: {e}")


async def send_protocol_selection(bot, chat_id):
    """Send protocol selection message with inline buttons"""
    try:
        text = "Отлично! Вот направления, с которыми я могу помочь 👇"

        # Create inline keyboard with protocol buttons
        markup = types.InlineKeyboardMarkup()

        for display_name, search_term, short_id in PROTOCOLS:
            btn = types.InlineKeyboardButton(
                display_name,
                callback_data=f"ps:{short_id}"  # Use short ID to avoid length limit
            )
            markup.add(btn)

        # Add menu button at the bottom
        btn_menu = types.InlineKeyboardButton(
            "📱 Главное меню",
            callback_data="menu:show"
        )
        markup.add(btn_menu)

        await bot.send_message(chat_id, text, reply_markup=markup)
        print(f"Protocol selection buttons sent to chat {chat_id}")

    except Exception as e:
        print(f"Error sending protocol selection: {e}")


async def handle_protocol_selection(bot, callback_query, protocol_id):
    """
    Handle protocol selection by user.
    Extract goal and exercises, prepare data for exercise module.

    Args:
        bot: Telegram bot instance
        callback_query: Callback query from button press
        protocol_id (str): Short ID of selected protocol (e.g., 'p1', 'p2')
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'

        # Get the actual protocol name from the ID
        if protocol_id not in PROTOCOL_ID_MAP:
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                callback_query.message.chat.id,
                "Неизвестный протокол. Пожалуйста, выберите из списка."
            )
            return

        protocol_name = PROTOCOL_ID_MAP[protocol_id]

        # Extract protocol data
        protocol_data = extract_protocol_data(protocol_name)

        if protocol_data is None:
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                callback_query.message.chat.id,
                f"Извините, не удалось найти информацию о протоколе '{protocol_name}'."
            )
            return

        # Store user state
        goal = protocol_data['goal']
        exercises = protocol_data['exercises']

        if not exercises:
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                callback_query.message.chat.id,
                f"Не найдены упражнения для протокола '{protocol_name}'."
            )
            return

        user_exercise_states[user_id] = {
            'protocol_id': protocol_id,
            'protocol_name': protocol_name,
            'exercises': exercises,
            'current_index': 0
        }

        # Answer callback
        await bot.answer_callback_query(callback_query.id)

        # Send protocol name and goal
        protocol_text = f"📘 Протокол \"{protocol_name}\"\n\n"
        protocol_text += f"Цель: {goal}"
        await bot.send_message(callback_query.message.chat.id, protocol_text)

        # Send first exercise
        await send_exercise(bot, callback_query.message.chat.id, user_id, username)

        print(f"Protocol selection handled for {protocol_name} by user {username}")

    except Exception as e:
        print(f"Error handling protocol selection: {e}")
        await bot.answer_callback_query(callback_query.id)
        await bot.send_message(
            callback_query.message.chat.id,
            "Произошла ошибка при обработке выбора протокола."
        )


async def send_exercise(bot, chat_id, user_id, username):
    """Send current exercise to user with buttons"""
    try:
        if user_id not in user_exercise_states:
            await bot.send_message(chat_id, "Не найдена информация о вашем протоколе.")
            return

        state = user_exercise_states[user_id]
        current_index = state['current_index']
        exercises = state['exercises']
        protocol_name = state['protocol_name']
        protocol_id = state['protocol_id']

        if current_index >= len(exercises):
            # All exercises completed
            await send_completion_message(bot, chat_id, user_id, username)
            return

        # Get current exercise
        current_exercise = exercises[current_index]

        # Send exercise header
        if current_index == 0:
            exercise_text = f"Первое упражнение:\n✍️ {current_exercise}"
        else:
            exercise_text = f"Упражнение {current_index + 1}:\n✍️ {current_exercise}"

        # Create buttons
        markup = types.InlineKeyboardMarkup()
        btn_start = types.InlineKeyboardButton(
            "🧩 Начать упражнение",
            callback_data=f"ex_start:{protocol_id}:{current_index}"
        )
        btn_skip = types.InlineKeyboardButton(
            "⏭ Пропустить",
            callback_data=f"ex_skip:{protocol_id}:{current_index}"
        )
        btn_menu = types.InlineKeyboardButton(
            "📱 Главное меню",
            callback_data="menu:show"
        )
        markup.row(btn_start)
        markup.row(btn_skip)
        markup.row(btn_menu)

        await bot.send_message(chat_id, exercise_text, reply_markup=markup)

    except Exception as e:
        print(f"Error sending exercise: {e}")


async def handle_exercise_start(bot, callback_query, protocol_id, exercise_index):
    """Handle exercise start button press"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(callback_query.message.chat.id, "Сессия истекла. Начните заново.")
            return

        state = user_exercise_states[user_id]

        # Verify protocol and index match
        if state['protocol_id'] != protocol_id or state['current_index'] != int(exercise_index):
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(callback_query.message.chat.id, "Неверная последовательность упражнений.")
            return

        exercise_name = state['exercises'][int(exercise_index)]
        protocol_name = state['protocol_name']

        # Save progress to Excel
        save_exercise_progress_to_excel(user_id, username, protocol_name, exercise_name, "started")

        # Get intervention description
        description = extract_intervention_description(exercise_name)

        await bot.answer_callback_query(callback_query.id)

        if description:
            # Clean the markdown for Telegram
            cleaned_description = clean_markdown_for_telegram(description)

            # Send as plain text (more reliable)
            await bot.send_message(callback_query.message.chat.id, cleaned_description)
        else:
            await bot.send_message(
                callback_query.message.chat.id,
                f"Описание не найдено для упражнения: {exercise_name}\n\n"
                f"Протокол: {protocol_name}"
            )

        # Move to next exercise
        state['current_index'] += 1

        # Send next exercise or completion
        await send_exercise(bot, callback_query.message.chat.id, user_id, username)

    except Exception as e:
        print(f"Error handling exercise start: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_exercise_skip(bot, callback_query, protocol_id, exercise_index):
    """Handle exercise skip button press"""
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(callback_query.message.chat.id, "Сессия истекла. Начните заново.")
            return

        state = user_exercise_states[user_id]

        # Verify protocol and index match
        if state['protocol_id'] != protocol_id or state['current_index'] != int(exercise_index):
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(callback_query.message.chat.id, "Неверная последовательность упражнений.")
            return

        exercise_name = state['exercises'][int(exercise_index)]
        protocol_name = state['protocol_name']

        # Save progress to Excel
        save_exercise_progress_to_excel(user_id, username, protocol_name, exercise_name, "skipped")

        await bot.answer_callback_query(callback_query.id, "Упражнение пропущено")

        # Move to next exercise
        state['current_index'] += 1

        # Send next exercise or completion
        await send_exercise(bot, callback_query.message.chat.id, user_id, username)

    except Exception as e:
        print(f"Error handling exercise skip: {e}")
        await bot.answer_callback_query(callback_query.id)


async def send_completion_message(bot, chat_id, user_id, username):
    """Send completion message when all exercises are done"""
    try:
        if user_id in user_exercise_states:
            protocol_name = user_exercise_states[user_id]['protocol_name']
            save_exercise_progress_to_excel(user_id, username, protocol_name, "All exercises", "completed")

            # Clear user state
            del user_exercise_states[user_id]

        await bot.send_message(chat_id, "Ты молодец! Захочешь — продолжим завтра :)")
        print(f"Protocol completed by user {username}")

        # Get user name from greeting state
        from greeting import user_states
        user_name = 'User'
        if user_id in user_states:
            user_name = user_states[user_id].get('user_name', 'User')

        # Show main menu
        from universal_menu import show_main_menu
        await show_main_menu(bot, chat_id, user_id, username, user_name)

    except Exception as e:
        print(f"Error sending completion message: {e}")
