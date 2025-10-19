# -*- coding: utf-8 -*-
"""
Safety and crisis detection module for therapy bot
Provides unified safety checks across all user inputs
"""

import json
import hashlib
from datetime import datetime, timedelta
from typing import Tuple, Optional, Dict, Any, List
from telebot import types
import os
from openpyxl import load_workbook

# Import LLM client for analysis
from openrouter import OpenRouterClient
from config import MODEL_SIMPLE, TEMPERATURE, TOP_P, TOP_K

# Cache for LLM responses to avoid repeated checks
safety_cache: Dict[str, Dict[str, Any]] = {}
CACHE_TTL_MINUTES = 30  # Shorter cache for safety checks

# Crisis keywords for quick detection (Russian)
CRISIS_KEYWORDS = [
    # Suicidal ideation
    'суицид', 'самоубийство', 'покончить с собой', 'не хочу жить',
    'умереть', 'убить себя', 'прыгнуть с', 'повеситься', 'отравиться',
    'вскрыть вены', 'таблетки выпить', 'нет смысла жить',

    # Self-harm
    'порезать себя', 'причинить себе боль', 'наказать себя',
    'резать руки', 'бить себя', 'самоповреждение',

    # Psychosis indicators
    'голоса в голове', 'преследуют', 'следят за мной', 'читают мысли',
    'управляют мной', 'заговор против меня', 'все против меня',

    # Severe dissociation
    'не чувствую тело', 'это не я', 'смотрю на себя со стороны',
    'не реально', 'все как во сне', 'отключаюсь от реальности',

    # Substance abuse crisis
    'передозировка', 'не могу остановиться', 'ломка', 'абстиненция',

    # Mania/extreme states
    'не сплю неделю', 'могу всё', 'я бог', 'не нужна еда',
    'энергия бьёт ключом', 'не могу остановиться'
]

# Help resources text
HELP_TEXT = """
🆘 **Экстренная помощь:**

**Горячие линии (круглосуточно, бесплатно):**
• 8-800-2000-122 - Детский телефон доверия
• 8-800-100-0191 - Кризисная линия доверия
• 051 - Телефон доверия (с городского)

**Онлайн-поддержка:**
• tvoyteritoriya.online - чат с психологом
• www.ya-roditel.ru - помощь родителям

**Экстренная помощь:**
• 112 - Единая служба экстренной помощи
• 103 - Скорая медицинская помощь

Помни: кризис временный, помощь доступна! 💙
"""


def get_cache_key(text: str, check_type: str) -> str:
    """Generate cache key for safety check results"""
    text_hash = hashlib.md5(text.encode()).hexdigest()[:8]
    return f"safety_{check_type}_{text_hash}"


def get_cached_result(cache_key: str) -> Optional[Tuple[bool, Optional[str]]]:
    """Get cached safety check result if still valid"""
    if cache_key in safety_cache:
        cached_data = safety_cache[cache_key]
        if datetime.now() - cached_data['timestamp'] < timedelta(minutes=CACHE_TTL_MINUTES):
            return cached_data['result']
    return None


def set_cached_result(cache_key: str, result: Tuple[bool, Optional[str]]) -> None:
    """Store safety check result in cache"""
    safety_cache[cache_key] = {
        'result': result,
        'timestamp': datetime.now()
    }


def quick_keyword_check(text: str) -> Tuple[bool, Optional[str]]:
    """
    Quick check for crisis keywords without LLM
    Returns: (crisis_detected, crisis_type)
    """
    text_lower = text.lower()

    # Check each keyword
    for keyword in CRISIS_KEYWORDS:
        if keyword in text_lower:
            # Determine crisis type
            if any(word in keyword for word in ['суицид', 'самоубийство', 'покончить', 'не хочу жить', 'умереть', 'убить себя']):
                return True, "Суицидальные мысли"
            elif any(word in keyword for word in ['порезать', 'причинить себе боль', 'резать руки']):
                return True, "Самоповреждение"
            elif any(word in keyword for word in ['голоса', 'преследуют', 'следят', 'читают мысли', 'управляют', 'заговор']):
                return True, "Психотические симптомы"
            elif any(word in keyword for word in ['не чувствую тело', 'это не я', 'не реально', 'как во сне']):
                return True, "Диссоциация"
            elif any(word in keyword for word in ['передозировка', 'ломка', 'абстиненция']):
                return True, "Кризис зависимости"
            elif any(word in keyword for word in ['не сплю неделю', 'я бог', 'энергия бьёт']):
                return True, "Маниакальное состояние"
            else:
                return True, "Кризисное состояние"

    return False, None


async def check_text_safety(text: str, context: str = "general") -> Tuple[bool, Optional[str], float]:
    """
    Check text for crisis indicators using both keyword and LLM analysis

    Args:
        text: Text to analyze
        context: Context of the text (exercise, diary, checkin, etc.)

    Returns:
        (crisis_detected, crisis_type, confidence_score)
    """
    if not text or len(text.strip()) < 3:
        return False, None, 0.0

    # Check cache first
    cache_key = get_cache_key(text, context)
    cached = get_cached_result(cache_key)
    if cached:
        crisis_detected, crisis_type = cached
        return crisis_detected, crisis_type, 1.0 if crisis_detected else 0.0

    # Quick keyword check first
    keyword_crisis, keyword_type = quick_keyword_check(text)
    if keyword_crisis:
        set_cached_result(cache_key, (True, keyword_type))
        return True, keyword_type, 0.95

    # LLM analysis for more nuanced detection
    try:
        system_prompt = """Ты опытный кризисный психолог. Проанализируй текст на наличие кризисных индикаторов.

КРИТИЧЕСКИЕ индикаторы (требуют немедленной помощи):
1. Суицидальные мысли/намерения/планы
2. Самоповреждение (порезы, ожоги, удары)
3. Психотические симптомы (галлюцинации, бред, паранойя)
4. Тяжёлая диссоциация/дереализация
5. Острый кризис зависимости
6. Маниакальные/экстремальные состояния

ВАЖНО:
- Будь внимателен к скрытым признакам (безнадёжность, ощущение ловушки)
- Учитывай контекст: упоминание прошлого опыта vs текущее состояние
- Избегай ложных срабатываний на обсуждение темы в общем

Ответь JSON:
{
  "crisis_detected": boolean,
  "crisis_type": "тип кризиса" или null,
  "confidence": float (0.0-1.0),
  "reasoning": "краткое обоснование"
}"""

        user_prompt = f"""Контекст: {context}

Текст для анализа:
{text}

Определи наличие кризисных индикаторов."""

        client = OpenRouterClient()
        response, _ = client.get_simple_response(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            model=MODEL_SIMPLE,
            temperature=0.1,  # Low temperature for consistency
            top_p=TOP_P,
            top_k=TOP_K
        )

        # Parse JSON response
        result = json.loads(response.strip())

        crisis_detected = result.get('crisis_detected', False)
        crisis_type = result.get('crisis_type')
        confidence = result.get('confidence', 0.0)

        # Only flag crisis if confidence is high enough
        if crisis_detected and confidence >= 0.7:
            set_cached_result(cache_key, (True, crisis_type))
            return True, crisis_type, confidence
        else:
            set_cached_result(cache_key, (False, None))
            return False, None, confidence

    except Exception as e:
        print(f"Error in LLM safety check: {e}")
        # On error, rely on keyword check only
        return keyword_crisis, keyword_type, 0.5 if keyword_crisis else 0.0


async def show_crisis_support(bot, chat_id: int, user_name: str, crisis_type: str,
                              context: str = "general", continue_after: bool = False):
    """
    Show crisis support message and resources

    Args:
        bot: Telegram bot instance
        chat_id: Chat ID to send message to
        user_name: User's name for personalization
        crisis_type: Type of crisis detected
        context: Where crisis was detected (exercise, diary, etc.)
        continue_after: Whether to allow continuing after showing support
    """
    try:
        # Ensure user_name is not empty or default
        if not user_name or user_name == "User" or user_name == "Друг":
            user_name = "Дорогой друг"

        # Personalized message based on crisis type
        intro_messages = {
            "Суицидальные мысли": f"{user_name}, я очень беспокоюсь за тебя. Эти мысли - сигнал о сильной боли.",
            "Самоповреждение": f"{user_name}, я вижу, что тебе очень тяжело. Боль, которую ты чувствуешь, реальна.",
            "Психотические симптомы": f"{user_name}, то, что ты описываешь, требует профессиональной поддержки.",
            "Диссоциация": f"{user_name}, ощущение отключения от реальности может быть очень пугающим.",
            "Кризис зависимости": f"{user_name}, борьба с зависимостью требует профессиональной помощи.",
            "Маниакальное состояние": f"{user_name}, важно стабилизировать твоё состояние с помощью специалиста.",
            "Кризисное состояние": f"{user_name}, я чувствую, что тебе сейчас очень трудно."
        }

        intro = intro_messages.get(crisis_type, f"{user_name}, я переживаю за тебя.")

        text = (
            f"{intro}\n\n"
            f"Сейчас самое важное - получить поддержку. "
            f"Ты не один/одна в этом.\n\n"
            f"{HELP_TEXT}"
        )

        # Create buttons
        markup = types.InlineKeyboardMarkup()

        # Emergency help button
        btn_help = types.InlineKeyboardButton(
            "🆘 Горячие линии",
            callback_data="safety:hotlines"
        )

        markup.add(btn_help)

        # Add continue option if allowed
        if continue_after:
            btn_continue = types.InlineKeyboardButton(
                "➡️ Продолжить позже",
                callback_data=f"safety:continue_{context}"
            )
            markup.add(btn_continue)

        # Always add main menu button
        btn_menu = types.InlineKeyboardButton(
            "📱 Главное меню",
            callback_data="menu:show"
        )
        markup.add(btn_menu)

        await bot.send_message(chat_id, text, reply_markup=markup, parse_mode='Markdown')

    except Exception as e:
        print(f"Error showing crisis support: {e}")




async def log_crisis_detection(user_id: int, username: str, crisis_type: str,
                               context: str, text_sample: str, file_path: str = None):
    """
    Log crisis detection to appropriate Excel file

    Args:
        user_id: User's Telegram ID
        username: User's username
        crisis_type: Type of crisis detected
        context: Where detected (exercise, diary, checkin)
        text_sample: Sample of text that triggered detection
        file_path: Excel file to log to (optional)
    """
    try:
        # Determine file based on context if not provided
        if not file_path:
            file_map = {
                'exercise': 'exercises.xlsx',
                'diary': 'diary.xlsx',
                'checkin': 'check_in.xlsx',
                'mvst': 'mvst.xlsx'
            }
            file_path = file_map.get(context, 'safety_log.xlsx')

        # Create safety log if needed
        if file_path == 'safety_log.xlsx' and not os.path.exists(file_path):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Safety Log'

            headers = [
                'User ID', 'Username', 'Detection Time', 'Crisis Type',
                'Context', 'Text Sample', 'Action Taken'
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(file_path)

        # Log to file
        if os.path.exists(file_path):
            wb = load_workbook(file_path)

            # Find or create safety sheet
            if 'Safety' in wb.sheetnames:
                ws = wb['Safety']
            else:
                ws = wb.create_sheet('Safety')
                # Add headers if new sheet
                headers = [
                    'User ID', 'Username', 'Detection Time', 'Crisis Type',
                    'Context', 'Text Sample'
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            # Add crisis record
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=user_id)
            ws.cell(row=next_row, column=2, value=username)
            ws.cell(row=next_row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=next_row, column=4, value=crisis_type)
            ws.cell(row=next_row, column=5, value=context)
            ws.cell(row=next_row, column=6, value=text_sample[:200])  # Limit sample length

            wb.save(file_path)
            print(f"Logged crisis detection for user {user_id} in {file_path}")

    except Exception as e:
        print(f"Error logging crisis detection: {e}")


def register_safety_handlers(bot):
    """Register safety-related callback handlers"""

    @bot.callback_query_handler(func=lambda call: call.data.startswith('safety:'))
    async def safety_callback_handler(callback_query):
        """Handle safety-related callbacks"""
        try:
            data = callback_query.data.replace('safety:', '')
            chat_id = callback_query.message.chat.id

            # Answer callback immediately
            await bot.answer_callback_query(callback_query.id, show_alert=False)

            if data == 'hotlines':
                # Show hotlines again with emphasis
                text = (
                    "📞 **Горячие линии помощи:**\n\n"
                    "**Бесплатно и анонимно:**\n"
                    "☎️ **8-800-2000-122** - Детский телефон доверия\n"
                    "☎️ **8-800-100-0191** - Кризисная линия\n"
                    "☎️ **051** - С городского телефона\n\n"
                    "**Экстренно:**\n"
                    "🚨 **112** - Единая служба\n"
                    "🚑 **103** - Скорая помощь\n\n"
                    "Не стесняйся обращаться за помощью. "
                    "Специалисты готовы поддержать тебя. 💙"
                )

                markup = types.InlineKeyboardMarkup()
                btn_menu = types.InlineKeyboardButton(
                    "📱 Главное меню",
                    callback_data="menu:show"
                )
                markup.add(btn_menu)

                await bot.send_message(chat_id, text, reply_markup=markup, parse_mode='Markdown')


            elif data.startswith('continue_'):
                # Continue with previous activity
                context = data.replace('continue_', '')

                text = (
                    "Я сохранил информацию о твоём состоянии. "
                    "Пожалуйста, обратись за профессиональной помощью как можно скорее.\n\n"
                    "Если почувствуешь ухудшение - сразу используй горячие линии."
                )

                markup = types.InlineKeyboardMarkup()
                
                # Add context-specific continue buttons
                if context == "goal_setting":
                    btn_continue = types.InlineKeyboardButton(
                        "➡️ Продолжить постановку цели",
                        callback_data="goal_continue:after_safety"
                    )
                    markup.add(btn_continue)
                elif context == "exercise":
                    btn_continue = types.InlineKeyboardButton(
                        "➡️ Продолжить упражнение",
                        callback_data="exercise_continue:after_safety"
                    )
                    markup.add(btn_continue)
                
                btn_menu = types.InlineKeyboardButton(
                    "📱 Главное меню",
                    callback_data="menu:show"
                )
                markup.add(btn_menu)

                await bot.send_message(chat_id, text, reply_markup=markup)

        except Exception as e:
            print(f"Error handling safety callback: {e}")
            await bot.answer_callback_query(callback_query.id)


# Export public interface
__all__ = [
    'check_text_safety',
    'show_crisis_support',
    'log_crisis_detection',
    'register_safety_handlers',
    'HELP_TEXT'
]