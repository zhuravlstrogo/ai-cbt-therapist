# -*- coding: utf-8 -*-
"""
Greeting module for AI Psychologist bot
Handles first interaction with users after /start command
New flow: Form of address (—Ç—ã/–í—ã) ‚Üí Name input ‚Üí Motivation message
"""

import os
import asyncio
from datetime import datetime
from openpyxl import load_workbook, Workbook
from telebot import types

# Store user states to track where they are in the greeting process
# Format: {user_id: {'stage': 'awaiting_consent'|'awaiting_form_choice'|'awaiting_name'|'ready_to_start', 'form': '—Ç—ã'|'–í—ã', 'user_name': str}}
user_states = {}

# Excel file path
EXCEL_FILE = 'messages.xlsx'


def init_greeting_excel_file():
    """Initialize Excel file with headers for greeting data"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Messages'
        ws['A1'] = 'User ID'
        ws['B1'] = 'Username'
        ws['C1'] = 'User Name'
        ws['D1'] = 'Message Text'
        ws['E1'] = 'Message Type'
        ws['F1'] = 'Form of Address'  # '—Ç—ã' or '–í—ã'
        ws['G1'] = 'Protocol Choice'
        ws['H1'] = 'Date Time'
        wb.save(EXCEL_FILE)


def update_excel_headers():
    """Update existing Excel file to include new columns if they don't exist"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            # Check if headers exist, if not add them
            if ws['A1'].value is None:
                ws['A1'] = 'User ID'
            if ws['C1'].value != 'User Name':
                ws['C1'] = 'User Name'
            if ws['E1'].value != 'Message Type':
                ws['E1'] = 'Message Type'
            if ws['F1'].value != 'Form of Address':
                ws['F1'] = 'Form of Address'
            if ws['G1'].value != 'Protocol Choice':
                ws['G1'] = 'Protocol Choice'

            wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error updating Excel headers: {e}")


def save_form_of_address_to_excel(user_id, username, form_of_address):
    """Save form of address (—Ç—ã/–í—ã) to Excel file"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            init_greeting_excel_file()
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        # Add form of address data
        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'F{next_row}'] = form_of_address
        ws[f'E{next_row}'] = 'form_of_address_choice'
        ws[f'H{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"Form of address saved to Excel: {username} - {form_of_address}")
    except Exception as e:
        print(f"Error saving form of address to Excel: {e}")


def get_form_of_address_from_excel(user_id):
    """Get form of address for user from Excel file (for subsequent /start calls)"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            # Find the last row for this user with form of address
            for row in range(ws.max_row, 0, -1):
                if ws[f'A{row}'].value == user_id and ws[f'F{row}'].value in ['—Ç—ã', '–í—ã']:
                    return ws[f'F{row}'].value

            return None
    except Exception as e:
        print(f"Error getting form of address from Excel: {e}")
        return None


def save_user_name_to_excel(user_id, username, user_name):
    """Save user name to Excel file"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            init_greeting_excel_file()
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        # Find the last row for this user and update it with the name
        for row in range(ws.max_row, 0, -1):
            if ws[f'A{row}'].value == user_id:
                ws[f'C{row}'] = user_name
                ws[f'D{row}'] = f"User provided name: {user_name}"
                ws[f'E{row}'] = 'name_input'
                ws[f'H{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                break

        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"User name saved to Excel: {username} - {user_name}")
    except Exception as e:
        print(f"Error saving user name to Excel: {e}")


def save_protocol_choice_to_excel(user_id, username, protocol_choice):
    """Save protocol choice to Excel file"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            # Find the last row for this user and update protocol choice
            for row in range(ws.max_row, 0, -1):
                if ws[f'A{row}'].value == user_id:
                    ws[f'G{row}'] = protocol_choice
                    break

            wb.save(EXCEL_FILE)
            print(f"Protocol choice saved to Excel: {username} - {protocol_choice}")
    except Exception as e:
        print(f"Error saving protocol choice to Excel: {e}")


async def send_greeting_messages(bot, chat_id, user_id, username):
    """Send initial greeting message with consent confirmation"""
    try:
        # Always send greeting text
        greeting_text = (
            "–ü—Ä–∏–≤–µ—Ç üëã Aide ‚Äì —Ç–≤–æ–π –ò–ò-–ø–æ–º–æ—â–Ω–∏–∫, —Ä–∞–±–æ—Ç–∞—é—â–∏–π –≤ —Ä–∞–º–∫–∞—Ö –º–µ—Ç–æ–¥–æ–≤ –∫–æ–≥–Ω–∏—Ç–∏–≤–Ω–æ-–ø–æ–≤–µ–¥–µ–Ω—á–µ—Å–∫–æ–π —Ç–µ—Ä–∞–ø–∏–∏ (–ö–ü–¢).\n\n"
            "–ö—Ä–∞—Ç–∫–æ –æ –ø–æ–¥—Ö–æ–¥–µ üìò\n"
            "–ö–ü–¢ ‚Äî –¥–æ–∫–∞–∑–∞—Ç–µ–ª—å–Ω—ã–π –º–µ—Ç–æ–¥ –ø—Å–∏—Ö–æ–ª–æ–≥–∏—á–µ—Å–∫–æ–π –ø–æ–º–æ—â–∏. –ú—ã –±—É–¥–µ–º –≤—ã—è–≤–ª—è—Ç—å –º—ã—Å–ª–∏ –∏ –ø–æ–≤–µ–¥–µ–Ω—á–µ—Å–∫–∏–µ –ø—Ä–∏–≤—ã—á–∫–∏, –∫–æ—Ç–æ—Ä—ã–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞—é—Ç –¥–∏—Å–∫–æ–º—Ñ–æ—Ä—Ç, –∏ —Å–∏—Å—Ç–µ–º–Ω–æ –º–µ–Ω—è—Ç—å –∏—Ö —Å –ø–æ–º–æ—â—å—é —Å—Ç—Ä—É–∫—Ç—É—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö —É–ø—Ä–∞–∂–Ω–µ–Ω–∏–π –∏ –∫–æ—Ä–æ—Ç–∫–∏—Ö –ø—Ä–∞–∫—Ç–∏–∫. –Ø –æ–±—ä—è—Å–Ω—è—é —à–∞–≥–∏ –ø—Ä–æ—Å—Ç—ã–º —è–∑—ã–∫–æ–º –∏ –ø—Ä–µ–¥–ª–∞–≥–∞—é —Ç–µ—Ö–Ω–∏–∫–∏, –∫–æ—Ç–æ—Ä—ã–µ –º–æ–∂–Ω–æ –ø—Ä–∏–º–µ–Ω—è—Ç—å –≤ –ø–æ–≤—Å–µ–¥–Ω–µ–≤–Ω–æ–π –∂–∏–∑–Ω–∏ üß≠\n\n"
            "‚ö†Ô∏è –í–∞–∂–Ω–æ:\n"
            "Aide –Ω–µ —è–≤–ª—è–µ—Ç—Å—è –∑–∞–º–µ–Ω–æ–π –ø—Ä–æ—Ñ–µ—Å—Å–∏–æ–Ω–∞–ª—å–Ω–æ–π –ø—Å–∏—Ö–æ—Ç–µ—Ä–∞–ø–µ–≤—Ç–∏—á–µ—Å–∫–æ–π –ø–æ–º–æ—â–∏, –Ω–µ –ø—Ä–µ–¥–Ω–∞–∑–Ω–∞—á–µ–Ω –¥–ª—è –ª–µ—á–µ–Ω–∏—è —Ç—è–∂–µ–ª—ã—Ö —Ä–∞—Å—Å—Ç—Ä–æ–π—Å—Ç–≤ –∏ –ø–æ–º–æ—â–∏ –≤ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö, –∂–∏–∑–Ω–µ—É–≥—Ä–æ–∂–∞—é—â–∏—Ö —Å–∏—Ç—É–∞—Ü–∏—è—Ö, –∞ —Ç–∞–∫–∂–µ –Ω–µ –ø—Ä–µ–¥–Ω–∞–∑–Ω–∞—á–µ–Ω–∞ –¥–ª—è –ø–æ–º–æ—â–∏ –ª—é–¥—è–º –º–ª–∞–¥—à–µ 18 –ª–µ—Ç.\n\n"
            "–ö–æ–Ω—Ñ–∏–¥–µ–Ω—Ü–∏–∞–ª—å–Ω–æ—Å—Ç—å üîí\n"
            "–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è, –∫–æ—Ç–æ—Ä—É—é —Ç—ã —Å–æ–æ–±—â–∞–µ—à—å –≤ —á–∞—Ç–µ, –Ω–∞–¥–µ–∂–Ω–æ –∑–∞—â–∏—â–µ–Ω–∞ –∏ –Ω–µ –ø—Ä–µ–¥–Ω–∞–∑–Ω–∞—á–µ–Ω–∞ –¥–ª—è –ø–µ—Ä–µ–¥–∞—á–∏ —Ç—Ä–µ—Ç—å–∏–º –ª–∏—Ü–∞–º.\n\n"
            "–ï—Å–ª–∏ —É—Å–ª–æ–≤–∏—è –ø–æ–Ω—è—Ç–Ω—ã –∏ –ø–æ–¥—Ö–æ–¥—è—Ç, –ø–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥—Ç–≤–µ—Ä–¥–∏ —Å–æ–≥–ª–∞—Å–∏–µ, –∏ –º—ã –Ω–∞—á–Ω—ë–º —Ä–∞–±–æ—Ç—É ‚úÖ"
        )

        # Create inline keyboard with consent confirmation
        markup = types.InlineKeyboardMarkup()
        btn_consent = types.InlineKeyboardButton(
            "–î–∞, –≤—Å–µ –ø–æ–Ω—è—Ç–Ω–æ",
            callback_data="consent_confirmed"
        )
        markup.add(btn_consent)

        await bot.send_message(chat_id, greeting_text, reply_markup=markup)

        # Set user state to awaiting consent confirmation
        user_states[user_id] = {'stage': 'awaiting_consent'}

        print(f"Greeting message with consent request sent to user {username} (ID: {user_id})")
    except Exception as e:
        print(f"Error sending greeting messages: {e}")


async def ask_for_form_of_address(bot, chat_id, user_id, username):
    """Ask user for form of address after consent is confirmed"""
    try:
        form_question = "–û—Ç–ª–∏—á–Ω–æ! –ö–∞–∫ —è –º–æ–≥—É –∫ —Ç–µ–±–µ –æ–±—Ä–∞—â–∞—Ç—å—Å—è?"

        # Create inline keyboard with form of address options
        markup = types.InlineKeyboardMarkup()
        btn_ty = types.InlineKeyboardButton(
            "–ú–æ–∂–Ω–æ –Ω–∞ —Ç—ã",
            callback_data="form_address:ty"
        )
        btn_vy = types.InlineKeyboardButton(
            "–ú–æ–∂–Ω–æ –Ω–∞ –í—ã",
            callback_data="form_address:vy"
        )
        markup.add(btn_ty)
        markup.add(btn_vy)

        await bot.send_message(chat_id, form_question, reply_markup=markup)

        # Set user state to awaiting form choice
        user_states[user_id] = {'stage': 'awaiting_form_choice'}

        print(f"Form of address question sent to user {username}")
    except Exception as e:
        print(f"Error asking for form of address: {e}")


async def ask_for_user_name(bot, chat_id, user_id, username, form_of_address):
    """Ask user for their name based on form of address"""
    try:
        if form_of_address == '—Ç—ã':
            name_question = "–ö–∞–∫ —Ç–µ–±—è –Ω–∞–∑—ã–≤–∞—Ç—å –≤ –¥–∏–∞–ª–æ–≥–µ? üìù –ù–∞–ø–∏—à–∏ –∏–º—è –∏–ª–∏ –Ω–∏–∫."
        else:  # –í—ã
            name_question = "–ö–∞–∫ –í–∞—Å –Ω–∞–∑—ã–≤–∞—Ç—å –≤ –¥–∏–∞–ª–æ–≥–µ? üìù –ù–∞–ø–∏—à–∏—Ç–µ –∏–º—è –∏–ª–∏ –Ω–∏–∫."

        # Add menu button for accessibility
        from universal_menu import get_menu_button
        markup = get_menu_button()

        await bot.send_message(chat_id, name_question, reply_markup=markup)

        # Set user state to awaiting name input
        user_states[user_id] = {
            'stage': 'awaiting_name',
            'form': form_of_address
        }

        print(f"Name question sent to user {username} with form: {form_of_address}")
    except Exception as e:
        print(f"Error asking for user name: {e}")


async def send_motivation_message(bot, chat_id, user_id, username, form_of_address, user_name):
    """Send motivation message with 'Ready to start?' button"""
    try:
        if form_of_address == '—Ç—ã':
            motivation_text = (
                f"–û—Ç–ª–∏—á–Ω–æ, {user_name}! üéØ\n\n"
                "–û—Ç —Ç–æ–≥–æ, –Ω–∞—Å–∫–æ–ª—å–∫–æ —Å–µ—Ä—å—ë–∑–Ω–æ —Ç—ã –ø–æ–¥–æ–π–¥—ë—à—å –∫ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—é –∑–∞–¥–∞–Ω–∏–π, "
                "–±—É–¥–µ—Ç –∑–∞–≤–∏—Å–µ—Ç—å —Å–∫–æ—Ä–æ—Å—Ç—å –¥–æ—Å—Ç–∏–∂–µ–Ω–∏—è —Ü–µ–ª–∏ ‚≠êÔ∏è\n\n"
                "–†–µ–∫–æ–º–µ–Ω–¥—É—é —É–¥–µ–ª–∏—Ç—å –≤—Ä–µ–º—è —Å–µ–±–µ –∏ –Ω–µ –æ—Ç–≤–ª–µ–∫–∞—Ç—å—Å—è –¥–æ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è —É–ø—Ä–∞–∂–Ω–µ–Ω–∏—è ‚Äî "
                "—Ç–∞–∫ –≥–ª—É–±–∂–µ –ø–æ–≥—Ä—É–∑–∏—à—å—Å—è –∏ –±—ã—Å—Ç—Ä–µ–µ –∑–∞–º–µ—Ç–∏—à—å —ç—Ñ—Ñ–µ–∫—Ç.\n\n"
                "–ì–æ—Ç–æ–≤(–∞) –Ω–∞—á–∞—Ç—å?"
            )
        else:  # –í—ã
            motivation_text = (
                f"–û—Ç–ª–∏—á–Ω–æ, {user_name}! üéØ\n\n"
                "–û—Ç —Ç–æ–≥–æ, –Ω–∞—Å–∫–æ–ª—å–∫–æ —Å–µ—Ä—å—ë–∑–Ω–æ –í—ã –ø–æ–¥–æ–π–¥—ë—Ç–µ –∫ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—é –∑–∞–¥–∞–Ω–∏–π, "
                "–±—É–¥–µ—Ç –∑–∞–≤–∏—Å–µ—Ç—å —Å–∫–æ—Ä–æ—Å—Ç—å –¥–æ—Å—Ç–∏–∂–µ–Ω–∏—è —Ü–µ–ª–∏ ‚≠êÔ∏è\n\n"
                "–†–µ–∫–æ–º–µ–Ω–¥—É—é —É–¥–µ–ª–∏—Ç—å –≤—Ä–µ–º—è —Å–µ–±–µ –∏ –Ω–µ –æ—Ç–≤–ª–µ–∫–∞—Ç—å—Å—è –¥–æ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è —É–ø—Ä–∞–∂–Ω–µ–Ω–∏—è ‚Äî "
                "—Ç–∞–∫ –≥–ª—É–±–∂–µ –ø–æ–≥—Ä—É–∑–∏—Ç–µ—Å—å –∏ –±—ã—Å—Ç—Ä–µ–µ –∑–∞–º–µ—Ç–∏—Ç–µ —ç—Ñ—Ñ–µ–∫—Ç.\n\n"
                "–ì–æ—Ç–æ–≤—ã –Ω–∞—á–∞—Ç—å?"
            )

        # Create inline keyboard with button
        markup = types.InlineKeyboardMarkup()
        btn_ready = types.InlineKeyboardButton(
            "–î–∞, –ø–æ–µ—Ö–∞–ª–∏",
            callback_data="ready_to_start"
        )
        markup.add(btn_ready)

        await bot.send_message(chat_id, motivation_text, reply_markup=markup)

        # Set user state to ready to start
        user_states[user_id] = {
            'stage': 'ready_to_start',
            'form': form_of_address
        }

        print(f"Motivation message sent to user {username}")
    except Exception as e:
        print(f"Error sending motivation message: {e}")


async def handle_consent_confirmation(bot, callback_query, user_id, username):
    """Handle consent confirmation"""
    try:
        choice = callback_query.data

        if choice == "consent_confirmed":
            # Save consent confirmation to Excel
            save_form_of_address_to_excel(user_id, username, 'consent_confirmed')

            # Answer callback and ask for form of address
            await bot.answer_callback_query(callback_query.id)
            print(f"DEBUG: About to ask for form of address for user {username}")
            await ask_for_form_of_address(bot, callback_query.message.chat.id, user_id, username)

            print(f"Consent confirmed for user {username}")
        else:
            response = "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π –≤—ã–±–æ—Ä. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ –ø—Ä–µ–¥–ª–æ–∂–µ–Ω–Ω—ã—Ö –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤."
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(callback_query.message.chat.id, response)
            return

    except Exception as e:
        print(f"Error handling consent confirmation: {e}")


async def handle_form_of_address_choice(bot, callback_query, user_id, username):
    """Handle form of address (—Ç—ã/–í—ã) selection"""
    try:
        choice = callback_query.data

        if choice == "form_address:ty":
            form_of_address = '—Ç—ã'
        elif choice == "form_address:vy":
            form_of_address = '–í—ã'
        else:
            response = "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π –≤—ã–±–æ—Ä. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ –ø—Ä–µ–¥–ª–æ–∂–µ–Ω–Ω—ã—Ö –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤."
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(callback_query.message.chat.id, response)
            return

        # Save form of address to Excel
        save_form_of_address_to_excel(user_id, username, form_of_address)

        # Answer callback and ask for name
        await bot.answer_callback_query(callback_query.id)
        print(f"DEBUG: About to ask for name for user {username}")
        await ask_for_user_name(bot, callback_query.message.chat.id, user_id, username, form_of_address)

        print(f"Form of address registered for user {username}: {form_of_address}")

    except Exception as e:
        print(f"Error handling form of address choice: {e}")


async def handle_name_input(bot, message, user_id, username):
    """Handle user name input and send motivation message"""
    try:
        user_name = message.text.strip()

        if not user_name or len(user_name) < 1:
            if user_id in user_states and user_states[user_id]['form'] == '—Ç—ã':
                error_msg = "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏ —Å–≤–æ—ë –∏–º—è –∏–ª–∏ –Ω–∏–∫."
            else:
                error_msg = "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ —Å–≤–æ—ë –∏–º—è –∏–ª–∏ –Ω–∏–∫."
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(message.chat.id, error_msg, reply_markup=markup)
            return False

        # Get form of address from state
        form_of_address = user_states[user_id].get('form', '—Ç—ã')

        # Save user name to Excel
        save_user_name_to_excel(user_id, username, user_name)

        # Send motivation message
        await send_motivation_message(bot, message.chat.id, user_id, username, form_of_address, user_name)

        # Update state with user_name
        user_states[user_id] = {
            'stage': 'awaiting_motivation_response',
            'form': form_of_address,
            'user_name': user_name
        }

        print(f"User name saved and motivation message sent to {username}: {user_name}")
        return True

    except Exception as e:
        print(f"Error handling name input: {e}")
        return False


async def handle_ready_to_start(bot, callback_query, user_id, username):
    """Handle 'Ready to start?' button click - display how we work message and then protocol selection"""
    try:
        how_we_work_text = (
            "üéØ –ö–∞–∫ –º—ã —Ä–∞–±–æ—Ç–∞–µ–º:\n\n"
            "üìå –û–ø—Ä–µ–¥–µ–ª–∏–º —Ü–µ–ª—å –∏ –ø–æ–≤–µ—Å—Ç–∫—É\n\n"
            "üîç –ü–æ–¥–±–µ—Ä—ë–º —Ä–µ–ª–µ–≤–∞–Ω—Ç–Ω—ã–µ —É–ø—Ä–∞–∂–Ω–µ–Ω–∏—è\n\n"
            "üìä –ü—Ä–æ–π–¥—ë–º –∏—Ö –ø–æ—à–∞–≥–æ–≤–æ —Å –ø–æ–¥–¥–µ—Ä–∂–∫–æ–π\n\n"
            "üí¨ –í –∫–æ–Ω—Ü–µ ‚Äî –æ–±—Ä–∞—Ç–Ω–∞—è —Å–≤—è–∑—å, –ø–æ–¥—ã—Ç–æ–∂–∏–º –∏ (–ø–æ –∂–µ–ª–∞–Ω–∏—é) –≤–∫–ª—é—á–∏–º –Ω–∞–ø–æ–º–∏–Ω–∞–Ω–∏—è –¥–ª—è –∑–∞–∫—Ä–µ–ø–ª–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞\n\n"
            "–ú—ã –≤–º–µ—Å—Ç–µ —Ñ–æ—Ä–º—É–ª–∏—Ä—É–µ–º –∏ –ø—Ä–æ–≤–µ—Ä—è–µ–º –≥–∏–ø–æ—Ç–µ–∑—ã, –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º –æ–±—Ä–∞—Ç–Ω—É—é —Å–≤—è–∑—å, —Å—É–º–º–∏—Ä—É–µ–º —É—Å–ª—ã—à–∞–Ω–Ω–æ–µ; "
            "–ø—Ä–∏ —Ç–≤–æ—ë–º —Å–æ–≥–ª–∞—Å–∏–∏ –º–æ–∂–µ–º –≤–æ–≤–ª–µ—á—å –±–ª–∏–∑–∫–∏—Ö –¥–ª—è –ø–æ–¥–¥–µ—Ä–∂–∫–∏."
        )

        await bot.answer_callback_query(callback_query.id)
        await bot.send_message(callback_query.message.chat.id, how_we_work_text)

        # Small delay for better UX
        await asyncio.sleep(1)

        # Start goal setting process
        import goal
        await goal.start_goal_setting(bot, callback_query.message.chat.id, user_id, username)

        # Update state
        if user_id in user_states:
            user_states[user_id]['stage'] = 'started'

        print(f"'How we work' message sent to user {username}, protocol selection initiated")

    except Exception as e:
        print(f"Error handling ready to start: {e}")


def reset_user_greeting_state(user_id):
    """Reset user greeting state when /start is called again"""
    if user_id in user_states:
        del user_states[user_id]
    print(f"User greeting state reset for user ID: {user_id}")
