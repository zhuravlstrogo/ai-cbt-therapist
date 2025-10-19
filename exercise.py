# -*- coding: utf-8 -*-
"""
Exercise selection and execution module for AI Psychologist bot
Handles exercise recommendations based on selected problems
"""

import asyncio
import os
import re
from datetime import datetime
from telebot import types
from openpyxl import load_workbook, Workbook
from difflib import SequenceMatcher

# File paths
PROTOCOL_MAP_FILE = 'protocol_and_interventions_map.md'
INTERVENTIONS_FILE = 'interventions.md'
EXERCISES_EXCEL_FILE = 'exercises.xlsx'

# List of emojis for different exercises
EXERCISE_EMOJIS = [
    '✍️', '🧠', '📈', '💬', '🎯', '💪',
    '🌟', '📊', '🎨', '🔥', '💡', '🚀'
]

# Store user exercise states
user_exercise_states = {}


def init_exercises_excel():
    """Initialize exercises Excel file with headers"""
    if not os.path.exists(EXERCISES_EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Exercises'
        ws['A1'] = 'User ID'
        ws['B1'] = 'Username'
        ws['C1'] = 'Exercise Name'
        ws['D1'] = 'Problem'
        ws['E1'] = 'Problem Rating'
        ws['F1'] = 'Exercise Start Time'
        ws['G1'] = 'Step Number'
        ws['H1'] = 'Step Text'
        ws['I1'] = 'Step Result'
        ws['J1'] = 'Step Completion Time'
        ws['K1'] = 'Insight'
        ws['L1'] = 'What Was Useful'
        ws['M1'] = 'Difficulty'
        ws['N1'] = 'Date Time'
        wb.save(EXERCISES_EXCEL_FILE)


def save_exercise_selection_to_excel(user_id, username, exercise_name, problem, rating):
    """Save exercise selection to exercises.xlsx"""
    try:
        if not os.path.exists(EXERCISES_EXCEL_FILE):
            init_exercises_excel()

        wb = load_workbook(EXERCISES_EXCEL_FILE)
        ws = wb.active

        next_row = ws.max_row + 1

        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = exercise_name
        ws[f'D{next_row}'] = problem
        ws[f'E{next_row}'] = rating
        ws[f'F{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'H{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save(EXERCISES_EXCEL_FILE)
        print(f"Exercise selection saved: {username} - {exercise_name}")

    except Exception as e:
        print(f"Error saving exercise selection to Excel: {e}")


def extract_steps_from_description(full_description):
    """
    Extract individual steps from exercise full description
    Returns list of tuples: (step_number, step_text)
    Example: [(1, 'Предвестники (твои маркеры).'), (2, 'Опоры (что/кто помогает).')]
    """
    try:
        if not full_description:
            return []

        steps = []
        lines = full_description.split('\n')

        for line in lines:
            stripped = line.strip()
            # Look for lines starting with number and dot (e.g., "1.", "2.")
            if stripped and stripped[0].isdigit() and '.' in stripped[:3]:
                # Extract step number
                match = re.match(r'^(\d+)\.\s+(.+)', stripped)
                if match:
                    step_num = int(match.group(1))
                    step_text = match.group(2).strip()
                    if step_text:
                        steps.append((step_num, step_text))

        return steps

    except Exception as e:
        print(f"Error extracting steps: {e}")
        return []


def save_exercise_step_to_excel(user_id, username, exercise_name, problem, rating, step_num, step_text, step_result):
    """Save exercise step data to exercises.xlsx (all in one row)"""
    try:
        if not os.path.exists(EXERCISES_EXCEL_FILE):
            init_exercises_excel()

        wb = load_workbook(EXERCISES_EXCEL_FILE)
        ws = wb.active

        next_row = ws.max_row + 1

        ws[f'A{next_row}'] = user_id
        ws[f'B{next_row}'] = username
        ws[f'C{next_row}'] = exercise_name
        ws[f'D{next_row}'] = problem
        ws[f'E{next_row}'] = rating
        ws[f'F{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'G{next_row}'] = step_num
        ws[f'H{next_row}'] = step_text
        ws[f'I{next_row}'] = step_result
        ws[f'J{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'N{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wb.save(EXERCISES_EXCEL_FILE)
        print(f"Exercise step saved: {username} - {exercise_name} - Step {step_num}")

    except Exception as e:
        print(f"Error saving exercise step to Excel: {e}")


def save_exercise_final_answers_to_excel(user_id, username, exercise_name, problem, rating, insight, useful, difficulty):
    """Save final answers (insight, useful, difficulty) to exercises.xlsx"""
    try:
        if not os.path.exists(EXERCISES_EXCEL_FILE):
            init_exercises_excel()

        wb = load_workbook(EXERCISES_EXCEL_FILE)
        ws = wb.active

        # Find the last row for this user/exercise and update it
        for row in range(ws.max_row, 0, -1):
            if (ws[f'A{row}'].value == user_id and
                ws[f'C{row}'].value == exercise_name):
                ws[f'K{row}'] = insight
                ws[f'L{row}'] = useful
                ws[f'M{row}'] = difficulty
                break

        wb.save(EXERCISES_EXCEL_FILE)
        print(f"Exercise final answers saved: {username} - {exercise_name}")

    except Exception as e:
        print(f"Error saving exercise final answers: {e}")


def save_exercise_text_to_excel(user_id, username, exercise_name, exercise_text):
    """Save exercise text input to exercises.xlsx"""
    try:
        if not os.path.exists(EXERCISES_EXCEL_FILE):
            init_exercises_excel()

        wb = load_workbook(EXERCISES_EXCEL_FILE)
        ws = wb.active

        for row in range(ws.max_row, 0, -1):
            if ws[f'A{row}'].value == user_id and ws[f'C{row}'].value == exercise_name:
                ws[f'G{row}'] = exercise_text
                break

        wb.save(EXERCISES_EXCEL_FILE)
        print(f"Exercise text saved: {username} - {exercise_name}")

    except Exception as e:
        print(f"Error saving exercise text to Excel: {e}")


def extract_exercises_for_problem(problem_name):
    """
    Extract exercises for a given problem from protocol_and_interventions_map.md
    """
    try:
        if not os.path.exists(PROTOCOL_MAP_FILE):
            print(f"Error: {PROTOCOL_MAP_FILE} not found")
            return []

        with open(PROTOCOL_MAP_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        lines = content.split('\n')
        problem_section_start = None

        for idx, line in enumerate(lines):
            if line.startswith('###') and problem_name in line:
                problem_section_start = idx
                break

        if problem_section_start is None:
            print(f"Problem '{problem_name}' not found in {PROTOCOL_MAP_FILE}")
            return []

        exercises = []
        for idx in range(problem_section_start + 1, len(lines)):
            line = lines[idx].strip()

            if line.startswith('###'):
                break

            if line.startswith('*'):
                exercise_text = line.lstrip('*').strip()
                exercise_text = re.sub(r'\s+', ' ', exercise_text)
                if exercise_text and any(c.isalpha() for c in exercise_text):
                    exercises.append(exercise_text)

        return exercises[:6]

    except Exception as e:
        print(f"Error extracting exercises: {e}")
        return []


def extract_exercise_goal(exercise_name):
    """
    Extract exercise goal from interventions.md using fuzzy matching
    """
    try:
        if not os.path.exists(INTERVENTIONS_FILE):
            print(f"Error: {INTERVENTIONS_FILE} not found")
            return None

        with open(INTERVENTIONS_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        lines = content.split('\n')

        # Handle "Exercise · Other" format - take only first part
        if '·' in exercise_name:
            exercise_name = exercise_name.split('·')[0].strip()

        search_term = exercise_name.split('(')[0].strip()
        search_term = search_term.rstrip('.!?,;:')
        search_term_lower = search_term.lower()

        # Find best match using fuzzy matching
        best_match_idx = None
        best_match_score = 0
        MATCH_THRESHOLD = 0.8  # 80% similarity threshold

        # Search only in section headers (lines starting with ##)
        for idx, line in enumerate(lines):
            if line.startswith('##'):
                line_clean = line.replace('##', '').strip()
                # Remove number with either . or ) after it
                line_clean = re.sub(r'^[^\w\u0400-\u04FF]*\d+[\.)\]]\s*', '', line_clean)

                # Check both the full name and abbreviation in parentheses
                # Extract main part (before parentheses)
                main_part = line_clean.split('(')[0].strip()

                # Extract abbreviation if exists (e.g., PST from "(PST)")
                abbreviation = None
                if '(' in line_clean and ')' in line_clean:
                    # Extract content between parentheses
                    abbreviation_match = re.search(r'\(([^)]+)\)', line_clean)
                    if abbreviation_match:
                        abbreviation = abbreviation_match.group(1).strip()

                # Calculate similarity score for main part
                score_main = SequenceMatcher(None, search_term_lower, main_part.lower()).ratio()

                # Calculate similarity score for abbreviation if it exists
                score_abbr = 0
                if abbreviation:
                    score_abbr = SequenceMatcher(None, search_term_lower, abbreviation.lower()).ratio()

                # Use the better score
                score = max(score_main, score_abbr)

                # Update best match if score is better
                if score > best_match_score:
                    best_match_score = score
                    best_match_idx = idx

                    # If we find an exact match, stop searching
                    if score == 1.0:
                        break

        # Use the best match if it meets the threshold
        if best_match_idx is not None and best_match_score >= MATCH_THRESHOLD:
            exercise_section_idx = best_match_idx
            print(f"Found '{search_term}' with score {best_match_score:.2f} at line {exercise_section_idx}")
        else:
            print(f"Exercise '{search_term}' not found in {INTERVENTIONS_FILE} (best score: {best_match_score:.2f})")
            return None

        # Extract goal from the found section
        for idx in range(exercise_section_idx, min(exercise_section_idx + 10, len(lines))):
            if 'Цель:' in lines[idx]:
                goal_text = lines[idx].replace('Цель:', '').strip()
                # Remove time information if present
                goal_text = re.sub(r'\s*Время:\s*\d+–\d+\s*мин\.?', '', goal_text)
                goal_text = re.sub(r'\s*Время:\s*[\d\w\s–]+\.?$', '', goal_text)
                return goal_text.strip()

        return None

    except Exception as e:
        print(f"Error extracting exercise goal: {e}")
        return None


def extract_exercise_full_description(exercise_name):
    """
    Extract full exercise description from interventions.md using fuzzy matching
    Returns all text from the exercise section until the next section marker
    """
    try:
        if not os.path.exists(INTERVENTIONS_FILE):
            print(f"Error: {INTERVENTIONS_FILE} not found")
            return None

        with open(INTERVENTIONS_FILE, 'r', encoding='utf-8') as f:
            content = f.read()

        lines = content.split('\n')

        # Handle "Exercise · Other" format - take only first part
        if '·' in exercise_name:
            exercise_name = exercise_name.split('·')[0].strip()

        search_term = exercise_name.split('(')[0].strip()
        search_term = search_term.rstrip('.!?,;:')
        search_term_lower = search_term.lower()

        # Find best match using fuzzy matching
        best_match_idx = None
        best_match_score = 0
        MATCH_THRESHOLD = 0.8  # 80% similarity threshold

        # Find exercise section header (##) using fuzzy matching
        for idx, line in enumerate(lines):
            if line.startswith('##'):
                line_clean = line.replace('##', '').strip()
                # Remove number with either . or ) after it
                line_clean = re.sub(r'^[^\w\u0400-\u04FF]*\d+[\.)\]]\s*', '', line_clean)

                # Check both the full name and abbreviation in parentheses
                # Extract main part (before parentheses)
                main_part = line_clean.split('(')[0].strip()

                # Extract abbreviation if exists (e.g., PST from "(PST)")
                abbreviation = None
                if '(' in line_clean and ')' in line_clean:
                    # Extract content between parentheses
                    abbreviation_match = re.search(r'\(([^)]+)\)', line_clean)
                    if abbreviation_match:
                        abbreviation = abbreviation_match.group(1).strip()

                # Calculate similarity score for main part
                score_main = SequenceMatcher(None, search_term_lower, main_part.lower()).ratio()

                # Calculate similarity score for abbreviation if it exists
                score_abbr = 0
                if abbreviation:
                    score_abbr = SequenceMatcher(None, search_term_lower, abbreviation.lower()).ratio()

                # Use the better score
                score = max(score_main, score_abbr)

                # Update best match if score is better
                if score > best_match_score:
                    best_match_score = score
                    best_match_idx = idx

                    # If we find an exact match, stop searching
                    if score == 1.0:
                        break

        # Use the best match if it meets the threshold
        if best_match_idx is not None and best_match_score >= MATCH_THRESHOLD:
            exercise_section_idx = best_match_idx
            print(f"Found '{search_term}' with score {best_match_score:.2f} at line {exercise_section_idx}")
        else:
            print(f"Exercise '{search_term}' not found in {INTERVENTIONS_FILE} (best score: {best_match_score:.2f})")
            return None

        # Extract content from exercise section until next section marker (*** or ##)
        description_lines = []
        skip_empty_header = True  # Flag to skip the empty ## header that comes after title

        for idx in range(exercise_section_idx + 1, len(lines)):
            line = lines[idx]

            # Skip the empty "##" header that comes right after the title
            if skip_empty_header and line.strip() == '##':
                skip_empty_header = False
                continue

            # Stop at next section marker (but not empty ##)
            if line.strip().startswith('***'):
                break

            # Stop at next numbered section header (## followed by number)
            if line.strip().startswith('##') and len(line.strip()) > 2:
                # Check if it's a numbered section (has digit after ##)
                header_content = line.replace('##', '').strip()
                if header_content and (header_content[0].isdigit() or header_content.startswith('0)')):
                    break

            # Skip empty lines at the beginning
            if not description_lines and not line.strip():
                continue

            description_lines.append(line)

        # Join and clean up
        description = '\n'.join(description_lines).strip()

        # Remove trailing empty lines
        while description.endswith('\n\n'):
            description = description[:-1]

        return description if description else None

    except Exception as e:
        print(f"Error extracting exercise full description: {e}")
        return None


async def show_exercise_recommendations(bot, chat_id, user_id, username, problems_with_ratings):
    """
    Show exercise recommendations based on selected problems
    """
    try:
        # Try to find exercises for any of the selected problems
        exercises = []
        problem_used = None

        for problem in problems_with_ratings.keys():
            exercises = extract_exercises_for_problem(problem)
            if exercises:
                problem_used = problem
                break

        if not exercises:
            # No exercises found for any problem - show message with menu button
            markup = types.InlineKeyboardMarkup()
            btn_menu = types.InlineKeyboardButton(
                "📱 Главное меню",
                callback_data="menu:show"
            )
            markup.add(btn_menu)

            # Format problem names for display
            problem_names = list(problems_with_ratings.keys())
            if len(problem_names) == 1:
                problem_text = f"проблемы '{problem_names[0]}'"
            else:
                problem_text = "выбранных проблем:\n" + "\n".join([f"• {p}" for p in problem_names])

            await bot.send_message(
                chat_id,
                f"К сожалению, не удалось найти упражнения для {problem_text}.\n\n"
                "Пока мы работаем над упражнениями для этой проблемы, но ты можешь вернуться в главное меню и попробовать другие функции.",
                reply_markup=markup
            )
            return

        user_exercise_states[user_id] = {
            'exercises': exercises,
            'problems': problems_with_ratings,
            'selected_exercise': None,
            'username': username,
            'completed_exercises': []  # Track completed exercises
        }

        header_text = "На основе твоих ответов рекомендую начать с этих упражнений:"
        await bot.send_message(chat_id, header_text)

        # Pause for 2 seconds
        await asyncio.sleep(2)

        for idx, exercise in enumerate(exercises):
            goal = extract_exercise_goal(exercise)
            emoji = EXERCISE_EMOJIS[idx % len(EXERCISE_EMOJIS)]

            # Remove "Время: X–Y мин." from goal
            if goal:
                # Remove the time part (e.g., "Время: 5–8 мин.")
                goal_clean = re.sub(r'\s*·\s*Время:\s*\d+–\d+\s*мин\.', '', goal)
                goal_clean = goal_clean.strip()
                card_text = f"{emoji} {exercise}\n{goal_clean}" if goal_clean else f"{emoji} {exercise}"
            else:
                card_text = f"{emoji} {exercise}"

            markup = types.InlineKeyboardMarkup()
            btn_select = types.InlineKeyboardButton(
                "Выбрать",
                callback_data=f"ex_select:{idx}"
            )
            markup.add(btn_select)

            await bot.send_message(chat_id, card_text, reply_markup=markup)

        # After all exercise cards, add menu button
        from universal_menu import get_menu_button
        menu_markup = get_menu_button()
        await bot.send_message(chat_id, "Выбери упражнение или вернись в меню", reply_markup=menu_markup)

    except Exception as e:
        print(f"Error showing exercise recommendations: {e}")


async def handle_exercise_select(bot, callback_query, exercise_idx):
    """
    Handle exercise selection
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        exercise_idx = int(exercise_idx)

        if exercise_idx >= len(state['exercises']):
            await bot.answer_callback_query(callback_query.id)
            return

        selected_exercise = state['exercises'][exercise_idx]
        state['selected_exercise'] = selected_exercise

        # Reset exercise execution state for new exercise
        state['steps'] = []
        state['current_step_idx'] = 0
        state['step_results'] = {}
        state['current_final_question'] = 0
        state['final_answers'] = {}
        state['awaiting_exercise_text'] = False
        state['awaiting_step_input'] = False
        state['awaiting_final_answer'] = False

        first_problem = list(state['problems'].keys())[0]
        first_rating = state['problems'][first_problem]
        save_exercise_selection_to_excel(user_id, username, selected_exercise, first_problem, first_rating)

        await bot.answer_callback_query(callback_query.id)

        markup = types.InlineKeyboardMarkup()

        btn_start = types.InlineKeyboardButton(
            "▶️ Начать это упражнение",
            callback_data="ex_start_exec"
        )
        btn_change = types.InlineKeyboardButton(
            "✏️ Изменить выбор",
            callback_data="ex_change_select"
        )
        btn_menu = types.InlineKeyboardButton(
            "📍 Вернуться в главное меню",
            callback_data="menu:show"
        )

        markup.add(btn_start)
        markup.add(btn_change)
        markup.add(btn_menu)

        nav_text = f"Выбранное упражнение: {selected_exercise}"
        await bot.send_message(chat_id, nav_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling exercise selection: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_exercise_start(bot, callback_query):
    """
    Handle exercise start - show steps one by one
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        selected_exercise = state['selected_exercise']

        if not selected_exercise:
            await bot.answer_callback_query(callback_query.id)
            return

        # Extract full description from interventions.md
        full_description = extract_exercise_full_description(selected_exercise)

        # Extract steps from full description
        steps = extract_steps_from_description(full_description)

        if not steps:
            # No steps found, show full description as before
            await bot.answer_callback_query(callback_query.id)
            exercise_idx = state['exercises'].index(selected_exercise)
            emoji = EXERCISE_EMOJIS[exercise_idx % len(EXERCISE_EMOJIS)]

            if full_description:
                exercise_text = f"{emoji} {selected_exercise}\n\n{full_description}"
            else:
                exercise_text = f"{emoji} {selected_exercise}"

            await bot.send_message(chat_id, exercise_text)
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(chat_id, "Поделись результатом упражнения:", reply_markup=markup)
            state['awaiting_exercise_text'] = True
            return

        # Store steps and initialize step navigation
        state['steps'] = steps
        state['current_step_idx'] = 0
        state['step_results'] = {}

        await bot.answer_callback_query(callback_query.id)

        # Show exercise header
        exercise_idx = state['exercises'].index(selected_exercise)
        emoji = EXERCISE_EMOJIS[exercise_idx % len(EXERCISE_EMOJIS)]
        await bot.send_message(chat_id, f"{emoji} {selected_exercise}")

        # Show first step
        await show_exercise_step(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error handling exercise start: {e}")
        await bot.answer_callback_query(callback_query.id)


async def show_exercise_step(bot, chat_id, user_id):
    """
    Show current exercise step
    """
    try:
        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]
        current_idx = state['current_step_idx']
        steps = state['steps']

        if current_idx >= len(steps):
            # All steps completed - show final questions
            await show_final_questions(bot, chat_id, user_id)
            return

        step_num, step_text = steps[current_idx]

        # Show step
        step_message = f"Шаг {current_idx + 1} из {len(steps)}:\n\n{step_text}"
        await bot.send_message(chat_id, step_message)

        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(chat_id, "Поделись результатом для этого шага:", reply_markup=markup)

        # Mark that we're awaiting step input
        state['awaiting_step_input'] = True

    except Exception as e:
        print(f"Error showing exercise step: {e}")


async def show_final_questions(bot, chat_id, user_id):
    """
    Show final questions after all steps
    """
    try:
        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]
        
        # Check for crisis indicators in all step results
        from safety_check import check_text_safety, show_crisis_support, log_crisis_detection
        from greeting import user_states
        
        # Combine all step results for safety check
        all_step_results = []
        for step_result in state.get('step_results', {}).values():
            if step_result:
                all_step_results.append(step_result)
        
        # Also check pending step result if exists
        if state.get('pending_step_result'):
            all_step_results.append(state['pending_step_result'])
        
        if all_step_results:
            combined_text = " ".join(all_step_results)
            
            crisis_detected, crisis_type, confidence = await check_text_safety(
                text=combined_text,
                context="exercise"
            )
            
            if crisis_detected and crisis_type:
                # Log crisis detection
                username = state.get('username', 'Unknown')
                await log_crisis_detection(
                    user_id=user_id,
                    username=username,
                    crisis_type=crisis_type,
                    context="exercise",
                    text_sample=combined_text[:200],
                    file_path='exercises.xlsx'
                )
                
                # Get user name
                user_name = 'Друг'
                if user_id in user_states:
                    user_name = user_states[user_id].get('user_name', 'Друг')
                
                # Show crisis support
                await show_crisis_support(
                    bot=bot,
                    chat_id=chat_id,
                    user_name=user_name,
                    crisis_type=crisis_type,
                    context="exercise",
                    continue_after=True  # Allow continuing with final questions
                )
                return
        
        state['current_final_question'] = 0
        state['final_answers'] = {}

        # Show first question
        await show_final_question(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error showing final questions: {e}")


async def show_final_question(bot, chat_id, user_id):
    """
    Show current final question
    """
    try:
        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]
        question_idx = state['current_final_question']

        questions = [
            "Какой инсайт ты получил?",
            "Что было полезно?",
            "Что вызвало трудность?"
        ]

        if question_idx >= len(questions):
            # All questions answered - show completion options
            await show_exercise_completion_options(bot, chat_id, user_id)
            return

        question = questions[question_idx]
        from universal_menu import get_menu_button
        markup = get_menu_button()
        await bot.send_message(chat_id, question, reply_markup=markup)

        state['awaiting_final_answer'] = True

    except Exception as e:
        print(f"Error showing final question: {e}")


async def show_exercise_completion_options(bot, chat_id, user_id):
    """
    Show completion options after all final questions are answered
    Options: Mark as completed → Next exercise or Main menu
    """
    try:
        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]

        # Show button to mark as completed
        markup = types.InlineKeyboardMarkup()

        btn_mark_complete = types.InlineKeyboardButton(
            "✅ Отметить как завершённое",
            callback_data="ex_mark_complete"
        )

        markup.add(btn_mark_complete)
        await bot.send_message(chat_id, "Отлично! Ты выполнила(а) упражнение.", reply_markup=markup)

    except Exception as e:
        print(f"Error showing exercise completion options: {e}")


async def finish_exercise(bot, chat_id, user_id):
    """
    Finish exercise and save answers, then show next exercise options
    """
    try:
        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]
        username = state.get('username', 'Unknown')
        selected_exercise = state['selected_exercise']
        first_problem = list(state['problems'].keys())[0]
        first_rating = state['problems'][first_problem]

        insight = state['final_answers'].get(0, '')
        useful = state['final_answers'].get(1, '')
        difficulty = state['final_answers'].get(2, '')

        # Save final answers
        save_exercise_final_answers_to_excel(user_id, username, selected_exercise, first_problem, first_rating, insight, useful, difficulty)

        # Check for crisis indicators in final answers
        from safety_check import check_text_safety, show_crisis_support, log_crisis_detection
        from greeting import user_states

        # Combine all final answers for safety check
        all_answers = f"{insight} {useful} {difficulty}"

        # Get user name
        user_name = 'Друг'
        if user_id in user_states:
            user_name = user_states[user_id].get('user_name', 'Друг')

        crisis_detected, crisis_type, confidence = await check_text_safety(
            text=all_answers,
            context="exercise"
        )

        if crisis_detected and crisis_type:
            # Log crisis detection
            await log_crisis_detection(
                user_id=user_id,
                username=username,
                crisis_type=crisis_type,
                context="exercise",
                text_sample=all_answers[:200],
                file_path='exercises.xlsx'
            )

            # Show completion message first
            text = "Спасибо! Я записал(а) твой опыт. Это отличная работа! 💪"
            await bot.send_message(chat_id, text)

            # Then show crisis support
            await show_crisis_support(
                bot=bot,
                chat_id=chat_id,
                user_name=user_name,
                crisis_type=crisis_type,
                context="exercise",
                continue_after=True  # Allow continuing to next exercise
            )
        else:
            # No crisis - show normal completion and next options
            text = "Спасибо! Я записал(а) твой опыт. Это отличная работа! 💪"
            await bot.send_message(chat_id, text)

            # Show next exercise options
            await show_next_exercise_options(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error finishing exercise: {e}")


async def show_next_exercise_options(bot, chat_id, user_id):
    """
    Show all remaining exercises for the problem after completing one
    """
    try:
        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]
        exercises = state.get('exercises', [])
        current_exercise = state.get('selected_exercise')

        # Track completed exercises in state
        if 'completed_exercises' not in state:
            state['completed_exercises'] = []

        # Mark current exercise as completed
        if current_exercise and current_exercise not in state['completed_exercises']:
            state['completed_exercises'].append(current_exercise)

        # Find remaining exercises
        remaining_exercises = [ex for ex in exercises if ex not in state['completed_exercises']]

        if remaining_exercises:
            # Show remaining exercises
            await bot.send_message(
                chat_id,
                "Отлично! ✨ Вот другие упражнения, которые могут помочь:"
            )

            # Display each remaining exercise with selection button
            for idx, exercise in enumerate(exercises):
                if exercise not in state['completed_exercises']:
                    # Get exercise goal
                    goal = extract_exercise_goal(exercise)
                    emoji = EXERCISE_EMOJIS[exercises.index(exercise) % len(EXERCISE_EMOJIS)]

                    # Create card text
                    if goal:
                        goal_clean = re.sub(r'\s*·\s*Время:\s*\d+–\d+\s*мин\.', '', goal)
                        goal_clean = goal_clean.strip()
                        card_text = f"{emoji} {exercise}\n{goal_clean}" if goal_clean else f"{emoji} {exercise}"
                    else:
                        card_text = f"{emoji} {exercise}"

                    # Create button
                    markup = types.InlineKeyboardMarkup()
                    btn_select = types.InlineKeyboardButton(
                        "Выбрать",
                        callback_data=f"ex_select:{exercises.index(exercise)}"
                    )
                    markup.add(btn_select)

                    await bot.send_message(chat_id, card_text, reply_markup=markup)

            # Add final menu button
            markup_final = types.InlineKeyboardMarkup()
            btn_menu = types.InlineKeyboardButton(
                "📍 Главное меню",
                callback_data="menu:show"
            )
            markup_final.add(btn_menu)
            await bot.send_message(
                chat_id,
                "Или вернись в главное меню:",
                reply_markup=markup_final
            )
        else:
            # All exercises completed
            markup = types.InlineKeyboardMarkup()
            btn_menu = types.InlineKeyboardButton(
                "📍 Главное меню",
                callback_data="menu:show"
            )
            markup.add(btn_menu)

            await bot.send_message(
                chat_id,
                "Поздравляю! 🎉 Ты выполнил(а) все рекомендуемые упражнения для этой проблемы!",
                reply_markup=markup
            )

        # Don't delete state yet - user might select next exercise
        # State will be cleared when user returns to menu or selects new exercise

    except Exception as e:
        print(f"Error showing next exercise options: {e}")


def validate_exercise_text(text):
    """
    Validate exercise text input.
    Returns: (is_valid, feedback_message)
    """
    text = text.strip()
    words = text.split()

    if len(words) == 1:
        if text.lower() in ['не', 'нет', 'да', 'не знаю']:
            return False, "Понимаю, что может быть сложновато! 🤝 Но давай разберёмся вместе. Расскажи, хотя бы:\n• Что было сложным?\n• Что заметил(а) во время упражнения?\n• Может, какой-то момент выделился?"
        else:
            return False, "Спасибо за ответ! 🙏 Но давай углубимся. Расскажи подробнее:\n• Что делал(а)?\n• Что почувствовал(а)?\n• Какие выводы?"

    meaningful_chars = sum(1 for c in text if c.isalpha())
    if meaningful_chars < 10:
        return False, "Твой ответ кажется очень коротким 📝 Давай расширим:\n• Как прошло упражнение?\n• Что изменилось в ощущениях?\n• Есть ли результат?"

    return True, None


async def handle_exercise_text_input(bot, message):
    """
    Handle exercise text input (for steps and final answers)
    """
    try:
        user_id = message.from_user.id
        username = message.from_user.username or 'Unknown'
        text = message.text

        if user_id not in user_exercise_states:
            return

        state = user_exercise_states[user_id]

        # Check if awaiting step input
        if state.get('awaiting_step_input'):
            await handle_step_input(bot, message, user_id, username, text, state)
            return

        # Check if awaiting final answer
        if state.get('awaiting_final_answer'):
            await handle_final_answer_input(bot, message, user_id, username, text, state)
            return

        # Legacy: handle exercise text input (backward compatibility)
        if state.get('awaiting_exercise_text'):
            selected_exercise = state['selected_exercise']

            # Validate input
            is_valid, feedback = validate_exercise_text(text)

            if not is_valid:
                await bot.send_message(message.chat.id, feedback)
                return

            # Store text temporarily
            state['pending_exercise_text'] = text

            # Show preview
            preview_text = f"📝 Вот что ты написал(а):\n\n{text}\n\nВсё верно?"

            markup = types.InlineKeyboardMarkup()

            btn_confirm = types.InlineKeyboardButton(
                "✅ Подтвердить",
                callback_data="ex_text_confirm:yes"
            )
            btn_edit = types.InlineKeyboardButton(
                "✏️ Изменить",
                callback_data="ex_text_confirm:edit"
            )
            btn_back = types.InlineKeyboardButton(
                "⬅️ Предыдущий шаг",
                callback_data="ex_text_confirm:back"
            )
            btn_menu = types.InlineKeyboardButton(
                "📍 Главное меню",
                callback_data="menu:show"
            )

            markup.add(btn_confirm)
            markup.add(btn_edit)
            markup.add(btn_back)
            markup.add(btn_menu)

            await bot.send_message(message.chat.id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling exercise text input: {e}")


async def handle_step_input(bot, message, user_id, username, text, state):
    """
    Handle step input during exercise execution
    """
    try:
        # Validate input
        is_valid, feedback = validate_exercise_text(text)

        if not is_valid:
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(message.chat.id, feedback, reply_markup=markup)
            return

        # Check for crisis indicators in step input
        from safety_check import check_text_safety, show_crisis_support, log_crisis_detection
        from greeting import user_states

        crisis_detected, crisis_type, confidence = await check_text_safety(
            text=text,
            context="exercise"
        )

        if crisis_detected and crisis_type:
            # Log crisis detection
            await log_crisis_detection(
                user_id=user_id,
                username=username,
                crisis_type=crisis_type,
                context="exercise",
                text_sample=text[:200],
                file_path='exercises.xlsx'
            )

            # Get user name
            user_name = 'Друг'
            if user_id in user_states:
                user_name = user_states[user_id].get('user_name', 'Друг')

            # Show crisis support
            await show_crisis_support(
                bot=bot,
                chat_id=message.chat.id,
                user_name=user_name,
                crisis_type=crisis_type,
                context="exercise",
                continue_after=True  # Allow continuing with exercise
            )
            return

        # Store step result temporarily
        state['pending_step_result'] = text

        # Show preview
        preview_text = f"📝 Вот что ты написал(а):\n\n{text}\n\nВсё верно?"

        markup = types.InlineKeyboardMarkup()

        btn_confirm = types.InlineKeyboardButton(
            "✅ Подтвердить",
            callback_data="ex_step_confirm:yes"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="ex_step_confirm:edit"
        )
        btn_back = types.InlineKeyboardButton(
            "⬅️ Вернуться",
            callback_data="ex_step_confirm:back"
        )
        btn_menu = types.InlineKeyboardButton(
            "📍 Меню",
            callback_data="menu:show"
        )

        markup.add(btn_confirm)
        markup.add(btn_edit)
        markup.add(btn_back)
        markup.add(btn_menu)

        await bot.send_message(message.chat.id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling step input: {e}")


async def handle_final_answer_input(bot, message, user_id, username, text, state):
    """
    Handle final answer input (for insight, useful, difficulty questions)
    """
    try:
        # Validate input
        is_valid, feedback = validate_exercise_text(text)

        if not is_valid:
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(message.chat.id, feedback, reply_markup=markup)
            return

        # Check for crisis indicators in final answer
        from safety_check import check_text_safety, show_crisis_support, log_crisis_detection
        from greeting import user_states

        crisis_detected, crisis_type, confidence = await check_text_safety(
            text=text,
            context="exercise"
        )

        if crisis_detected and crisis_type:
            # Log crisis detection
            await log_crisis_detection(
                user_id=user_id,
                username=username,
                crisis_type=crisis_type,
                context="exercise",
                text_sample=text[:200],
                file_path='exercises.xlsx'
            )

            # Get user name
            user_name = 'Друг'
            if user_id in user_states:
                user_name = user_states[user_id].get('user_name', 'Друг')

            # Show crisis support
            await show_crisis_support(
                bot=bot,
                chat_id=message.chat.id,
                user_name=user_name,
                crisis_type=crisis_type,
                context="exercise",
                continue_after=True  # Allow continuing with final questions
            )
            return

        # Store answer temporarily
        state['pending_final_answer'] = text

        # Show preview
        preview_text = f"📝 Вот что ты написал(а):\n\n{text}\n\nВсё верно?"

        markup = types.InlineKeyboardMarkup()

        btn_confirm = types.InlineKeyboardButton(
            "✅ Подтвердить",
            callback_data="ex_answer_confirm:yes"
        )
        btn_edit = types.InlineKeyboardButton(
            "✏️ Изменить",
            callback_data="ex_answer_confirm:edit"
        )
        btn_menu = types.InlineKeyboardButton(
            "📍 Меню",
            callback_data="menu:show"
        )

        markup.add(btn_confirm)
        markup.add(btn_edit)
        markup.add(btn_menu)

        await bot.send_message(message.chat.id, preview_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling final answer input: {e}")


async def handle_exercise_text_confirm(bot, callback_query, action):
    """
    Handle exercise text confirmation
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        pending_text = state.get('pending_exercise_text')

        if action == "yes":
            # Save the text
            selected_exercise = state['selected_exercise']
            save_exercise_text_to_excel(user_id, username, selected_exercise, pending_text)

            await bot.answer_callback_query(callback_query.id, "Спасибо! Сохранено.")

            # Show completion message
            await bot.send_message(
                chat_id,
                "Спасибо! Я записал(а) твой опыт. Это отличная работа! 💪"
            )

            # Show main menu
            from universal_menu import show_main_menu
            from greeting import user_states

            user_name = 'User'
            form_of_address = 'ты'
            if user_id in user_states:
                user_name = user_states[user_id].get('user_name', 'User')
                form_of_address = user_states[user_id].get('form', 'ты')

            del user_exercise_states[user_id]

            await show_main_menu(bot, chat_id, user_id, username, user_name, form_of_address)

        elif action == "edit":
            # Ask to re-enter
            state['pending_exercise_text'] = None
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                chat_id,
                "Окей, введи свой ответ заново:"
            )

        elif action == "back":
            # Go back to exercise selection
            await bot.answer_callback_query(callback_query.id)
            del user_exercise_states[user_id]

            from universal_menu import show_main_menu
            from greeting import user_states

            user_name = 'User'
            form_of_address = 'ты'
            if user_id in user_states:
                user_name = user_states[user_id].get('user_name', 'User')
                form_of_address = user_states[user_id].get('form', 'ты')

            await show_main_menu(bot, chat_id, user_id, username, user_name, form_of_address)

    except Exception as e:
        print(f"Error handling exercise text confirm: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_step_confirm(bot, callback_query, action):
    """
    Handle step confirmation during exercise execution
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        pending_result = state.get('pending_step_result')

        if action == "yes":
            # Save the step result to Excel
            current_idx = state['current_step_idx']
            steps = state['steps']
            step_num, step_text = steps[current_idx]

            selected_exercise = state['selected_exercise']
            first_problem = list(state['problems'].keys())[0]
            first_rating = state['problems'][first_problem]

            save_exercise_step_to_excel(
                user_id, username, selected_exercise, first_problem, first_rating,
                step_num, step_text, pending_result
            )

            # Store step result for safety checking
            state['step_results'][current_idx] = pending_result

            await bot.answer_callback_query(callback_query.id, "Спасибо! Сохранено.")

            # Move to next step
            state['current_step_idx'] += 1
            state['awaiting_step_input'] = False
            state['pending_step_result'] = None

            # Show next step or final questions
            await show_exercise_step(bot, chat_id, user_id)

        elif action == "edit":
            # Ask to re-enter
            state['pending_step_result'] = None
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                chat_id,
                "Окей, введи свой ответ заново:"
            )

        elif action == "back":
            # Go back to previous step
            if state['current_step_idx'] > 0:
                state['current_step_idx'] -= 1
                state['awaiting_step_input'] = False
                state['pending_step_result'] = None

                await bot.answer_callback_query(callback_query.id)
                await show_exercise_step(bot, chat_id, user_id)
            else:
                # No previous step - return to exercise selection
                await bot.answer_callback_query(callback_query.id)
                state['selected_exercise'] = None
                state['awaiting_step_input'] = False

                # Re-show exercise recommendations
                from greeting import user_states
                user_name = user_states.get(user_id, {}).get('user_name', 'User')

                header_text = "Выбери другое упражнение:"
                await bot.send_message(chat_id, header_text)

                for idx, exercise in enumerate(state['exercises']):
                    goal = extract_exercise_goal(exercise)
                    emoji = EXERCISE_EMOJIS[idx % len(EXERCISE_EMOJIS)]

                    if goal:
                        card_text = f"{emoji} {exercise}\n{goal}"
                    else:
                        card_text = f"{emoji} {exercise}"

                    markup = types.InlineKeyboardMarkup()
                    btn_select = types.InlineKeyboardButton(
                        "Выбрать",
                        callback_data=f"ex_select:{idx}"
                    )
                    markup.add(btn_select)

                    await bot.send_message(chat_id, card_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling step confirm: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_answer_confirm(bot, callback_query, action):
    """
    Handle final answer confirmation
    """
    try:
        user_id = callback_query.from_user.id
        username = callback_query.from_user.username or 'Unknown'
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        pending_answer = state.get('pending_final_answer')
        question_idx = state['current_final_question']

        if action == "yes":
            # Save the answer
            state['final_answers'][question_idx] = pending_answer

            await bot.answer_callback_query(callback_query.id, "Спасибо! Записано.")

            # Move to next final question
            state['current_final_question'] += 1
            state['awaiting_final_answer'] = False
            state['pending_final_answer'] = None

            # Show next question or finish
            await show_final_question(bot, chat_id, user_id)

        elif action == "edit":
            # Ask to re-enter
            state['pending_final_answer'] = None
            await bot.answer_callback_query(callback_query.id)
            await bot.send_message(
                chat_id,
                "Окей, введи свой ответ заново:"
            )

    except Exception as e:
        print(f"Error handling answer confirm: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_mark_exercise_complete(bot, callback_query):
    """
    Handle marking exercise as completed
    After this, show next exercise options
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        await bot.answer_callback_query(callback_query.id, "✅ Упражнение завершено!")

        # Finish the exercise (save data)
        await finish_exercise(bot, chat_id, user_id)

    except Exception as e:
        print(f"Error marking exercise complete: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_exercise_change_select(bot, callback_query):
    """
    Handle going back to exercise selection
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        state['selected_exercise'] = None

        await bot.answer_callback_query(callback_query.id)

        header_text = "Выбери другое упражнение:"
        await bot.send_message(chat_id, header_text)

        for idx, exercise in enumerate(state['exercises']):
            goal = extract_exercise_goal(exercise)
            emoji = EXERCISE_EMOJIS[idx % len(EXERCISE_EMOJIS)]

            if goal:
                card_text = f"{emoji} {exercise}\n{goal}"
            else:
                card_text = f"{emoji} {exercise}"

            markup = types.InlineKeyboardMarkup()
            btn_select = types.InlineKeyboardButton(
                "Выбрать",
                callback_data=f"ex_select:{idx}"
            )
            markup.add(btn_select)

            await bot.send_message(chat_id, card_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling exercise change selection: {e}")
        await bot.answer_callback_query(callback_query.id)


async def handle_exercise_continue_after_safety(bot, callback_query):
    """
    Handle continuing exercise after safety check
    """
    try:
        user_id = callback_query.from_user.id
        chat_id = callback_query.message.chat.id

        if user_id not in user_exercise_states:
            await bot.answer_callback_query(callback_query.id)
            return

        state = user_exercise_states[user_id]
        await bot.answer_callback_query(callback_query.id)

        # Determine where to continue based on current state
        if state.get('awaiting_step_input'):
            # Continue with current step
            current_idx = state.get('current_step_idx', 0)
            steps = state.get('steps', [])
            if current_idx < len(steps):
                step_num, step_text = steps[current_idx]
                step_message = f"Шаг {current_idx + 1} из {len(steps)}:\n\n{step_text}"
                await bot.send_message(chat_id, step_message)
                
                from universal_menu import get_menu_button
                markup = get_menu_button()
                await bot.send_message(chat_id, "Поделись результатом для этого шага:", reply_markup=markup)
        
        elif state.get('awaiting_final_answer'):
            # Continue with final questions
            await show_final_question(bot, chat_id, user_id)
        
        elif state.get('awaiting_exercise_text'):
            # Continue with exercise text input
            from universal_menu import get_menu_button
            markup = get_menu_button()
            await bot.send_message(chat_id, "Поделись результатом упражнения:", reply_markup=markup)
        
        else:
            # Default: show exercise selection
            header_text = "Выбери упражнение:"
            await bot.send_message(chat_id, header_text)

            for idx, exercise in enumerate(state.get('exercises', [])):
                goal = extract_exercise_goal(exercise)
                emoji = EXERCISE_EMOJIS[idx % len(EXERCISE_EMOJIS)]

                if goal:
                    card_text = f"{emoji} {exercise}\n{goal}"
                else:
                    card_text = f"{emoji} {exercise}"

                markup = types.InlineKeyboardMarkup()
                btn_select = types.InlineKeyboardButton(
                    "Выбрать",
                    callback_data=f"ex_select:{idx}"
                )
                markup.add(btn_select)

                await bot.send_message(chat_id, card_text, reply_markup=markup)

    except Exception as e:
        print(f"Error handling exercise continue after safety: {e}")
        await bot.answer_callback_query(callback_query.id)
